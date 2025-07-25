from flask import Blueprint, request, jsonify, url_for
import anthropic
from flask import Flask, render_template_string, request, jsonify, send_file, Response
import os
import re
import json
import csv
import io
import subprocess
import string
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path
import threading
from typing import Dict, List, Tuple, Optional
import time
from urllib.parse import quote
import html
from collections import OrderedDict
import uuid
import asyncio
import queue
from routes.grep_analyzer import AndroidLogAnalyzer, LimitedCache
import shutil
import pandas as pd
import requests
import atexit
import tempfile
from routes.analysisLockManager import AnalysisLockManager

# 創建全域的鎖管理器實例
analysis_lock_manager = AnalysisLockManager()

# 創建一個藍圖實例
main_page_bp = Blueprint('main_page_bp', __name__)

# Global storage for analysis results
analysis_cache = LimitedCache(max_size=100, max_age_hours=24)
analyzer = AndroidLogAnalyzer()
analysis_lock = threading.Lock()

# 全域變數追蹤臨時目錄
temp_directories = set()

# HTML template with beautiful charts
HTML_TEMPLATE = r'''
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Android ANR/Tombstone Analyzer</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
    /* ===== 全域重置和基礎樣式 ===== */
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }

    body {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;
        background-color: #f0f2f5;
        color: #333;
        min-height: 100vh;
        display: flex;
        flex-direction: column;
    }

    .container {
        max-width: 1400px;
        margin: 0 auto;
        padding: 20px;
        padding-bottom: 200px;
        flex: 1;
    }

    /* ===== 頁首樣式 ===== */
    .header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 30px 30px 50px 30px;  /* 增加底部 padding */
        border-radius: 12px;
        margin-bottom: 30px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        position: relative;
        min-height: 170px;  /* 設定最小高度 */
    }

    .header h1 {
        font-size: 2.5rem;
        margin-bottom: 10px;
    }

    .load-excel-btn {
        background: #17a2b8;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        display: inline-flex;
        align-items: center;
        gap: 8px;
    }

    .load-excel-btn:hover {
        background: #138496;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4);
    }

    .export-excel-report-btn {
        position: absolute;
        top: 90px;  /* 改到第三排 */
        right: 30px;
        background: #ff6b6b;
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .export-excel-report-btn:hover {
        background: rgba(255, 107, 107, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    @media (max-width: 768px) {
        .export-excel-report-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
    }

    .export-html-btn {
        position: absolute;
        top: 90px;  /* 改到第三排 */
        right: 460px;  /* 在匯出 Excel 左邊 */
        background: rgba(255, 255, 255, 0.2);
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .export-html-btn:hover {
        background: rgba(255, 255, 255, 0.3);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    @media (max-width: 768px) {
        .export-html-btn {
            position: static;
            margin-top: 20px;
            display: block;
            width: 100%;
        }
    }

    .export-excel-btn {
        position: absolute;
        top: 90px;  /* 改到第三排 */
        right: 330px;  /* 在合併 Excel 左邊 */
        background: #28a745;
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    /* 查看分析結果區塊 */
    .analysis-result-section {
        margin-bottom: 20px;
        padding: 15px;
        background: #f0f8ff;
        border-radius: 8px;
        border: 2px solid #4a90e2;
    }

    .view-analysis-btn {
        background: #6f42c1;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: transform 0.2s, box-shadow 0.2s;
    }

    .view-analysis-btn:hover {
        background: #5a32a3;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(111, 66, 193, 0.4);
    }

    /* 匯出歷史區塊 */
    .export-history-section {
        margin-top: 20px;
        padding: 15px;
        background: #e8f4f8;
        border-radius: 8px;
        border: 2px solid #17a2b8;
    }

    .export-all-btn {
        background: #17a2b8;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        display: flex;
        align-items: center;
        gap: 8px;
        margin: 0 auto;
    }

    .export-all-btn:hover {
        background: #138496;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4);
    }

    @media (max-width: 768px) {
        .export-html-btn,
        .export-excel-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
    }

    /* 分隔線樣式 */
    .header-separator {
        position: absolute;
        top: 75px;
        right: 30px;
        left: auto;
        width: 260px;
        height: 1px;
        background: rgba(255, 255, 255, 0.3);
        display: none;
    }

    .export-excel-btn:hover {
        background: rgba(40, 167, 69, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    .merge-excel-btn {
        position: absolute;
        top: 90px;  /* 改到第三排 */
        right: 200px;  /* 在匯出 Excel 報表左邊 */
        background: #17a2b8;
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .merge-excel-btn:hover {
        background: rgba(23, 162, 184, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    @media (max-width: 768px) {
        .export-html-btn,
        .export-excel-btn,
        .merge-excel-btn,
        .view-existing-analysis-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
        
        .header-separator {
            position: static;
            width: 100%;
            margin: 10px 0;
        }
    }

    .export-all-excel-btn {
        background: #17a2b8;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s ease;
        display: flex;
        align-items: center;
        gap: 8px;
        margin: 0;  /* 移除 auto margin */
    }

    /* 查看已有分析結果的按鈕 */
    .view-existing-analysis-btn {
        position: absolute;
        top: 30px;
        right: 300px;  /* 在其他按鈕左邊 */
        background: #6f42c1;
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .view-existing-analysis-btn:hover {
        background: rgba(111, 66, 193, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    @media (max-width: 768px) {
        .export-html-btn,
        .export-excel-btn,
        .export-all-excel-btn,
        .view-existing-analysis-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
        
        .header-separator {
            position: static;
            width: 100%;
            margin: 10px 0;
        }
    }
    
    .export-all-excel-btn:hover {
        background: #138496;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4);
    }

    /* 檔案路徑資訊 */
    .file-path-info {
        margin-top: 10px;
        padding: 10px;
        background: #f8f9fa;
        border-radius: 6px;
        font-size: 14px;
    }

    .file-path-info code {
        background: #e9ecef;
        padding: 2px 6px;
        border-radius: 3px;
        font-size: 13px;
    }

    .export-all-excel-with-update-btn {
        background: #ffc107;
        color: #212529;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s ease;
        display: flex;
        align-items: center;
        gap: 8px;
        margin: 0;  /* 移除 auto margin */
    }

    .export-all-excel-with-update-btn:hover {
        background: #e0a800;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(255, 193, 7, 0.4);
    }

    /* 強調動畫 */
    @keyframes highlightPulse {
        0% {
            transform: scale(1);
            box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4);
        }
        50% {
            transform: scale(1.05);
            box-shadow: 0 6px 20px rgba(23, 162, 184, 0.6);
        }
        100% {
            transform: scale(1);
            box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4);
        }
    }

    .export-all-excel-btn.highlight {
        animation: highlightPulse 0.6s ease-in-out 2;
    }

    /* 更新資訊樣式 */
    .update-info {
        color: #28a745;
        font-weight: 600;
        margin-left: 10px;
        opacity: 0;
        transition: opacity 0.3s ease;
    }

    .update-info.show {
        opacity: 1;
    }

    @media (max-width: 768px) {
        .export-all-excel-btn,
        .export-all-excel-with-update-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
    }

    .export-csv-btn {
        position: absolute;
        top: 70px;  /* 在 HTML 按鈕下方 */
        right: 30px;
        background: #28a745;  /* 綠色背景 */
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .export-csv-btn:hover {
        background: rgba(40, 167, 69, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    .export-csv-btn:disabled {
        background: #ccc;
        cursor: not-allowed;
        transform: none;
    }

    @media (max-width: 768px) {
        .export-csv-btn {
            position: static;
            margin-top: 10px;
            display: block;
            width: 100%;
        }
    }

    /* ===== 導航欄樣式 ===== */
    .nav-bar {
        position: fixed;
        right: 90px;
        bottom: 90px;
        background: white;
        border: 1px solid #e1e4e8;
        border-radius: 12px;
        padding: 15px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        z-index: 100;
        max-width: 200px;
        transform: translateX(300px);
        opacity: 0;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        visibility: hidden;
    }

    .nav-bar.show {
        transform: translateX(0);
        opacity: 1;
        visibility: visible;
    }

    .nav-title {
        font-weight: 600;
        color: #333;
        margin-bottom: 10px;
        font-size: 14px;
    }

    .nav-links {
        display: flex;
        flex-direction: column;
        gap: 8px;
    }

    .nav-link {
        background: #f0f0f0;
        color: #667eea;
        padding: 10px 14px;
        border-radius: 8px;
        text-decoration: none;
        font-size: 13px;
        transition: all 0.2s;
        white-space: nowrap;
        text-align: center;
    }

    .nav-link:hover {
        background: #667eea;
        color: white;
        transform: translateY(-2px);
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }

    /* ===== 浮動按鈕樣式 ===== */
    .back-to-top,
    .nav-toggle-btn,
    .global-toggle-btn {
        position: fixed !important;
        width: 50px;
        height: 50px;
        border-radius: 50%;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        font-size: 24px;
        cursor: pointer;
        transition: all 0.3s;
        z-index: 9999 !important;
        right: 30px !important;
        left: auto !important;
        background: #667eea;
        color: white;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
        text-align: center;
        line-height: 1;
        opacity: 0;
        visibility: hidden;
    }

    .back-to-top {
        bottom: 30px !important;
    }

    .nav-toggle-btn {
        bottom: 90px !important;
    }

    .global-toggle-btn {
        bottom: 150px !important;
    }

    .back-to-top.show,
    .nav-toggle-btn.show,
    .global-toggle-btn.show {
        opacity: 1;
        visibility: visible;
    }

    .back-to-top:hover,
    .nav-toggle-btn:hover,
    .global-toggle-btn:hover {
        background: #5a67d8;
        transform: translateY(-5px) !important;
        box-shadow: 0 6px 16px rgba(102, 126, 234, 0.6);
    }

    .nav-toggle-btn.active {
        background: #5a67d8;
        transform: rotate(180deg);
    }

    /* ===== 動畫效果 ===== */
    @keyframes slideIn {
        from {
            transform: translateX(300px);
            opacity: 0;
        }
        to {
            transform: translateX(0);
            opacity: 1;
        }
    }

    @keyframes slideOut {
        from {
            transform: translateX(0);
            opacity: 1;
        }
        to {
            transform: translateX(300px);
            opacity: 0;
        }
    }

    @keyframes dots {
        0%, 20% { content: ''; }
        40% { content: '.'; }
        60% { content: '..'; }
        80%, 100% { content: '...'; }
    }

    .nav-bar.animating-in {
        animation: slideIn 0.3s ease-out forwards;
    }

    .nav-bar.animating-out {
        animation: slideOut 0.3s ease-in forwards;
    }

    /* ===== 控制面板樣式 ===== */
    .control-panel {
        background-color: white;
        padding: 25px;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        margin-bottom: 30px;
    }

    .input-group {
        margin-bottom: 20px;
        position: relative;
    }

    label {
        display: block;
        margin-bottom: 8px;
        font-weight: 600;
        color: #555;
    }

    input[type="text"] {
        width: 100%;
        padding: 12px;
        border: 2px solid #e1e4e8;
        border-radius: 8px;
        font-size: 16px;
        transition: border-color 0.3s;
    }

    input[type="text"]:focus {
        outline: none;
        border-color: #667eea;
    }

    #pathInput {
        background-color: #f8f9fa;
        padding: 15px;
        border: 2px solid #e1e4e8;
        border-left: 5px solid #28a745;
        border-radius: 8px;
        margin-top: 5px;
        font-family: monospace;
        font-size: 90%;
        white-space: nowrap;      /* 新增：保持單行顯示 */
        overflow-x: auto;         /* 新增：允許水平捲動 */
        overflow-y: hidden;       /* 新增：隱藏垂直捲軸 */
    }

    /* 新增：美化輸入框的捲軸 */
    #pathInput::-webkit-scrollbar {
        height: 6px;
    }

    #pathInput::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 3px;
    }

    #pathInput::-webkit-scrollbar-thumb {
        background: #999;
        border-radius: 3px;
    }

    #pathInput::-webkit-scrollbar-thumb:hover {
        background: #666;
    }

    #pathInput:focus ~ .path-autocomplete {
        border-color: #667eea;
    }

    #pathInput.autocomplete-open {
        border-bottom-left-radius: 0;
        border-bottom-right-radius: 0;
    }

    /* ===== 路徑自動完成 ===== */
    .path-autocomplete {
        position: absolute;
        top: 100%;
        left: 0;
        right: 0;
        background: white;
        border: 2px solid #e1e4e8;
        border-top: none;
        border-radius: 0 0 8px 8px;
        max-height: 300px;
        overflow-y: auto;
        overflow-x: hidden;         /* 容器本身不需要水平捲軸 */
        display: none;
        z-index: 1000;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }

    /* 為自動完成框添加捲軸樣式 */
    .path-autocomplete::-webkit-scrollbar {
        width: 8px;
    }

    .path-autocomplete::-webkit-scrollbar-track {
        background: #f1f1f1;
    }

    .path-autocomplete::-webkit-scrollbar-thumb {
        background: #888;
        border-radius: 4px;
    }

    .path-autocomplete::-webkit-scrollbar-thumb:hover {
        background: #555;
    }

    .path-suggestion {
        padding: 12px;
        cursor: pointer;
        border-bottom: 1px solid #f0f0f0;
        font-family: monospace;
        font-size: 14px;
        color: #333;
        transition: background-color 0.2s;
        white-space: nowrap;        /* 防止折行 */
        overflow-x: auto;           /* 允許水平捲動 */
        overflow-y: hidden;         /* 隱藏垂直捲軸 */
        text-overflow: ellipsis;    /* 文字過長時顯示省略號 */
        max-width: 100%;            /* 確保不超出容器 */     
    }

    .path-suggestion:hover {
        background-color: #f8f9fa;
        overflow-x: auto;           /* 保持水平捲動 */
        position: relative;
        z-index: 10;          
    }

    /* 美化路徑建議的捲軸 */
    .path-suggestion::-webkit-scrollbar {
        height: 4px;
    }

    .path-suggestion::-webkit-scrollbar-track {
        background: transparent;
    }

    .path-suggestion::-webkit-scrollbar-thumb {
        background: #ccc;
        border-radius: 2px;
    }

    .path-suggestion::-webkit-scrollbar-thumb:hover {
        background: #999;
    }

    .path-suggestion.selected {
        background-color: #e8eaf6;
    }

    .path-suggestion .star {
        color: #ffd700;
        margin-left: 5px;
    }

    .path-loading {
        padding: 12px;
        text-align: center;
        color: #667eea;
        font-size: 14px;
    }

    .path-format {
        background-color: #f8f9fa;
        padding: 12px 15px;
        border: 2px solid #e1e4e8;
        border-left: 5px solid #3498db;
        border-radius: 8px;
        margin-top: 5px;
    }

    .path-format ul {
        line-height: 1.6;
    }

    .path-format ul li {
        color: #555;
        font-size: 14px;
    }

    .path-format strong {
        color: #3498db;
    }

    /* ===== 按鈕樣式 ===== */
    .button-group {
        display: flex;
        gap: 10px;
        margin-top: 20px;
        margin-bottom: 5px;
        flex-wrap: wrap;
        align-items: center;  /* 垂直居中對齊 */
    }

    button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: transform 0.2s, box-shadow 0.2s;
    }

    button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
    }

    button:disabled {
        background: #ccc;
        cursor: not-allowed;
        transform: none;
    }

    /* ===== 載入和狀態訊息 ===== */
    .loading {
        display: none;
        text-align: center;
        padding: 20px;
        color: #667eea;
    }

    .loading::after {
        content: '...';
        animation: dots 1.5s steps(4, end) infinite;
    }

    .error {
        color: #e53e3e;
        padding: 15px;
        background-color: #fff5f5;
        border: 1px solid #feb2b2;
        border-radius: 8px;
        margin-bottom: 20px;
    }

    .success {
        color: #38a169;
        padding: 10px 15px;
        background-color: #f0fff4;
        border: 1px solid #9ae6b4;
        border-radius: 8px;
        margin-bottom: 15px;
        line-height: 1.4;
    }

    .info {
        color: #3182ce;
        padding: 15px;
        background-color: #ebf8ff;
        border: 1px solid #90cdf4;
        border-radius: 8px;
        margin-bottom: 20px;
    }

    /* ===== 統計卡片 ===== */
    .stats-summary {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 20px;
        margin-bottom: 30px;
    }

    .stat-card {
        background: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        text-align: center;
    }

    .stat-card h3 {
        color: #667eea;
        font-size: 2.5rem;
        margin-bottom: 5px;
    }

    .stat-card p {
        color: #666;
        font-size: 0.9rem;
    }

    .stat-card.highlight {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
    }

    .stat-card.highlight h3,
    .stat-card.highlight p {
        color: white;
    }

    .stat-card.highlight p {
        color: rgba(255, 255, 255, 0.9);
    }

    /* ===== 圖表樣式 ===== */
    .chart-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
        gap: 30px;
        margin-bottom: 30px;
    }

    .chart-container {
        background: white;
        padding: 25px;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }

    .chart-container h3 {
        margin-bottom: 20px;
        color: #333;
        font-size: 1.3rem;
    }

    .chart-wrapper {
        position: relative;
        height: 300px;
    }

    /* ===== 區塊樣式 ===== */
    .section-container,
    .logs-table {
        position: relative;
        margin-bottom: 30px;
        z-index: 1;
    }

    .section-container > div:first-child,
    .logs-table {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }

    .section-content {
        transition: all 0.3s ease-in-out;
        overflow: hidden;
    }

    .section-container.collapsed .section-content {
        max-height: 0;
        opacity: 0;
        padding: 0;
        margin: 0;
    }

    .section-container.collapsed .table-header {
        border-radius: 12px;
        margin-bottom: 0;
    }

    /* ===== 區塊標題 ===== */
    .section-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 15px;
        padding-right: 60px;
    }

    .section-title {
        font-size: 1.2rem;
        font-weight: 600;
        color: #333;
    }

    .top-link {
        color: #667eea;
        text-decoration: none;
        font-size: 14px;
        padding: 5px 10px;
        border-radius: 5px;
        transition: all 0.2s;
    }

    .top-link:hover {
        background: #f0f0f0;
        color: #5a67d8;
    }

    /* ===== 表格標題 ===== */
    .table-header {
        position: relative;
        padding: 20px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        border-bottom: 2px solid #e1e4e8;
    }

    .table-header .section-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 0 !important;
        margin: 0;
        position: relative;
    }

    .table-header h3 {
        color: white;
        margin: 0;
        font-size: 1.2rem;
        padding-right: 150px;
    }

    .table-header .top-link {
        position: absolute;
        right: 60px;
        top: 50%;
        transform: translateY(-50%);
        color: white;
        text-decoration: none;
        font-size: 14px;
        padding: 5px 10px;
        border-radius: 5px;
        transition: all 0.2s;
        white-space: nowrap;
    }

    .table-header .top-link:hover {
        background: rgba(255, 255, 255, 0.2);
    }

    /* ===== 區塊開合按鈕 ===== */
    .section-toggle {
        position: absolute;
        right: 20px;
        top: 50%;
        transform: translateY(-50%);
        width: 36px;
        height: 36px;
        background: rgba(255, 255, 255, 0.2);
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        cursor: pointer;
        transition: all 0.3s;
        z-index: 10;
        border: 2px solid rgba(255, 255, 255, 0.3);
    }

    .section-toggle:hover {
        background: rgba(255, 255, 255, 0.3);
        transform: translateY(-50%) scale(1.1);
    }

    .section-toggle .toggle-icon {
        font-size: 20px;
        color: white;
        line-height: 1;
        font-weight: normal;
    }

    .section-toggle .toggle-icon::before {
        content: "x";
        display: block;
    }

    .section-container.collapsed .section-toggle .toggle-icon::before {
        content: "+";
    }

    .logs-table .section-toggle {
        top: 15px;
        right: 15px;
        transform: none;
    }

    /* ===== Tooltip 樣式 ===== */
    .section-toggle[data-tooltip]::before {
        content: attr(data-tooltip);
        position: absolute;
        left: calc(100% + 15px);
        top: 50%;
        transform: translateY(-50%);
        background: rgba(0, 0, 0, 0.9);
        color: white;
        padding: 6px 12px;
        border-radius: 4px;
        font-size: 12px;
        white-space: nowrap;
        opacity: 0;
        visibility: hidden;
        transition: all 0.2s ease;
        pointer-events: none;
        z-index: 10000;
    }

    .section-toggle[data-tooltip]::after {
        content: '';
        position: absolute;
        left: calc(100% + 7px);
        top: 50%;
        transform: translateY(-50%);
        width: 0;
        height: 0;
        border-style: solid;
        border-width: 5px 8px 5px 0;
        border-color: transparent rgba(0, 0, 0, 0.9) transparent transparent;
        opacity: 0;
        visibility: hidden;
        transition: all 0.2s ease;
        pointer-events: none;
    }

    .section-toggle[data-tooltip]:hover::before,
    .section-toggle[data-tooltip]:hover::after {
        opacity: 1;
        visibility: visible;
    }

    .floating-tooltip {
        position: absolute;
        right: calc(100% + 15px);
        top: 50%;
        transform: translateY(-50%);
        background: rgba(0, 0, 0, 0.9);
        color: white;
        padding: 8px 12px;
        border-radius: 6px;
        font-size: 13px;
        font-weight: 500;
        white-space: nowrap;
        opacity: 0;
        visibility: hidden;
        transition: all 0.3s ease;
        pointer-events: none;
        z-index: 10000;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.3);
    }

    .floating-tooltip::after {
        content: '';
        position: absolute;
        left: 100%;
        top: 50%;
        transform: translateY(-50%);
        width: 0;
        height: 0;
        border-style: solid;
        border-width: 6px 0 6px 8px;
        border-color: transparent transparent transparent rgba(0, 0, 0, 0.9);
    }

    /* ===== 表格樣式 ===== */
    .summary-table,
    .logs-table {
        background-color: white;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        overflow: hidden;
        margin-bottom: 30px;
    }

    .table-controls {
        padding: 15px 20px;
        background-color: #f8f9fa;
        border-bottom: 1px solid #e1e4e8;
        display: flex;
        justify-content: space-between;
        align-items: center;
        flex-wrap: wrap;
        gap: 10px;
    }

    .search-box {
        position: relative;
        flex: 1;
    }

    .search-box input {
        width: 100%;
        padding: 8px 35px 8px 12px;
        border: 1px solid #ddd;
        border-radius: 6px;
        font-size: 14px;
    }

    .search-box::after {
        content: '🔍';
        position: absolute;
        right: 10px;
        top: 50%;
        transform: translateY(-50%);
        opacity: 0.5;
    }

    .regex-toggle {
        display: flex;
        align-items: center;
        gap: 5px;
        color: #666;
        font-size: 12px;
        cursor: pointer;
        user-select: none;
        margin-left: 10px;
        margin-top: 10px;
    }

    .regex-toggle input {
        cursor: pointer;
    }

    .regex-toggle:hover {
        color: #667eea;
    }

    /* ===== 分頁樣式 ===== */
    .pagination {
        display: flex;
        gap: 5px;
        align-items: center;
        flex-wrap: wrap;
    }

    .pagination button {
        padding: 5px 10px;
        font-size: 14px;
        background: white;
        color: #667eea;
        border: 1px solid #667eea;
        min-width: 35px;
    }

    .pagination button:hover:not(:disabled) {
        background: #667eea;
        color: white;
    }

    .pagination button:disabled {
        background: #f0f0f0;
        color: #999;
        border-color: #ddd;
        cursor: not-allowed;
    }

    .pagination button.active {
        background: #667eea;
        color: white;
    }

    .pagination .page-numbers {
        display: flex;
        gap: 5px;
        align-items: center;
    }

    .pagination span {
        padding: 0 10px;
        font-size: 14px;
        color: #666;
    }

    /* ===== 表格內容樣式 ===== */
    .table-wrapper,
    .logs-table-content {
        overflow-x: auto;
    }

    table {
        width: 100%;
        border-collapse: collapse;
    }

    thead {
        position: sticky;
        top: 0;
        z-index: 10;
    }

    th {
        background-color: #f8f9fa;
        color: #333;
        font-weight: 600;
        padding: 12px;
        text-align: left;
        border-bottom: 2px solid #e1e4e8;
        position: sticky;
        top: 0;
        white-space: nowrap;
    }

    td {
        padding: 12px;
        border-bottom: 1px solid #e1e4e8;
    }

    tr:nth-child(even) {
        background-color: #f8f9fa;
    }

    tr:hover {
        background-color: #f3f4fb;  /* 更淺的藍色 */
    }

    /* ===== 表格內容特定樣式 ===== */
    .rank-number {
        font-weight: bold;
        color: #667eea;
        text-align: center;
    }

    .process-name {
        font-weight: 600;
        color: #667eea;
    }

    .process-name div {
        padding: 2px 0;
    }

    .process-name div:not(:last-child) {
        border-bottom: 1px solid #f0f0f0;
    }

    .line-number {
        font-weight: bold;
        color: #667eea;
        text-align: center;
        padding: 0 15px;
        cursor: pointer;
        position: relative;
        transition: all 0.2s;
    }

    .line-number:hover::after {
        content: "📌";
        position: absolute;
        right: 2px;
        font-size: 10px;
        opacity: 0.5;
    }

    .line-number.bookmarked:hover::after {
        content: "❌";
        opacity: 0.7;
    }

    .file-link {
        color: #667eea;
        text-decoration: none;
        cursor: pointer;
        transition: all 0.2s;
    }

    .file-link:hover {
        text-decoration: underline;
        color: #5a67d8;
    }

    .cmdline-cell {
        font-size: 0.9em;
        max-width: 600px;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }

    .folder-path-cell {
        font-size: 0.9em;
        color: #999;
        font-family: monospace;
    }

    .filesize {
        color: #666;
        font-size: 0.9em;
    }

    /* ===== 排序功能 ===== */
    .sortable {
        cursor: pointer;
        user-select: none;
        position: relative;
    }

    .sortable:hover {
        background-color: #e8eaf6;
    }

    .sort-indicator {
        position: absolute;
        right: 5px;
        color: #667eea;
        font-size: 12px;
    }

    /* ===== 標籤樣式 ===== */
    .grep-badge {
        display: inline-block;
        background: #48bb78;
        color: white;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.75rem;
        margin-left: 8px;
        vertical-align: middle;
    }

    .no-grep-badge {
        background: #f56565;
    }

    /* ===== 其他元素樣式 ===== */
    #results {
        display: none;
    }

    .filter-input {
        width: 300px;
        padding: 8px 12px;
        border: 2px solid rgba(255, 255, 255, 0.8);
        border-radius: 8px;
        font-size: 14px;
        background-color: rgba(255, 255, 255, 0.95);
    }

    ul {
        list-style: none;
        padding: 0;
        margin-left: 40px;
    }

    .icon {
        padding: 2px;
        margin: 2px;
    }

    #globalToggleIcon,
    .nav-icon {
        font-size: 20px;
        line-height: 1;
    }

    /* ===== 頁尾樣式 ===== */
    .footer {
        background-color: #2d3748;
        color: #a0aec0;
        text-align: center;
        padding: 10px;
        margin-top: 10px;
        font-size: 14px;
    }

    .footer p {
        margin: 0;
    }

    #results {
        display: none;
    }
    
    .grep-badge {
        display: inline-block;
        background: #48bb78;
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.8rem;
        margin-left: 10px;
    }
    
    .no-grep-badge {
        background: #f56565;
    }
    
    .footer {
        background-color: #2d3748;
        color: #a0aec0;
        text-align: center;
        padding: 10px;
        margin-top: 10px;
        font-size: 14px;
    }
    
    .footer p {
        margin: 0;
    }
    
    .filesize {
        color: #666;
        font-size: 0.9em;
    }
    
    .line-number {
        font-weight: bold;
        color: #667eea;
        text-align: center;
        transition: all 0.2s;
    }

    ul {
        list-style: none;
        padding: 0;
        /* 讓整個列表向右縮排 */
        margin-left: 40px;
    }
    
    .icon {
        padding:2px;
        margin:2px;
    }
    
    .path-format {
        background-color: #f8f9fa;
        padding: 15px;
        border: 2px solid #e1e4e8;
        border-left: 5px solid #3498db;
        border-radius: 8px;
        margin-top: 5px;
        display: flex;
        align-items: flex-start;
    }
    
    .path-format strong {
        color: #3498db;
    }

    .line-number {
        padding: 0 15px;
        cursor: pointer;
        position: relative;
        transition: all 0.2s;
    }

    .line-number:hover::after {
        content: "📌";
        position: absolute;
        right: 2px;
        font-size: 10px;
        opacity: 0.5;
    }

    .line-number.bookmarked:hover::after {
        content: "❌";
        opacity: 0.7;
    }

    .process-name div {
        padding: 2px 0;
    }

    .process-name div:not(:last-child) {
        border-bottom: 1px solid #f0f0f0;
    }

    .table-highlight {
        background-color: rgba(102, 126, 234, 0.1) !important;  /* 透明的藍紫色 */
        color: #333;
        font-weight: 600;
    }

    .logs-table-content table thead tr th {
        white-space: nowrap
    }

    /* Analysis Result Button */
    .analysis-result-btn {
        position: fixed;
        bottom: 30px;
        right: 90px;  /* 在返回頂部按鈕左邊 */
        background: #28a745;  /* 綠色背景 */
        color: white;
        width: 50px;
        height: 50px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 24px;
        cursor: pointer;
        transition: all 0.3s;
        opacity: 0;
        visibility: hidden;
        z-index: 1000;
        box-shadow: 0 4px 12px rgba(40, 167, 69, 0.4);
        text-decoration: none;
    }

    .analysis-result-btn.show {
        opacity: 1;
        visibility: visible;
    }

    .analysis-result-btn:hover {
        background: #218838;
        transform: translateY(-5px);
        box-shadow: 0 6px 16px rgba(40, 167, 69, 0.6);
        color: white;
        text-decoration: none;
    }

    /* 當有多個按鈕時的排列 */
    @media (max-width: 768px) {
        .analysis-result-btn {
            right: 30px;
            bottom: 90px;  /* 在返回頂部按鈕上方 */
        }
        
        .nav-toggle-btn {
            bottom: 150px;  /* 再往上移 */
        }
    }

    /* Tooltip styles */
    .folder-path-cell {
        font-size: 0.9em;
        color: #999;
        font-family: monospace;
        cursor: help;
        position: relative;
        display: inline-block;  /* 加入這行 */
        line-height: 1.5;       /* 加入這行 */
        vertical-align: middle; /* 加入這行 */
    }

    /* Tooltip 容器 */
    .tooltip-container {
        position: relative;
        display: table-cell;    /* 改為 table-cell */
        cursor: help;
        vertical-align: middle; /* 加入這行 */
    }

    .tooltip-container:hover .tooltip-text {
        visibility: visible;
        opacity: 1;
        animation: fadeIn 0.3s ease-in-out;
    }

    .tooltip-text {
        visibility: hidden;
        opacity: 0;
        position: absolute;
        bottom: 125%;  /* 預設在上方 */
        left: 50%;
        transform: translateX(-50%);
        background-color: rgba(0, 0, 0, 0.95);
        color: white;
        padding: 10px 15px;
        border-radius: 8px;
        font-size: 13px;
        font-family: monospace;
        white-space: pre-wrap;
        word-break: break-all;
        z-index: 1000;
        transition: opacity 0.3s, visibility 0.3s;
        min-width: 300px;
        max-width: 800px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.3);
        border: 1px solid rgba(255, 255, 255, 0.1);
    }

    .tooltip-text::after {
        content: "";
        position: absolute;
        top: 100%;
        left: 50%;
        margin-left: -8px;
        border-width: 8px;
        border-style: solid;
        border-color: rgba(0, 0, 0, 0.95) transparent transparent transparent;
    }

    @keyframes fadeIn {
        from {
            opacity: 0;
            transform: translateX(-50%) translateY(-5px);
        }
        to {
            opacity: 1;
            transform: translateX(-50%) translateY(0);
        }
    }

    /* 針對前幾行，將 tooltip 顯示在下方 */
    tr:nth-child(-n+3) .tooltip-text {
        bottom: auto;
        top: 125%;
    }

    /* 調整箭頭方向 - 預設向下 */
    .tooltip-text::after {
        content: "";
        position: absolute;
        top: 100%;
        left: 50%;
        margin-left: -8px;
        border-width: 8px;
        border-style: solid;
        border-color: rgba(0, 0, 0, 0.95) transparent transparent transparent;
    }

    /* 前幾行的箭頭向上 */
    tr:nth-child(-n+3) .tooltip-text::after {
        top: auto;
        bottom: 100%;
        border-color: transparent transparent rgba(0, 0, 0, 0.95) transparent;
    }

    /* 當 tooltip 在下方時的箭頭樣式 */
    .tooltip-text.tooltip-below::after {
        top: auto;
        bottom: 100%;
        border-color: transparent transparent rgba(0, 0, 0, 0.95) transparent;
    }

    /* 合併 Excel 彈出視窗樣式 */
    .merge-dialog-overlay {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        width: 100%;
        height: 100%;
        background: rgba(0, 0, 0, 0.5);
        display: flex;
        align-items: center;
        justify-content: center;
        z-index: 10000;
        overflow: hidden;  /* 防止滾動 */
    }

    .merge-dialog {
        background: white;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.15);
        width: 90%;
        max-width: 700px;
        max-height: 90vh;
        display: flex;
        flex-direction: column;
        overflow: hidden;  /* 防止內容溢出 */
    }

    .merge-dialog-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 20px 25px;
        border-radius: 12px 12px 0 0;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }

    .merge-dialog-header h3 {
        margin: 0;
        font-size: 1.5rem;
        font-weight: 600;
    }

    .merge-dialog-close {
        background: rgba(255, 255, 255, 0.2);
        border: none;
        font-size: 28px;
        cursor: pointer;
        color: white;
        padding: 0;
        width: 36px;
        height: 36px;
        display: flex;
        align-items: center;
        justify-content: center;
        border-radius: 50%;
        transition: all 0.2s;
    }

    .merge-dialog-close:hover {
        background: rgba(255, 255, 255, 0.3);
        transform: scale(1.1);
    }

    .merge-dialog-body {
        padding: 25px;
        flex: 1;
        overflow-y: auto;
        background-color: #f0f2f5;
    }

    /* 拖曳區域樣式 */
        .merge-drop-zone {
        border: 3px dashed #e1e4e8;
        border-radius: 12px;
        padding: 15px;  /* 從 25px 改為 15px */
        text-align: center;
        transition: all 0.3s;
        background: white;
        cursor: pointer;
    }

    .merge-drop-zone:hover {
        border-color: #667eea;
        background: #f8f9ff;
    }

    .merge-drop-zone.drag-over {
        border-color: #667eea;
        background: #e8eaf6;
        transform: scale(1.02);
    }

    .drop-zone-content {
        pointer-events: none;
    }

    .drop-icon {
        font-size: 28px;  /* 從 36px 改為 28px */
        margin-bottom: 5px;  /* 從 8px 改為 5px */
    }

    .drop-zone-hint {
        color: #999;
        font-size: 12px;  /* 從 13px 改為 12px */
        margin: 5px 0;  /* 從 8px 改為 5px */
    }

    .btn-select-file {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 10px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        pointer-events: auto;
    }

    .btn-select-file:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
    }

    /* 分隔線 */
    .merge-separator {
        text-align: center;
        margin: 20px 0;
        position: relative;
    }

    .merge-separator span {
        background: #f0f2f5;
        padding: 0 15px;
        color: #999;
        font-size: 14px;
    }

    .merge-separator::before {
        content: '';
        position: absolute;
        top: 50%;
        left: 0;
        right: 0;
        height: 1px;
        background: #e1e4e8;
        z-index: -1;
    }

    /* 路徑輸入樣式（複用主介面樣式） */
    .merge-input-group {
        position: relative;
        margin-bottom: 15px;
    }

    .merge-input-group label {
        display: block;
        margin-bottom: 8px;
        font-weight: 600;
        color: #555;
    }

    .merge-input-group input {
        width: 100%;
        padding: 12px;
        border: 2px solid #e1e4e8;
        border-radius: 8px;
        font-size: 16px;
        transition: border-color 0.3s;
        background-color: white;
    }

    .merge-input-group input:focus {
        outline: none;
        border-color: #667eea;
    }

    .merge-input-group input.autocomplete-open {
        border-bottom-left-radius: 0;
        border-bottom-right-radius: 0;
    }

    /* 檔案資訊顯示 */
    .merge-file-info {
        background: #d4edda;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #28a745;
        margin-top: 15px;
    }

    .file-info-content {
        display: flex;
        justify-content: space-between;
        align-items: center;
    }

    .merge-file-info code {
        background: rgba(0, 0, 0, 0.05);
        padding: 4px 8px;
        border-radius: 4px;
        font-size: 14px;
        word-break: break-all;
    }

    .btn-clear {
        background: #dc3545;
        color: white;
        border: none;
        padding: 5px 12px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    .btn-clear:hover {
        background: #c82333;
    }

    /* 對話框底部 */
    .merge-dialog-footer {
        padding: 20px 25px;
        border-top: 1px solid #e1e4e8;
        display: flex;
        justify-content: flex-end;
        gap: 10px;
        background: white;
        border-radius: 0 0 12px 12px;
    }

    /* 修改路徑建議項目樣式 - 允許橫向捲動 */
    .merge-path-autocomplete .path-suggestion {
        padding: 10px 12px;
        cursor: pointer;
        border-bottom: 1px solid #f0f0f0;
        font-family: monospace;
        font-size: 13px;
        color: #333;
        transition: background-color 0.2s;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        position: relative;  /* 重要：設定相對定位 */
    }

    /* 確保最後一個項目沒有底部邊框 */
    .merge-path-autocomplete .path-suggestion:last-child {
        border-bottom: none;
    }

    /* 為路徑建議添加內部容器 */
    .merge-path-autocomplete-inner {
        min-width: 100%;
        width: max-content;
    }

    /* 修改 merge-path-autocomplete 的樣式 */
    #mergePathAutocomplete {
        position: absolute;
        top: 100%;
        left: 0;
        right: 0;
        background: white;
        border: 2px solid #e1e4e8;
        border-top: none;
        border-radius: 0 0 8px 8px;
        max-height: 150px;
        overflow-y: auto;
        overflow-x: auto;
        display: none;
        z-index: 1000;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }

    /* 懸停時的樣式 - 不改變大小 */
    .merge-path-autocomplete .path-suggestion:hover {
        background-color: #f8f9fa;
    }

    /* 顯示完整路徑的 tooltip */
    .merge-path-autocomplete .path-suggestion::after {
        content: attr(title);
        position: absolute;
        left: 0;
        top: 100%;
        background: rgba(0, 0, 0, 0.9);
        color: white;
        padding: 5px 10px;
        border-radius: 4px;
        font-size: 12px;
        white-space: nowrap;
        opacity: 0;
        pointer-events: none;
        transition: opacity 0.2s;
        z-index: 1001;
        display: none;
    }

    .merge-path-autocomplete .path-suggestion:hover::after {
        opacity: 1;
        display: block;
    }

    /* 捲軸樣式 */
    #mergePathAutocomplete::-webkit-scrollbar {
        height: 6px;
        width: 6px;
    }

    #mergePathAutocomplete::-webkit-scrollbar-track {
        background: #f1f1f1;
    }

    #mergePathAutocomplete::-webkit-scrollbar-thumb {
        background: #999;
        border-radius: 3px;
    }

    #mergePathAutocomplete::-webkit-scrollbar-thumb:hover {
        background: #666;
    }

    /* 防止選中時的跳動 */
    .merge-path-autocomplete .path-suggestion.selected {
        background-color: #e8eaf6;
    }

    .btn-select-file:focus {
        outline: 2px solid #667eea;
        outline-offset: 2px;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.2);
    }

    /* 選中項目的樣式 */
    .merge-path-autocomplete .path-suggestion.selected {
        background-color: #e8eaf6;
        position: relative;  /* 確保選中項目顯示在上層 */
    }

    /* 讓捲軸更明顯 */
    #mergePathAutocomplete::-webkit-scrollbar {
        height: 8px;
        width: 8px;
    }

    #mergePathAutocomplete::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }

    #mergePathAutocomplete::-webkit-scrollbar-thumb {
        background: #888;
        border-radius: 4px;
    }

    #mergePathAutocomplete::-webkit-scrollbar-thumb:hover {
        background: #555;
    }

    /* 支援格式容器 */
    .merge-dialog-body h2 {
        font-size: 16px;
        color: #333;
        margin-bottom: 10px;
        font-weight: 600;
    }

    /* 只針對支援格式的框 */
    .support-format-box {
        background: #f8f9fa;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #667eea;
    }

    /* 確保拖曳區域內容沒有邊框 */
    .drop-zone-content {
        border: none;
        background: transparent;
        padding: 0;
    }

    /* 統一字體樣式 */
    .merge-dialog-body {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;
        font-size: 14px;
        color: #333;
    }

    .merge-dialog-body strong {
        font-weight: 600;
        color: #333;
    }

    /* 統一列表項目樣式 */
    .support-format-section ul li {
        font-size: 14px;
        color: #555;
        line-height: 1.6;
        padding: 5px 0;
    }

    /* 圖示統一大小 */
    .icon {
        display: inline-block;
        width: 24px;
        text-align: center;
        font-size: 16px;
    }

    /* 列表項目樣式 */
    .merge-dialog-body ul {
        list-style: none !important;
        padding: 0 !important;
        margin: 0 !important;
    }

    .merge-dialog-body ul li {
        list-style-type: none !important;
        padding: 5px 0;
        font-size: 14px;
        color: #555;
        line-height: 1.6;
    }

    .merge-dialog-body ul li strong {
        color: #333;
        font-weight: 600;
    }

    /* 圖示樣式 */
    .merge-dialog-body .icon {
        display: inline-block;
        width: 24px;
        margin-right: 5px;
        text-align: center;
        font-size: 16px;
        vertical-align: middle;
    }

    /* 確保合併對話框內的所有列表都沒有項目符號 */
    .merge-dialog-body ul li::before {
        content: none !important;
    }

    .merge-dialog-body ul li::marker {
        content: none !important;
    }

    /* 調整支援格式的樣式 */
    .support-format-section {
        margin-top: 20px;
        padding: 0;
    }

    .support-format-section h2 {
        font-size: 16px;
        color: #333;
        margin-bottom: 10px;
        font-weight: 600;
    }

    .support-format-list {
        background: transparent;
        padding: 0;
        border: none;
    }

    .support-format-list ul {
        list-style: none !important;
        padding: 0 !important;
        margin: 0 !important;
    }

    .support-format-list ul li {
        list-style-type: none !important;
        padding: 8px 0;
        font-size: 14px;
        color: #555;
        line-height: 1.6;
        display: flex;
        align-items: center;
    }

    .support-format-list .icon {
        display: inline-flex;
        min-width: 24px;
        margin-right: 8px;
        font-size: 16px;
        justify-content: center;
    }

    #mergePathInput {
        background-color: #f8f9fa;
        padding: 15px;
        border: 2px solid #e1e4e8;
        border-left: 5px solid #28a745;
        border-radius: 8px;
        margin-top: 5px;
        display: flex;
        align-items: flex-start;
        font-family: monospace;
        font-size: 90%;        
    }

    .download-zip-btn {
        position: absolute;
        top: 130px;  /* 改到第四排 */
        right: 30px;
        background: #fd7e14;
        color: white;
        border: 2px solid rgba(255, 255, 255, 0.5);
        padding: 10px 20px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.3s;
        display: none;
    }

    .download-zip-btn:hover {
        background: rgba(253, 126, 20, 0.8);
        border-color: rgba(255, 255, 255, 0.8);
        transform: translateY(-2px);
    }

    /* 確保歷史區塊的所有按鈕使用相同樣式 */
    #historySection button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: transform 0.2s, box-shadow 0.2s;
    }

    #historySection button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
    }

    #historySection {
        margin-bottom: 30px;  /* 增加底部間距 */
    }

    /* 或者在統計摘要容器上增加頂部間距 */
    #stats-section-container {
        margin-top: 20px;  /* 增加頂部間距 */
    }

    /* 為不同功能的按鈕設定不同顏色 */
    #viewIndexBtn {
        background: #6f42c1 !important;  /* 紫色 */
    }

    #downloadExcelBtn {
        background: #28a745 !important;  /* 綠色 */
    }

    #viewHTMLBtn {
        background: #17a2b8 !important;  /* 藍綠色 */
    }

    #viewExcelReportBtn {
        background: #ff6b6b !important;  /* 紅色 */
    }

    #downloadZipBtn {
        background: #fd7e14 !important;  /* 橙色 */
    }

    /* 確保主控制面板的合併 Excel 按鈕樣式正確 */
    #mergeExcelMainBtn {
        background: #17a2b8 !important;
        color: white !important;
        border: none !important;
        padding: 12px 24px !important;
        border-radius: 8px !important;
        cursor: pointer !important;
        font-size: 16px !important;
        font-weight: 600 !important;
        transition: all 0.2s !important;
        display: inline-flex !important;
        align-items: center !important;
        gap: 8px !important;
        position: static !important;
        margin: 0 !important;
    }

    #mergeExcelMainBtn:hover {
        background: #138496 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(23, 162, 184, 0.4) !important;
    }

    #viewExcelReportBtn {
        background: #ff6b6b !important;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        display: none;
        align-items: center;
        gap: 8px;
    }

    #viewExcelReportBtn:hover {
        background: #ff5252 !important;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(255, 107, 107, 0.4);
    }

    .load-excel-btn {
        background: #6f42c1;  /* 改為紫色，與合併 Excel 的藍綠色區分 */
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        display: inline-flex;
        align-items: center;
        gap: 8px;
    }

    .load-excel-btn:hover {
        background: #5a32a3;  /* 深紫色 */
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(111, 66, 193, 0.4);
    }

    .header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 30px 30px 50px 30px;
        border-radius: 12px;
        margin-bottom: 30px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        position: relative;
        min-height: 200px;  /* 增加高度以容納更多按鈕 */
    }

    /* 重新調整按鈕位置 */
    .export-html-btn {
        position: absolute;
        top: 30px;
        right: 30px;
    }

    .export-excel-btn {
        position: absolute;
        top: 30px;
        right: 170px;  /* 調整間距 */
    }

    .merge-excel-btn {
        position: absolute;
        top: 30px;
        right: 310px;  /* 調整間距 */
    }

    .export-excel-report-btn {
        position: absolute;
        top: 90px;  /* 移到第二排 */
        right: 30px;
    }

    .download-zip-btn {
        position: absolute;
        top: 90px;  /* 第二排 */
        right: 200px;
    }

    .select-files-btn {
        background: #6f42c1;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        transition: all 0.2s;
        display: inline-flex;
        align-items: center;
        gap: 8px;
    }

    .select-files-btn:hover {
        background: #5a32a3;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(111, 66, 193, 0.4);
    }

    .selected-items-list {
        max-height: 300px;
        overflow-y: auto;
        border: 1px solid #e1e4e8;
        border-radius: 8px;
        padding: 10px;
        background: #f8f9fa;
    }

    .selected-item {
        padding: 10px;
        background: white;
        margin: 5px 0;
        border-radius: 6px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        border: 1px solid #e1e4e8;
    }

    .selected-item:hover {
        background: #f0f0f0;
    }

    /* ===== 精緻版頁籤樣式 (紫色主題) ===== */
    .tabs-container {
        display: flex;
        margin-bottom: 30px;
        padding: 5px;
        background: #f0f2f5;
        border-radius: 16px;
        position: relative;
    }

    .tab-button {
        flex: 1;
        padding: 16px 24px;
        background: transparent;
        border: none;
        cursor: pointer;
        font-size: 16px;
        font-weight: 600;
        color: #666;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        z-index: 2;
        border-radius: 12px;
        margin: 0 2px;
    }

    .tab-button:hover:not(.active) {
        color: #667eea;
        background: rgba(102, 126, 234, 0.08);
    }

    .tab-button.active {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
    }

    /* 移除滑動指示器，因為我們直接用漸變背景 */
    .tabs-container::before {
        display: none;
    }

    /* 為活動頁籤添加光澤效果 */
    .tab-button.active::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 50%;
        background: linear-gradient(to bottom, rgba(255,255,255,0.2), transparent);
        border-radius: 12px 12px 0 0;
        pointer-events: none;
    }

    /* 頁籤內容 */
    .tab-content {
        display: none;
        padding: 30px;
        background: white;
        border-radius: 16px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }

    .tab-content.active {
        display: block;
        animation: slideIn 0.4s ease-out;
    }

    @keyframes slideIn {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }

    /* 響應式設計 */
    @media (max-width: 768px) {
        .tabs-container {
            flex-direction: column;
            padding: 8px;
        }
        
        .tab-button {
            margin: 4px 0;
            width: 100%;
        }
    }

    /* 鎖定提示對話框樣式 */
    .lock-dialog-overlay {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: rgba(0, 0, 0, 0.6);
        display: flex;
        align-items: center;
        justify-content: center;
        z-index: 10001;
        backdrop-filter: blur(5px);
        animation: fadeIn 0.3s ease-out;
    }

    @keyframes fadeIn {
        from {
            opacity: 0;
        }
        to {
            opacity: 1;
        }
    }

    .lock-dialog {
        background: white;
        border-radius: 16px;
        box-shadow: 0 20px 40px rgba(0, 0, 0, 0.2);
        width: 90%;
        max-width: 480px;
        overflow: hidden;
        animation: slideUp 0.3s ease-out;
    }

    @keyframes slideUp {
        from {
            transform: translateY(50px);
            opacity: 0;
        }
        to {
            transform: translateY(0);
            opacity: 1;
        }
    }

    .lock-dialog-header {
        background: linear-gradient(135deg, #fbbf24 0%, #f59e0b 100%);
        color: white;
        padding: 20px 25px;
        display: flex;
        align-items: center;
        justify-content: center;
    }

    .lock-dialog-header h3 {
        margin: 0;
        font-size: 1.5rem;
        font-weight: 600;
        display: flex;
        align-items: center;
        gap: 10px;
    }

    .lock-dialog-body {
        padding: 30px 25px;
        text-align: center;
    }

    .lock-warning-icon {
        margin-bottom: 20px;
    }

    .lock-warning-icon svg {
        filter: drop-shadow(0 4px 6px rgba(251, 191, 36, 0.3));
    }

    .lock-message {
        font-size: 18px;
        color: #333;
        margin-bottom: 20px;
        font-weight: 500;
    }

    .lock-time-info {
        background: #fef3c7;
        border: 2px solid #fbbf24;
        border-radius: 12px;
        padding: 20px;
        margin: 20px 0;
    }

    .time-display {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 10px;
    }

    .time-label {
        font-size: 16px;
        color: #92400e;
    }

    .time-value {
        font-size: 24px;
        font-weight: bold;
        color: #d97706;
        font-variant-numeric: tabular-nums;
    }

    .lock-question {
        font-size: 16px;
        color: #666;
        margin-top: 20px;
    }

    .lock-dialog-footer {
        padding: 20px 25px;
        background: #f9fafb;
        display: flex;
        gap: 12px;
        justify-content: center;
    }

    .lock-dialog-footer button {
        padding: 12px 24px;
        border-radius: 10px;
        font-size: 16px;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.2s;
        display: flex;
        align-items: center;
        gap: 8px;
        border: none;
    }

    .btn-wait {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.3);
    }

    .btn-wait:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 16px rgba(16, 185, 129, 0.4);
    }

    .btn-cancel {
        background: #e5e7eb;
        color: #6b7280;
    }

    .btn-cancel:hover {
        background: #d1d5db;
    }

    .btn-icon {
        font-size: 18px;
    }

    /* 等待中動畫 */
    @keyframes pulse {
        0% {
            opacity: 1;
        }
        50% {
            opacity: 0.5;
        }
        100% {
            opacity: 1;
        }
    }

    .waiting-animation {
        animation: pulse 2s ease-in-out infinite;
    }

    /* 修改合併對話框的檔案資訊區域 */
    .merge-file-info {
        background: transparent;
        padding: 0;
        border: none;
        margin-top: 15px;
    }

    .merge-file-info .file-info-content {
        display: block;
    }

    .merge-file-info .file-info-content strong {
        display: block;
        margin-bottom: 10px;
        font-size: 16px;
        color: #333;
    }

    /* 確保選擇項目的樣式一致 */
    .merge-dialog .selected-item {
        padding: 10px;
        background: white;
        margin: 5px 0;
        border-radius: 6px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        border: 1px solid #e1e4e8;
        transition: all 0.2s;
    }

    .merge-dialog .selected-item:hover {
        background: #f0f0f0;
        border-color: #d0d0d0;
    }

    .merge-dialog .selected-item label {
        display: flex;
        align-items: center;
        flex: 1;
        cursor: pointer;
        margin: 0;
    }

    .merge-dialog .selected-item input[type="checkbox"] {
        cursor: pointer;
    }

    /* 清除按鈕樣式一致性 */
    .merge-dialog .btn-clear {
        background: #dc3545;
        color: white;
        border: none;
        padding: 5px 12px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    .merge-dialog .btn-clear:hover {
        background: #c82333;
        transform: translateY(-1px);
    }

    /* 全選按鈕樣式 */
    .merge-dialog .btn-select-all {
        background: #17a2b8;
        color: white;
        border: none;
        padding: 5px 12px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    .merge-dialog .btn-select-all:hover {
        background: #138496;
        transform: translateY(-1px);
    }

    .merge-dialog .btn-remove-selected {
        background: #ffc107;
        color: #212529;
        border: none;
        padding: 5px 12px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    .merge-dialog .btn-remove-selected:hover {
        background: #e0a800;
        transform: translateY(-1px);
    }

    /* 主頁面選擇區域的按鈕樣式 */
    #mainSelectedItemsSection .btn-select-all {
        background: #17a2b8;
        color: white;
        border: none;
        padding: 8px 16px;
        border-radius: 6px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    #mainSelectedItemsSection .btn-select-all:hover {
        background: #138496;
        transform: translateY(-1px);
    }

    #mainSelectedItemsSection .btn-remove-selected {
        background: #ffc107;
        color: #212529;
        border: none;
        padding: 8px 16px;
        border-radius: 6px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    #mainSelectedItemsSection .btn-remove-selected:hover {
        background: #e0a800;
        transform: translateY(-1px);
    }

    #mainSelectedItemsSection .btn-clear {
        background: #dc3545;
        color: white;
        border: none;
        padding: 8px 16px;
        border-radius: 6px;
        cursor: pointer;
        font-size: 14px;
        transition: all 0.2s;
    }

    #mainSelectedItemsSection .btn-clear:hover {
        background: #c82333;
        transform: translateY(-1px);
    }

    /* 確保選擇項目有正確的樣式 */
    #mainSelectedItemsList .selected-item {
        padding: 10px;
        background: white;
        margin: 5px 0;
        border-radius: 6px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        border: 1px solid #e1e4e8;
        transition: all 0.2s;
    }

    #mainSelectedItemsList .selected-item:hover {
        background: #f0f0f0;
        border-color: #d0d0d0;
    }

    #mainSelectedItemsList .selected-item label {
        display: flex;
        align-items: center;
        flex: 1;
        cursor: pointer;
        margin: 0;
    }

    /* 對話框內的訊息樣式 */
    #mergeMessage {
        padding: 10px 15px;
        border-radius: 6px;
        margin-bottom: 15px;
        display: none;
        font-size: 14px;
    }

    #mergeMessage.error {
        color: #e53e3e;
        background-color: #fff5f5;
        border: 1px solid #feb2b2;
    }

    #mergeMessage.success {
        color: #38a169;
        background-color: #f0fff4;
        border: 1px solid #9ae6b4;
    }

    #mergeMessage.info {
        color: #3182ce;
        background-color: #ebf8ff;
        border: 1px solid #90cdf4;
    }

    </style>      
</head>
<body>
    <div class="container">
        <div class="header" id="top">
            <h1>Android ANR/Tombstone Analyzer</h1>
            <p>分析 anr/ 和 tombstones/ 資料夾中的 Cmd line: / Cmdline: 統計資訊</p>
            <button class="export-excel-btn" id="exportExcelBtn" onclick="exportAIResults()" style="display: none;">匯出 Excel</button>
            <button class="export-excel-report-btn" id="exportExcelReportBtn" onclick="exportExcelReport()" style="display: none;">匯出 Excel 報表</button>
            <button class="merge-excel-btn" id="mergeExcelBtn" onclick="openMergeDialog()" style="display: none;">合併 Excel</button>
            <button class="export-html-btn" id="exportHtmlBtn" onclick="exportResults('html')">匯出 HTML</button>
            <button class="download-zip-btn" id="downloadCurrentZipBtn" onclick="downloadCurrentAnalysisZip()" style="display: none;">
                📦 分析結果打包
            </button>            
            <div class="header-separator" id="headerSeparator"></div>
        </div>
        
        <!-- Navigation Bar -->
        <div class="nav-bar" id="navBar">
            <div class="nav-title">快速導覽</div>
            <div class="nav-links">
                <a href="#stats-section" class="nav-link">📊 統計摘要</a>
                <a href="#charts-section" class="nav-link">📈 圖表分析</a>
                <a href="#process-summary-section" class="nav-link">🔧 程序統計</a>
                <a href="#summary-section" class="nav-link">📋 彙整資訊</a>
                <a href="#files-section" class="nav-link">📁 檔案統計</a>
                <a href="#logs-section" class="nav-link">📝 詳細記錄</a>
            </div>
        </div>
        
        <!-- Back to Top Button -->
        <!-- Analysis Result Button -->
        <a class="analysis-result-btn" id="analysisResultBtn" href="" target="_blank" title="查看詳細分析報告">📊</a>        
        <div class="back-to-top" id="backToTop" onclick="scrollToTop()">↑</div>
        <div class="global-toggle-btn" id="globalToggleBtn" onclick="toggleAllSections()">
            <span id="globalToggleIcon">⊕</span>
        </div>        
        <div class="nav-toggle-btn" id="navToggleBtn" onclick="toggleNavBar()">
            <span class="nav-icon">☰</span>
        </div>        
        <div class="control-panel">
            <!-- 頁籤按鈕 -->
            <div class="tabs-container">
                <button class="tab-button active" onclick="switchAnalysisTab('path')" id="pathTabBtn">
                    📁 路徑分析
                </button>
                <!-- 加入第二個頁籤按鈕 -->
                <button class="tab-button" onclick="switchAnalysisTab('files')" id="filesTabBtn">
                    📂 選擇檔案/資料夾
                </button>
            </div>
            
            <!-- 路徑分析頁籤內容 -->
            <div class="tab-content active" id="pathTabContent">
                <div class="input-group">
                    <label for="pathInput">📁 <span style="margin-left: 5px;">選擇基礎路徑 (包含 anr/ 或 tombstones/ 子資料夾):</span></label>
                    <input type="text" id="pathInput" placeholder="/path/to/logs" value="/R306_ShareFolder/nightrun_log/Demo_stress_Test_log/2025" autocomplete="off">
                    <div id="pathAutocomplete" class="path-autocomplete"></div>
                </div>
                <small style="display: block; margin-top: 8px;">
                    <h2 style="margin-bottom:10px">✨ 功能特色</h2>
                    <ul>
                        <li><span class="icon">🔍</span> <strong>路徑自動建議：</strong> 當您輸入時，工具會自動建議可用的子資料夾，讓您更輕鬆地導航到所需的目錄。</li>
                        <li><span class="icon">📂</span> <strong>自動解壓縮 ZIP 檔案：</strong> 指定路徑下的所有 ZIP 檔案將會自動解壓縮，方便您的操作。</li>
                        <li><span class="icon">🔄</span> <strong>遞迴資料夾搜尋：</strong> 工具會遞迴搜尋所有 <strong>anr</strong> 和 <strong>tombstones</strong> 資料夾，確保不會遺漏任何相關的紀錄檔資料。</li>
                        <li><span class="icon">📜</span> <strong>彈性解析：</strong> ANR 檔案搜尋 "Subject:"，Tombstone 檔案搜尋 "Cmd line:" 或 "Cmdline:"</li>
                        <li><span class="icon">👆</span> <strong>可點擊檔案名稱：</strong> 只需點擊檔案名稱，即可輕鬆查看任何紀錄檔的內容。</li>
                    </ul>
                    <h2 style="margin-top:10px;margin-bottom:10px">💻 支援路徑格式</h2>
                    <div class="path-format">
                        <p><strong>Linux/Unix：</strong> <code>/R306_ShareFolder/nightrun_log/Demo_stress_Test_log</code></p>
                    </div>
                </small>
                <div class="button-group">
                    <button onclick="analyzeLogs()" id="analyzeBtn">開始分析</button>
                    <button onclick="openLoadExcelDialog()" id="loadExcelBtn" class="load-excel-btn">📊 載入 Excel</button>
                    <button onclick="openMergeDialog()" id="mergeExcelMainBtn" class="merge-excel-btn" style="display: inline-flex; position: static; background: #17a2b8;">
                        💹 合併 Excel
                    </button>
                </div>
            </div>
            
            <!-- 選擇檔案/資料夾頁籤內容 -->
            <div class="tab-content" id="filesTabContent">
                <div class="files-selection-area">
                    <!-- 拖曳區域 -->
                    <div class="merge-drop-zone" id="mainFileSelectDropZone" style="margin-bottom: 20px;">
                        <div class="drop-zone-content">
                            <div class="drop-icon">📂</div>
                            <p>拖曳檔案或資料夾到這裡</p>
                            <p class="drop-zone-hint">支援任何檔案格式</p>
                            <input type="file" id="mainFileSelectInput" style="display: none;" multiple>
                            <input type="file" id="mainFolderSelectInput" style="display: none;" webkitdirectory directory multiple>
                            <button class="btn-select-file" id="mainSelectLocalFilesBtn">選擇檔案</button>
                            <button class="btn-select-file" id="mainSelectLocalFolderBtn" style="margin-left: 10px;">選擇資料夾</button>
                        </div>
                    </div>
                    
                    <!-- 已選擇的檔案/資料夾列表 -->
                    <div class="selected-items-section" id="mainSelectedItemsSection" style="display: none;">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                            <h3 style="margin: 0;">已選擇的項目</h3>
                            <div style="display: flex; gap: 10px;">
                                <button class="btn-select-all" onclick="toggleAllMainFiles()">全選</button>
                                <button class="btn-remove-selected" onclick="removeSelectedMainFiles()">移除選中</button>
                                <button class="btn-clear" onclick="clearMainFileSelection()">清除全部</button>
                            </div>
                        </div>
                        <div class="selected-items-list" id="mainSelectedItemsList"></div>
                    </div>
                    
                    <!-- 選項設定 -->
                    <div class="options-section" style="margin-top: 20px;">
                        <label style="display: flex; align-items: center; cursor: pointer;">
                            <input type="checkbox" id="mainAutoGroupFiles" checked style="margin-right: 8px;">
                            自動將獨立的 ANR/Tombstone 檔案分組
                        </label>
                    </div>
                </div>
                <div class="button-group" style="margin-top: 20px;">
                    <button onclick="executeMainFileAnalysis()" id="mainFileAnalysisBtn">開始分析</button>
                    <button onclick="openLoadExcelDialog()" class="load-excel-btn">📊 載入 Excel</button>
                    <button onclick="openMergeDialog()" class="merge-excel-btn" style="display: inline-flex; position: static; background: #17a2b8;">
                        💹 合併 Excel
                    </button>
                </div>
            </div>
            
            <div class="loading" id="loading">
                正在分析中
            </div>
            
            <div id="message" style="margin-top:8px"></div>
        </div>
        <!-- 修改歷史分析文件區塊 -->
        <div id="historySection" style="display: none; margin-top: 20px;">
            <h2 style="margin-bottom: 10px; color: #333;">📚 歷史分析文件</h2>
            <div style="background: #f8f9fa; padding: 15px; border-radius: 8px;">
                <div class="button-group" style="margin-top: 10px; display: flex; flex-wrap: wrap; gap: 10px;">
                    <button onclick="viewHistoryIndex()" id="viewIndexBtn" style="display: none; background: #6f42c1;">
                        📊 查看已有分析結果
                    </button>
                    <button onclick="downloadExistingExcel()" id="downloadExcelBtn" style="display: none; background: #28a745;">
                        📥 匯出 Excel
                    </button>
                    <button onclick="viewExistingHTML()" id="viewHTMLBtn" style="display: none; background: #17a2b8;">
                        📈 已統計分析
                    </button>
                    <button onclick="viewExcelReport()" id="viewExcelReportBtn" style="display: none; background: #ff6b6b;">
                        📊 Excel 報表
                    </button>
                    <button onclick="downloadAnalysisZip()" id="downloadZipBtn" style="display: none; background: #fd7e14;">
                        📦 打包分析結果
                    </button>
                </div>
            </div>
        </div>    
        <div id="results">
            <div class="section-container" id="stats-section-container">
                <div class="logs-table" id="stats-section">
                    <div class="table-header" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                        <div class="section-header" style="padding: 0;">
                            <h3 style="color: white;">統計摘要</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('stats-section-container')">
                            <span class="toggle-icon"></span>
                        </div>
                    </div>
                    <div class="section-content">
                        <div class="stats-summary" id="statsSummary" style="padding: 20px;"></div>
                    </div>
                </div>
            </div>
            
            <div class="section-container" id="charts-section-container">
                <div class="logs-table" id="charts-section" style="margin-top: 40px;">
                    <div class="table-header" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                        <div class="section-header" style="padding: 0;">
                            <h3 style="color: white;">圖表分析</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('charts-section-container')">
                            <span class="toggle-icon"></span>
                        </div>
                    </div>
                    <div class="section-content">
                        <div class="chart-grid" style="padding: 20px;">
                            <div class="chart-container">
                                <h3>Top 10 問題程序 (Process)</h3>
                                <div class="chart-wrapper">
                                    <canvas id="processChart"></canvas>
                                </div>
                            </div>
                            
                            <div class="chart-container">
                                <h3>問題類型分佈</h3>
                                <div class="chart-wrapper">
                                    <canvas id="typeChart"></canvas>
                                </div>
                            </div>
                            
                            <div class="chart-container">
                                <h3>每日問題趨勢</h3>
                                <div class="chart-wrapper">
                                    <canvas id="dailyChart"></canvas>
                                </div>
                            </div>
                            
                            <div class="chart-container">
                                <h3>每小時分佈</h3>
                                <div class="chart-wrapper">
                                    <canvas id="hourlyChart"></canvas>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
 
            <div class="section-container" id="process-summary-section-container">
                <div class="logs-table" id="process-summary-section" style="margin-top: 40px;">
                    <div class="table-header">
                        <div class="section-header" style="padding: 0;">
                            <h3>彙整資訊 - 程序統計</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('process-summary-section-container')">
                            <span class="toggle-icon"></span>
                        </div>
                    </div>
                    <div class="section-content">
                        <div class="table-controls">
                            <div class="search-wrapper" style="display: flex; align-items: center; gap: 10px; flex: 1; max-width: 500px;">
                                <div class="search-box">
                                    <input type="text" id="processSummarySearchInput" placeholder="搜尋程序..." style="flex: 1;">
                                </div>
                                <label class="regex-toggle">
                                    <input type="checkbox" id="processSummaryRegexToggle">
                                    Regex
                                </label>
                                <span class="search-count" id="processSummarySearchCount" style="color: #667eea; font-weight: bold; font-size: 12px; white-space: nowrap; display: none;"></span>
                            </div>
                            <div class="pagination" id="processSummaryPagination">
                                <button onclick="changeProcessSummaryPage('first')" id="processSummaryFirstBtn">第一頁</button>
                                <button onclick="changeProcessSummaryPage(-1)" id="processSummaryPrevBtn">上一頁</button>
                                <span id="processSummaryPageInfo">第 1 頁 / 共 1 頁</span>
                                <button onclick="changeProcessSummaryPage(1)" id="processSummaryNextBtn">下一頁</button>
                                <button onclick="changeProcessSummaryPage('last')" id="processSummaryLastBtn">最後一頁</button>
                            </div>
                        </div>
                        <div class="table-wrapper">
                            <table>
                                <thead>
                                    <tr>
                                        <th style="width: 80px; text-align: center;" class="sortable" onclick="sortProcessSummaryTable('rank')">
                                            排名 <span class="sort-indicator" data-column="rank"></span>
                                        </th>
                                        <th class="sortable" onclick="sortProcessSummaryTable('process')">
                                            程序 <span class="sort-indicator" data-column="process"></span>
                                        </th>
                                        <th class="sortable" onclick="sortProcessSummaryTable('set')">
                                            問題 set <span class="sort-indicator" data-column="set"></span>
                                        </th>                                        
                                        <th style="width: 100px; text-align: center;" class="sortable" onclick="sortProcessSummaryTable('count')">
                                            次數 <span class="sort-indicator" data-column="count">▼</span>
                                        </th>
                                    </tr>
                                </thead>
                                <tbody id="processSummaryTableBody">
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <div class="section-container" id="summary-section-container">
                <div class="logs-table" id="summary-section" style="margin-top: 40px;">
                    <div class="table-header">
                        <div class="section-header" style="padding: 0;">
                            <h3>彙整資訊 - 按類型和程序統計</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('summary-section-container')">
                            <span class="toggle-icon"></span>
                        </div>
                    </div>
                    <div class="section-content">
                        <div class="table-controls">
                            <div class="search-wrapper" style="display: flex; align-items: center; gap: 10px; flex: 1; max-width: 500px;">
                                <div class="search-box">
                                    <input type="text" id="summarySearchInput" placeholder="搜尋類型或程序..." style="flex: 1;">
                                </div>
                                <label class="regex-toggle">
                                    <input type="checkbox" id="summaryRegexToggle">
                                    Regex
                                </label>
                                <span class="search-count" id="summarySearchCount" style="color: #667eea; font-weight: bold; font-size: 12px; white-space: nowrap; display: none;"></span>
                            </div>
                            <div class="pagination" id="summaryPagination">
                                <button onclick="changeSummaryPage('first')" id="summaryFirstBtn">第一頁</button>
                                <button onclick="changeSummaryPage(-1)" id="summaryPrevBtn">上一頁</button>
                                <span id="summaryPageInfo">第 1 頁 / 共 1 頁</span>
                                <button onclick="changeSummaryPage(1)" id="summaryNextBtn">下一頁</button>
                                <button onclick="changeSummaryPage('last')" id="summaryLastBtn">最後一頁</button>
                            </div>
                        </div>
                        <div class="table-wrapper">
                            <table>
                                <thead>
                                    <tr>
                                        <th style="width: 80px; text-align: center;" class="sortable" onclick="sortSummaryTable('rank')">
                                            排名 <span class="sort-indicator" data-column="rank"></span>
                                        </th>
                                        <th style="width: 120px;" class="sortable" onclick="sortSummaryTable('type')">
                                            類型 <span class="sort-indicator" data-column="type"></span>
                                        </th>
                                        <th class="sortable" onclick="sortSummaryTable('process')">
                                            程序 <span class="sort-indicator" data-column="process"></span>
                                        </th>
                                        <th class="sortable" onclick="sortSummaryTable('set')">
                                            問題 set <span class="sort-indicator" data-column="set"></span>
                                        </th>
                                        <th style="width: 100px; text-align: center;" class="sortable" onclick="sortSummaryTable('count')">
                                            次數 <span class="sort-indicator" data-column="count">▼</span>
                                        </th>
                                    </tr>
                                </thead>
                                <tbody id="summaryTableBody">
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="section-container" id="files-section-container">
                <div class="logs-table" id="files-section" style="margin-top: 40px;">
                    <div class="table-header">
                        <div class="section-header" style="padding: 0;">
                            <h3>詳細記錄 (依檔案)</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('files-section-container')">
                            <span class="toggle-icon"></span>
                        </div>                     
                    </div>
                    <div class="section-content">
                        <div class="table-controls">
                            <div class="search-wrapper" style="display: flex; align-items: center; gap: 10px; flex: 1; max-width: 500px;">
                                <div class="search-box">
                                    <input type="text" id="filesSearchInput" placeholder="搜尋檔案名稱..." style="flex: 1;">
                                </div>
                                <label class="regex-toggle">
                                    <input type="checkbox" id="filesRegexToggle">
                                    Regex
                                </label>
                                <span class="search-count" id="filesSearchCount" style="color: #667eea; font-weight: bold; font-size: 12px; white-space: nowrap; display: none;"></span>
                            </div>
                            <div class="pagination" id="filesPagination">
                                <button onclick="changeFilesPage('first')" id="filesFirstBtn">第一頁</button>
                                <button onclick="changeFilesPage(-1)" id="filesPrevBtn">上一頁</button>
                                <span id="filesPageInfo">第 1 頁 / 共 1 頁</span>
                                <button onclick="changeFilesPage(1)" id="filesNextBtn">下一頁</button>
                                <button onclick="changeFilesPage('last')" id="filesLastBtn">最後一頁</button>
                            </div>
                        </div>
                        <div class="logs-table-content">
                            <table>
                                <thead>
                                    <tr>
                                        <th style="width: 60px;" class="sortable" onclick="sortFilesTable('index')">
                                            編號 <span class="sort-indicator" data-column="index"></span>
                                        </th>
                                        <th style="width: 100px;" class="sortable" onclick="sortFilesTable('type')">
                                            類型 <span class="sort-indicator" data-column="type"></span>
                                        </th>
                                        <th class="sortable" onclick="sortFilesTable('processes')">
                                            相關程序 <span class="sort-indicator" data-column="processes"></span>
                                        </th>
                                        <th class="sortable" onclick="sortFilesTable('set')">
                                            問題 set <span class="sort-indicator" data-column="set"></span>
                                        </th>                                  
                                        <th style="width: 200px;" class="sortable" onclick="sortFilesTable('folder_path')">
                                            資料夾路徑 <span class="sort-indicator" data-column="folder_path"></span>
                                        </th>
                                        <th style="width: 250px;" class="sortable" onclick="sortFilesTable('filename')">
                                            檔案 <span class="sort-indicator" data-column="filename"></span>
                                        </th>
                                        <th style="width: 80px;" class="sortable" onclick="sortFilesTable('count')">
                                            次數 <span class="sort-indicator" data-column="count">▼</span>
                                        </th>
                                        <th style="width: 180px;" class="sortable" onclick="sortFilesTable('timestamp')">
                                            時間戳記 <span class="sort-indicator" data-column="timestamp"></span>
                                        </th>
                                    </tr>
                                </thead>
                                <tbody id="filesTableBody">
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="section-container" id="logs-section-container">
                <div class="logs-table" id="logs-section" style="margin-top: 40px;">
                    <div class="table-header">
                        <div class="section-header" style="padding: 0;">
                            <h3>詳細記錄 (依行號)</h3>
                            <a href="#top" class="top-link" style="color: white;">⬆ 回到頂部</a>
                        </div>
                        <div class="section-toggle" onclick="toggleSection('logs-section-container')">
                            <span class="toggle-icon"></span>
                        </div>                    
                    </div>
                    <div class="section-content">
                        <div class="table-controls">
                            <div class="search-wrapper" style="display: flex; align-items: center; gap: 10px; flex: 1; max-width: 500px;">
                                <div class="search-box">
                                    <input type="text" id="logsSearchInput" placeholder="搜尋類型、程序、檔案名稱或資料夾路徑..." style="flex: 1;">
                                </div>
                                <label class="regex-toggle">
                                    <input type="checkbox" id="logsRegexToggle">
                                    Regex
                                </label>
                                <span class="search-count" id="logsSearchCount" style="color: #667eea; font-weight: bold; font-size: 12px; white-space: nowrap; display: none;"></span>
                            </div>
                            <div class="pagination" id="logsPagination">
                                <button onclick="changeLogsPage('first')" id="logsFirstBtn">第一頁</button>
                                <button onclick="changeLogsPage(-1)" id="logsPrevBtn">上一頁</button>
                                <span id="logsPageInfo">第 1 頁 / 共 1 頁</span>
                                <button onclick="changeLogsPage(1)" id="logsNextBtn">下一頁</button>
                                <button onclick="changeLogsPage('last')" id="logsLastBtn">最後一頁</button>
                            </div>
                        </div>
                        <div class="logs-table-content">
                            <table>
                                <thead>
                                    <tr>
                                        <th style="width: 60px;" class="sortable" onclick="sortLogsTable('index')">
                                            編號 <span class="sort-indicator" data-column="index"></span>
                                        </th>
                                        <th style="width: 100px;" class="sortable" onclick="sortLogsTable('type')">
                                            類型 <span class="sort-indicator" data-column="type"></span>
                                        </th>
                                        <th class="sortable" onclick="sortLogsTable('process')">
                                            相關程序 <span class="sort-indicator" data-column="process"></span>
                                        </th>
                                        <th style="width: 80px;" class="sortable" onclick="sortLogsTable('line_number')">
                                            行號 <span class="sort-indicator" data-column="line_number"></span>
                                        </th>
                                        <th class="sortable" onclick="sortLogsTable('set')">
                                            問題 set <span class="sort-indicator" data-column="set"></span>
                                        </th>                                        
                                        <th style="width: 200px;" class="sortable" onclick="sortLogsTable('folder_path')">
                                            資料夾路徑 <span class="sort-indicator" data-column="folder_path"></span>
                                        </th>
                                        <th style="width: 250px;" class="sortable" onclick="sortLogsTable('filename')">
                                            檔案 <span class="sort-indicator" data-column="filename"></span>
                                        </th>
                                        <th style="width: 180px;" class="sortable" onclick="sortLogsTable('timestamp')">
                                            時間戳記 <span class="sort-indicator" data-column="timestamp"></span>
                                        </th>
                                    </tr>
                                </thead>
                                <tbody id="logsTableBody">
                                </tbody>
                            </table>
                        </div>
                   </div>
                </div>
            </div>
        </div>
        <!-- 合併 Excel 彈出視窗 -->
        <div class="merge-dialog-overlay" id="mergeDialogOverlay" style="display: none;">
            <div class="merge-dialog">
                <div class="merge-dialog-header">
                    <h3>💹 合併 Excel 檔案</h3>
                    <button class="merge-dialog-close" onclick="closeMergeDialog()">×</button>
                </div>
                <!-- 合併 Excel 彈出視窗的 HTML -->
                <div class="merge-dialog-body">
                    <!-- 訊息顯示區域 -->
                    <div id="mergeMessage" style="margin-bottom: 15px;"></div>                
                    <!-- 拖曳區域 -->
                    <div class="merge-drop-zone" id="mergeDropZone">
                        <div class="drop-zone-content">
                            <div class="drop-icon">💹</div>
                            <p>拖曳 Excel 檔案到這裡</p>
                            <p class="drop-zone-hint">或</p>
                            <input type="file" id="mergeFileInput" accept=".xlsx" style="display: none;" multiple>
                            <button class="btn-select-file" id="selectFileBtn">選擇檔案</button>
                        </div>
                    </div>
                    
                    <!-- 分隔線 -->
                    <div class="merge-separator">
                        <span>或輸入伺服器路徑</span>
                    </div>
                    
                    <!-- 路徑輸入區域 -->
                    <div class="merge-input-group">
                        <label for="mergePathInput">💹 <span style="margin-left: 5px;">選擇要合併的 Excel 檔案：</span></label>
                        <input type="text" id="mergePathInput" placeholder="/path/to/excel/file.xlsx" autocomplete="off">
                        <div id="mergePathAutocomplete" class="path-autocomplete"></div>
                    </div>
                    
                    <!-- 檔案資訊顯示 -->
                    <div class="merge-file-info" id="mergeFileInfo" style="display: none;">
                        <div class="file-info-content">
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                                <strong>已選擇的項目</strong>
                                <div style="display: flex; gap: 10px;">
                                    <button class="btn-select-all" onclick="toggleAllMergeFiles()">全選</button>
                                    <button class="btn-remove-selected" onclick="removeSelectedFiles()">移除選中</button>
                                    <button class="btn-clear" onclick="clearMergeSelection()">清除全部</button>
                                </div>
                            </div>
                            <div id="selectedMergeFiles"></div>
                        </div>
                    </div>
                    
                    <!-- 支援格式 -->
                    <div class="support-format-section">
                        <h2>✨ 支援格式</h2>
                        <div class="support-format-box">
                              <ul>
                                <li><span class="icon">🔍</span><strong>選擇檔案：</strong>點擊「選擇檔案」按鈕瀏覽本地檔案</li>
                                <li><span class="icon">💹</span><strong>拖曳上傳：</strong>拖曳本地 Excel 檔案到上方區域</li>
                                <li><span class="icon">💹</span><strong>伺服器路徑：</strong>輸入伺服器上的 Excel 檔案路徑</li>
                            </ul>
                        </div>
                    </div>
                </div>
                <div class="merge-dialog-footer">
                    <button class="btn-primary" onclick="executeMerge()" id="mergeExecuteBtn">匯出</button>
                    <button class="btn-secondary" onclick="closeMergeDialog()">取消</button>
                </div>
            </div>
        </div>
        <!-- 選擇檔案/資料夾彈出視窗 -->
        <div class="merge-dialog-overlay" id="fileSelectDialogOverlay" style="display: none;">
            <div class="merge-dialog">
                <div class="merge-dialog-header">
                    <h3>📂 選擇檔案/資料夾</h3>
                    <button class="merge-dialog-close" onclick="closeFileSelectDialog()">×</button>
                </div>
                <div class="merge-dialog-body">
                    <!-- 拖曳區域 -->
                    <div class="merge-drop-zone" id="fileSelectDropZone">
                        <div class="drop-zone-content">
                            <div class="drop-icon">📂</div>
                            <p>拖曳檔案或資料夾到這裡</p>
                            <p class="drop-zone-hint">支援任何檔案格式</p>
                            <input type="file" id="fileSelectInput" style="display: none;" multiple>
                            <input type="file" id="folderSelectInput" style="display: none;" webkitdirectory directory multiple>
                            <button class="btn-select-file" id="selectLocalFilesBtn">選擇檔案</button>
                            <button class="btn-select-file" id="selectLocalFolderBtn" style="margin-left: 10px;">選擇資料夾</button>
                        </div>
                    </div>
                    
                    <!-- 已選擇的檔案/資料夾列表 -->
                    <div class="selected-items-section" id="selectedItemsSection" style="display: none; margin-top: 20px;">
                        <h2>已選擇的項目</h2>
                        <div class="selected-items-list" id="selectedItemsList"></div>
                    </div>
                    
                    <!-- 選項設定 -->
                    <div class="options-section" style="margin-top: 20px;">
                        <label style="display: flex; align-items: center; cursor: pointer;">
                            <input type="checkbox" id="autoGroupFiles" checked style="margin-right: 8px;">
                            自動將獨立的 ANR/Tombstone 檔案分組
                        </label>
                    </div>
                </div>
                <div class="merge-dialog-footer">
                    <button class="btn-primary" onclick="executeFileAnalysis()" id="fileAnalysisExecuteBtn">開始分析</button>
                    <button class="btn-secondary" onclick="closeFileSelectDialog()">取消</button>
                </div>
            </div>
        </div>
        <!-- 鎖定提示對話框 -->
        <div class="lock-dialog-overlay" id="lockDialogOverlay" style="display: none;">
            <div class="lock-dialog">
                <div class="lock-dialog-header">
                    <h3>⚠️ 路徑使用中</h3>
                </div>
                <div class="lock-dialog-body">
                    <div class="lock-warning-icon">
                        <svg width="80" height="80" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
                            <path d="M12 9V13M12 17H12.01M12 3L2 20H22L12 3Z" stroke="#fbbf24" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
                        </svg>
                    </div>
                    <p class="lock-message" id="lockMessage">此路徑正在被其他使用者分析中</p>
                    <div class="lock-time-info" id="lockTimeInfo">
                        <div class="time-display">
                            <span class="time-label">預計剩餘時間：</span>
                            <span class="time-value" id="remainingTime">計算中...</span>
                        </div>
                    </div>
                    <p class="lock-question">是否要自動等待並在分析完成後開始？</p>
                </div>
                <div class="lock-dialog-footer">
                    <button class="btn-wait" onclick="confirmWait()" id="btnWait">
                        <span class="btn-icon">⏳</span> 自動等待
                    </button>
                    <button class="btn-cancel" onclick="cancelWait()">
                        <span class="btn-icon">✖</span> 取消
                    </button>
                </div>
            </div>
        </div>                    
    <footer class="footer">
        <p>&copy; 2025 Copyright by Vince. All rights reserved.</p>
    </footer>
    <script>
        // 全域變數
        let lockDialogResolver = null;
        let waitingForUnlock = false;
        let countdownInterval = null;  // 新增：倒數計時器

        // 全選/取消全選所有檔案
        function toggleAllMergeFiles() {
            const checkboxes = document.querySelectorAll('#selectedMergeFiles input[type="checkbox"]');
            
            if (checkboxes.length === 0) return;
            
            // 檢查是否全部已選中
            const allChecked = Array.from(checkboxes).every(cb => cb.checked);
            
            // 切換選擇狀態
            checkboxes.forEach(cb => {
                cb.checked = !allChecked;
            });
            
            // 更新按鈕文字（可選）
            updateSelectAllButtonText();
        }

        // 更新全選按鈕的文字
        function updateSelectAllButtonText() {
            const checkboxes = document.querySelectorAll('#selectedMergeFiles input[type="checkbox"]');
            const checkedCount = Array.from(checkboxes).filter(cb => cb.checked).length;
            const button = document.querySelector('.btn-select-all');
            
            if (button) {
                if (checkedCount === 0) {
                    button.textContent = '全選';
                } else if (checkedCount === checkboxes.length) {
                    button.textContent = '取消全選';
                } else {
                    button.textContent = `全選 (${checkedCount}/${checkboxes.length})`;
                }
            }
        }

        // 取得選中的檔案索引
        function getSelectedMergeFiles() {
            const selectedIndices = {
                local: [],
                path: []
            };
            
            const items = document.querySelectorAll('#selectedMergeFiles .selected-item');
            items.forEach((item, index) => {
                const checkbox = item.querySelector('input[type="checkbox"]');
                if (checkbox && checkbox.checked) {
                    // 判斷是本地檔案還是路徑檔案
                    const removeBtn = item.querySelector('.btn-clear');
                    if (removeBtn && removeBtn.onclick) {
                        const onclickStr = removeBtn.onclick.toString();
                        if (onclickStr.includes("'local'")) {
                            selectedIndices.local.push(index);
                        } else if (onclickStr.includes("'path'")) {
                            selectedIndices.path.push(index - selectedMergeFiles.length);
                        }
                    }
                }
            });
            
            return selectedIndices;
        }

        // 批量移除選中的檔案
        function removeSelectedFiles() {
            const selectedIndices = getSelectedMergeFiles();
            
            // 從後往前刪除，避免索引錯位
            selectedIndices.local.sort((a, b) => b - a).forEach(index => {
                selectedMergeFiles.splice(index, 1);
            });
            
            selectedIndices.path.sort((a, b) => b - a).forEach(index => {
                selectedMergeFilePaths.splice(index, 1);
            });
            
            // 更新顯示
            updateSelectedFilesDisplay();
        }

        // 對話框專用的訊息顯示函數
        function showDialogMessage(message, type, dialogId = 'mergeMessage') {
            const messageDiv = document.getElementById(dialogId);
            if (messageDiv) {
                messageDiv.className = type;
                messageDiv.innerHTML = message;
                messageDiv.style.display = 'block';
                
                // 3秒後自動隱藏成功訊息
                if (type === 'success') {
                    setTimeout(() => {
                        messageDiv.style.display = 'none';
                    }, 3000);
                }
            }
        }

        // 清除對話框訊息
        function clearDialogMessage(dialogId = 'mergeMessage') {
            const messageDiv = document.getElementById(dialogId);
            if (messageDiv) {
                messageDiv.innerHTML = '';
                messageDiv.className = '';
                messageDiv.style.display = 'none';
            }
        }

        // 顯示鎖定對話框
        function showLockDialog(remainingTime) {
            const overlay = document.getElementById('lockDialogOverlay');
            const timeElement = document.getElementById('remainingTime');
            
            // 清除之前的計時器
            if (countdownInterval) {
                clearInterval(countdownInterval);
            }
            
            let currentTime = remainingTime;
            
            // 更新時間顯示的函數
            function updateTimeDisplay() {
                const minutes = Math.floor(currentTime / 60);
                const seconds = currentTime % 60;
                
                if (minutes > 0) {
                    timeElement.textContent = `${minutes} 分 ${seconds} 秒`;
                } else {
                    timeElement.textContent = `${seconds} 秒`;
                }
                
                // 倒數
                currentTime--;
                
                // 如果時間到了，停止倒數
                if (currentTime < 0) {
                    clearInterval(countdownInterval);
                    timeElement.textContent = '即將完成...';
                }
            }
            
            // 初始顯示
            updateTimeDisplay();
            
            // 每秒更新
            countdownInterval = setInterval(updateTimeDisplay, 1000);
            
            overlay.style.display = 'flex';
            
            // 返回 Promise，等待用戶選擇
            return new Promise((resolve) => {
                lockDialogResolver = resolve;
            });
        }

        // 確認等待
        function confirmWait() {
            const overlay = document.getElementById('lockDialogOverlay');
            overlay.style.display = 'none';
            
            // 清除計時器
            if (countdownInterval) {
                clearInterval(countdownInterval);
                countdownInterval = null;
            }
            
            if (lockDialogResolver) {
                lockDialogResolver(true);
                lockDialogResolver = null;
            }
        }

        // 取消等待
        function cancelWait() {
            const overlay = document.getElementById('lockDialogOverlay');
            overlay.style.display = 'none';
            waitingForUnlock = false;
            
            // 清除計時器
            if (countdownInterval) {
                clearInterval(countdownInterval);
                countdownInterval = null;
            }
            
            if (lockDialogResolver) {
                lockDialogResolver(false);
                lockDialogResolver = null;
            }
        }

        // 添加錯誤對話框
        function showErrorDialog(title, message, question) {
            return new Promise((resolve) => {
                // 重用鎖定對話框，但修改內容
                const overlay = document.getElementById('lockDialogOverlay');
                const header = overlay.querySelector('.lock-dialog-header h3');
                const messageElement = overlay.querySelector('.lock-message');
                const timeInfo = overlay.querySelector('.lock-time-info');
                const questionElement = overlay.querySelector('.lock-question');
                const btnWait = document.getElementById('btnWait');
                
                // 修改內容
                header.innerHTML = '❌ ' + title;
                messageElement.textContent = message;
                timeInfo.style.display = 'none';
                questionElement.textContent = question;
                btnWait.innerHTML = '<span class="btn-icon">✓</span> 繼續';
                
                // 臨時修改 header 背景色
                const headerElement = overlay.querySelector('.lock-dialog-header');
                headerElement.style.background = 'linear-gradient(135deg, #ef4444 0%, #dc2626 100%)';
                
                overlay.style.display = 'flex';
                
                // 設置臨時的處理函數
                const originalConfirm = window.confirmWait;
                window.confirmWait = function() {
                    overlay.style.display = 'none';
                    // 恢復原始設置
                    headerElement.style.background = '';
                    timeInfo.style.display = '';
                    btnWait.innerHTML = '<span class="btn-icon">⏳</span> 自動等待';
                    window.confirmWait = originalConfirm;
                    resolve(true);
                };
                
                const originalCancel = window.cancelWait;
                window.cancelWait = function() {
                    overlay.style.display = 'none';
                    // 恢復原始設置
                    headerElement.style.background = '';
                    timeInfo.style.display = '';
                    btnWait.innerHTML = '<span class="btn-icon">⏳</span> 自動等待';
                    window.cancelWait = originalCancel;
                    resolve(false);
                };
                
                lockDialogResolver = null;
            });
        }

        // 修改 waitForAnalysisUnlock 函數，改善錯誤處理
        async function waitForAnalysisUnlock(path, maxWaitTime = 300000) {
            const startTime = Date.now();
            const pollInterval = 1000;
            waitingForUnlock = true;
            let initialRemainingTime = 0;
            
            showMessage(
                '其他使用者正在分析此路徑，系統將自動等待...<br>' +
                '<button onclick="cancelWaiting()" class="btn btn-sm btn-danger" style="margin-top: 10px;">取消等待</button>' +
                createWaitingProgressBar(),
                'info'
            );
            
            return new Promise((resolve, reject) => {
                const checkLock = async () => {
                    if (!waitingForUnlock) {
                        showMessage('已取消等待', 'info');
                        reject(new Error('使用者取消等待'));
                        return;
                    }
                    
                    try {
                        const response = await fetch('/check-analysis-lock', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({ path: path })
                        });
                        
                        if (!response.ok) {
                            throw new Error(`HTTP error! status: ${response.status}`);
                        }
                        
                        const data = await response.json();
                        
                        if (!data.locked) {
                            showMessage('分析鎖已釋放，開始進行分析...', 'success');
                            resolve();
                        } else {
                            const remainingTime = data.remaining_time || 0;
                            const minutes = Math.floor(remainingTime / 60);
                            const seconds = remainingTime % 60;
                            
                            // 初始化初始剩餘時間
                            if (initialRemainingTime === 0) {
                                initialRemainingTime = remainingTime;
                            }
                            
                            // 計算進度
                            const progress = Math.max(0, Math.min(100, 
                                ((initialRemainingTime - remainingTime) / initialRemainingTime) * 100
                            ));
                            
                            // 更新進度條
                            const progressBar = document.getElementById('waitingProgressBar');
                            const progressText = document.getElementById('waitingProgressText');
                            
                            if (progressBar) {
                                progressBar.style.width = progress + '%';
                            }
                            
                            if (progressText) {
                                progressText.textContent = `${Math.round(progress)}%`;
                            }
                            
                            const message = '正在等待其他分析完成...<br>' +
                                '預計剩餘時間: ' + minutes + ' 分 ' + seconds + ' 秒<br>' +
                                '<small>系統會自動開始分析，請勿關閉此頁面</small><br>' +
                                '<button onclick="cancelWaiting()" class="btn btn-sm btn-danger" style="margin-top: 10px;">取消等待</button>';
                            
                            showMessage(message, 'info');
                            
                            // 檢查是否超時
                            if (Date.now() - startTime > maxWaitTime) {
                                throw new Error('等待超時');
                            }
                            
                            setTimeout(checkLock, pollInterval);
                        }
                    } catch (error) {
                        console.error('檢查鎖狀態時發生錯誤:', error);
                        // 不要立即失敗，繼續重試
                        if (Date.now() - startTime < maxWaitTime) {
                            setTimeout(checkLock, pollInterval);
                        } else {
                            showMessage('檢查鎖狀態失敗，已超過最大等待時間', 'error');
                            reject(error);
                        }
                    }
                };
                
                checkLock();
            });
        }

        // 切換所有檔案的選擇狀態
        function toggleAllFiles() {
            const checkboxes = document.querySelectorAll('#selectedMergeFiles input[type="checkbox"]');
            const allChecked = Array.from(checkboxes).every(cb => cb.checked);
            
            checkboxes.forEach(cb => {
                cb.checked = !allChecked;
            });
        }

        // 取得選中的檔案
        function getSelectedFiles() {
            const checkboxes = document.querySelectorAll('#selectedMergeFiles input[type="checkbox"]:checked');
            const selectedIndices = [];
            
            checkboxes.forEach((cb, index) => {
                if (cb.checked) {
                    selectedIndices.push(index);
                }
            });
            
            return selectedIndices;
        }

    </script>
    <script>
        let currentAnalysisId = null;
        let allLogs = [];
        let allSummary = [];
        let allFileStats = [];
        let charts = {};
        
        // Pagination state
        let summaryPage = 1;
        let logsPage = 1;
        let filesPage = 1;
        const itemsPerPage = 10;
        
        // Filtered data
        let filteredSummary = [];
        let filteredLogs = [];
        let filteredFiles = [];
        
        // Autocomplete state
        let selectedSuggestionIndex = -1;
        let currentSuggestions = [];
        let autocompleteTimeout = null;

        let allProcessSummary = [];
        let filteredProcessSummary = [];
        let processSummaryPage = 1;
        let processSummarySort = { column: 'count', order: 'desc' };

        // 排序狀態
        let summarySort = { column: 'count', order: 'desc' };
        let filesSort = { column: 'count', order: 'desc' };
        let logsSort = { column: null, order: 'asc' };
        
        // 添加導覽列開關狀態
        let navBarOpen = false;
        
        // 區塊收縮狀態
        let sectionStates = {
            'stats-section-container': false,
            'charts-section-container': false,
            'process-summary-section-container': false,
            'summary-section-container': false,
            'files-section-container': false,  // 改為 -container
            'logs-section-container': false     // 改為 -container
        };

        // 切換單個區塊
        function toggleSection(sectionId) {
            let container;
            let actualSectionId;
            
            // 先嘗試直接找到容器
            container = document.getElementById(sectionId);
            
            // 如果沒找到，嘗試加上 -container 後綴
            if (!container && !sectionId.endsWith('-container')) {
                container = document.getElementById(sectionId + '-container');
                actualSectionId = sectionId;
            } else if (container && sectionId.endsWith('-container')) {
                // 如果找到了且有 -container 後綴，去掉它
                actualSectionId = sectionId.replace('-container', '');
            } else if (container) {
                // 找到了但沒有 -container 後綴
                actualSectionId = sectionId;
            }
            
            if (!container) {
                console.error('找不到區塊容器:', sectionId);
                return;
            }
            
            container.classList.toggle('collapsed');
            sectionStates[actualSectionId] = container.classList.contains('collapsed');
            
            // 如果是圖表區塊，需要重新渲染圖表
            if (actualSectionId === 'charts-section' && !sectionStates[actualSectionId]) {
                setTimeout(() => {
                    // 重新渲染所有圖表
                    Object.values(charts).forEach(chart => {
                        if (chart) chart.resize();
                    });
                }, 300);
            }
            
            updateToggleTooltips();
        }

        // 切換所有區塊
        function toggleAllSections() {
            const allCollapsed = Object.values(sectionStates).every(state => state);
            const icon = document.getElementById('globalToggleIcon');
            
            // 如果全部收縮，則全部展開；否則全部收縮
            const newState = !allCollapsed;
            
            Object.keys(sectionStates).forEach(sectionId => {
                let container = document.getElementById(sectionId + '-container');
                // 如果沒找到，嘗試不帶 -container 的
                if (!container) {
                    container = document.getElementById(sectionId);
                }
                
                if (container) {
                    if (newState) {
                        container.classList.add('collapsed');
                    } else {
                        container.classList.remove('collapsed');
                    }
                    sectionStates[sectionId] = newState;
                }
            });
            
            // 更新圖標
            icon.textContent = newState ? '⊖' : '⊕';
            
            // 如果展開圖表區塊，重新渲染
            if (!newState) {
                setTimeout(() => {
                    Object.values(charts).forEach(chart => {
                        if (chart) chart.resize();
                    });
                }, 300);
            }
            
            // 更新全局按鈕的 Tooltip
            const globalBtn = document.getElementById('globalToggleBtn');
            if (globalBtn) {
                globalBtn.setAttribute('data-tooltip', allCollapsed ? '全部展開' : '全部收合');
            }
            
            // 更新 tooltip
            updateToggleTooltips();
            updateGlobalToggleTooltip();
    
        }
        
        let analysisIndexPath = null;  // 儲存分析結果的 index.html 路徑

        // Initialize autocomplete
        document.addEventListener('DOMContentLoaded', function() {
            const pathInput = document.getElementById('pathInput');
            const autocompleteDiv = document.getElementById('pathAutocomplete');

            // 輸入框失去焦點時檢查
            pathInput.addEventListener('blur', function() {
                checkExistingAnalysis(this.value);
            });
            
            // 按下 Enter 時也檢查
            pathInput.addEventListener('keypress', function(e) {
                if (e.key === 'Enter') {
                    e.preventDefault();
                    checkExistingAnalysis(this.value);
                }
            });
            
            // 初始檢查
            if (pathInput.value) {
                checkExistingAnalysis(pathInput.value);
            }
                                    
            // Handle input events
            pathInput.addEventListener('input', function(e) {
                clearTimeout(autocompleteTimeout);
                const value = e.target.value;
                
                // Always fetch suggestions, even for empty input
                autocompleteTimeout = setTimeout(() => {
                    fetchPathSuggestions(value);
                }, 300);
            });
            
            // Handle keyboard navigation
            pathInput.addEventListener('keydown', function(e) {
                switch(e.key) {
                    case 'ArrowDown':
                        if (currentSuggestions.length > 0) {
                            e.preventDefault();
                            selectSuggestion(selectedSuggestionIndex + 1);
                        }
                        break;
                    case 'ArrowUp':
                        if (currentSuggestions.length > 0) {
                            e.preventDefault();
                            selectSuggestion(selectedSuggestionIndex - 1);
                        }
                        break;
                    case 'Enter':
                        if (currentSuggestions.length > 0 && selectedSuggestionIndex >= 0) {
                            e.preventDefault();
                            const cleanPath = currentSuggestions[selectedSuggestionIndex].replace(/ [⭐📁]$/, '');
                            applySuggestion(cleanPath);
                        } else if (currentSuggestions.length === 0 || selectedSuggestionIndex === -1) {
                            // 如果沒有選擇任何建議，就關閉提示框
                            hideAutocomplete();
                        }
                        break;
                    case 'Tab':
                        if (currentSuggestions.length > 0) {
                            e.preventDefault();
                            // If no selection, select first suggestion
                            const index = selectedSuggestionIndex >= 0 ? selectedSuggestionIndex : 0;
                            applySuggestion(currentSuggestions[index].replace(/ [⭐📁]$/, ''));
                        }
                        break;
                    case 'Escape':
                        hideAutocomplete();
                        break;
                }
            });
            
            // Handle focus
            pathInput.addEventListener('focus', function() {
                if (currentSuggestions.length > 0) {
                    showAutocomplete();
                } else {
                    // Fetch suggestions for current value
                    fetchPathSuggestions(this.value);
                }
            });

            pathInput.addEventListener('blur', function(e) {
                // 延遲執行，給予時間點擊提示項
                setTimeout(() => {
                    // 檢查當前焦點是否在自動完成框內
                    if (!document.activeElement.closest('#pathAutocomplete')) {
                        hideAutocomplete();
                    }
                }, 200);
            });
            
            // Handle click outside
            document.addEventListener('click', function(e) {
                if (navBarOpen && 
                    !e.target.closest('.nav-bar') && 
                    !e.target.closest('.nav-toggle-btn')) {
                    toggleNavBar();
                }
                const pathInput = document.getElementById('pathInput');
                const autocompleteDiv = document.getElementById('pathAutocomplete');
                
                // 如果點擊的不是輸入框或自動完成框，則隱藏提示框
                if (!pathInput.contains(e.target) && !autocompleteDiv.contains(e.target)) {
                    hideAutocomplete();
                }                
            });
            
            // Back to top button functionality
            window.addEventListener('scroll', function() {
                const backToTopBtn = document.getElementById('backToTop');
                const navToggleBtn = document.getElementById('navToggleBtn');
                const analysisResultBtn = document.getElementById('analysisResultBtn');
                
                if (window.pageYOffset > 300) {
                    backToTopBtn.classList.add('show');
                    if (document.getElementById('results').style.display !== 'none') {
                        navToggleBtn.classList.add('show');
                        // 如果有分析結果，顯示分析結果按鈕
                        if (window.vpAnalyzeSuccess && analysisIndexPath) {
                            analysisResultBtn.classList.add('show');
                        }
                    }
                } else {
                    backToTopBtn.classList.remove('show');
                    navToggleBtn.classList.remove('show');
                    analysisResultBtn.classList.remove('show');
                    // 滾動到頂部時關閉導覽列
                    if (navBarOpen) {
                        toggleNavBar();
                    }
                }
            });

            const tooltipContainers = document.querySelectorAll('.tooltip-container');
            tooltipContainers.forEach(container => {
                container.addEventListener('mouseenter', function(e) {
                    const tooltip = this.querySelector('.tooltip-text');
                    if (!tooltip) return;
                    
                    const rect = this.getBoundingClientRect();
                    const tooltipHeight = 150; // 預估的 tooltip 高度
                    
                    // 如果上方空間不足，改為顯示在下方
                    if (rect.top < tooltipHeight) {
                        tooltip.style.bottom = 'auto';
                        tooltip.style.top = '125%';
                        
                        // 調整箭頭
                        const arrow = window.getComputedStyle(tooltip, '::after');
                        if (arrow) {
                            tooltip.classList.add('tooltip-below');
                        }
                    } else {
                        // 恢復預設（顯示在上方）
                        tooltip.style.bottom = '125%';
                        tooltip.style.top = 'auto';
                        tooltip.classList.remove('tooltip-below');
                    }
                });
            });

            // 為右下角浮動按鈕添加 tooltip
            const globalToggleBtn = document.getElementById('globalToggleBtn');
            const backToTopBtn = document.getElementById('backToTop');
            const navToggleBtn = document.getElementById('navToggleBtn');
            
            if (backToTopBtn) backToTopBtn.setAttribute('data-tooltip', '回到頂部');
            if (navToggleBtn) navToggleBtn.setAttribute('data-tooltip', '快速導覽');
            if (globalToggleBtn) globalToggleBtn.setAttribute('data-tooltip', '全部展開/收合');
            
            // 添加 tooltip-container class
            if (backToTopBtn) {
                backToTopBtn.classList.add('tooltip-container');
                backToTopBtn.setAttribute('data-tooltip', '回到頂部');
            }
            if (navToggleBtn) {
                navToggleBtn.classList.add('tooltip-container');
                navToggleBtn.setAttribute('data-tooltip', '快速導覽');
            }
            if (globalToggleBtn) {
                globalToggleBtn.classList.add('tooltip-container');
                globalToggleBtn.setAttribute('data-tooltip', '全部展開/收合');
            }
    
            // 強制重置浮動按鈕位置
            const floatingButtons = [
                document.getElementById('backToTop'),
                document.getElementById('navToggleBtn'),
                document.getElementById('globalToggleBtn')
            ];
            
            floatingButtons.forEach(btn => {
                if (btn) {
                    // 移除可能的內聯樣式
                    btn.style.removeProperty('left');
                    btn.style.removeProperty('top');
                    btn.style.removeProperty('transform');
                }
            });
            
            // 為所有開合按鈕設置 Tooltip
            updateToggleTooltips();
            
            // 為右下角浮動按鈕創建 tooltip
            setupFloatingTooltips();
            
            // 為區塊開合按鈕設置 Tooltip
            updateToggleTooltips();            
            
        });

        // 新增查看 Excel 報表的函數
        function viewExcelReport() {
            if (historyAnalysisInfo && historyAnalysisInfo.excel_report_path) {
                console.log('開啟 Excel 報表:', historyAnalysisInfo.excel_report_path);
                try {
                    // 使用與已統計分析相同的路由
                    window.open('/view-analysis-html?path=' + encodeURIComponent(historyAnalysisInfo.excel_report_path), '_blank');
                } catch (error) {
                    console.error('開啟 Excel 報表時發生錯誤:', error);
                    showMessage('開啟 Excel 報表失敗', 'error');
                }
            } else {
                showMessage('找不到 Excel 報表', 'error');
            }
        }

        // 查看已有分析結果
        function viewExistingAnalysis() {
            if (window.existingAnalysisPath) {
                window.open('/view-analysis-report?path=' + encodeURIComponent(window.existingAnalysisPath), '_blank');
            } else {
                showMessage('找不到分析結果', 'error');
            }
        }

        // 新增函數：設置浮動按鈕的 tooltip
        function setupFloatingTooltips() {
            const tooltipData = [
                { id: 'backToTop', text: '回到頂部' },
                { id: 'navToggleBtn', text: '快速導覽' },
                { id: 'globalToggleBtn', text: '全部展開/收合' }
            ];
            
            tooltipData.forEach(item => {
                const btn = document.getElementById(item.id);
                if (btn) {
                    // 創建 tooltip 元素
                    const tooltip = document.createElement('span');
                    tooltip.className = 'floating-tooltip';
                    tooltip.textContent = item.text;
                    btn.appendChild(tooltip);
                    
                    // 特別處理導覽按鈕
                    if (item.id === 'navToggleBtn') {
                        btn.addEventListener('mouseenter', function() {
                            // 只有在菜單關閉時才顯示 tooltip
                            if (!navBarOpen) {
                                tooltip.style.opacity = '1';
                                tooltip.style.visibility = 'visible';
                            }
                        });
                    } else {
                        // 其他按鈕正常處理
                        btn.addEventListener('mouseenter', function() {
                            tooltip.style.opacity = '1';
                            tooltip.style.visibility = 'visible';
                        });
                    }
                    
                    btn.addEventListener('mouseleave', function() {
                        // 導覽按鈕特殊處理
                        if (item.id === 'navToggleBtn' && navBarOpen) {
                            return; // 菜單開啟時不恢復 tooltip
                        }
                        tooltip.style.opacity = '0';
                        tooltip.style.visibility = 'hidden';
                    });
                }
            });
        }

        // 更新全局按鈕的 tooltip 文字
        function updateGlobalToggleTooltip() {
            const globalBtn = document.getElementById('globalToggleBtn');
            if (globalBtn) {
                const tooltip = globalBtn.querySelector('.floating-tooltip');
                if (tooltip) {
                    const allCollapsed = Object.values(sectionStates).every(state => state);
                    tooltip.textContent = allCollapsed ? '全部展開' : '全部收合';
                }
            }
        }

        // 新增函數：更新開合按鈕的 Tooltip
        function updateToggleTooltips() {
            document.querySelectorAll('.section-toggle').forEach(toggle => {
                const container = toggle.closest('.section-container');
                if (container && container.classList.contains('collapsed')) {
                    toggle.setAttribute('data-tooltip', '展開區塊');
                } else {
                    toggle.setAttribute('data-tooltip', '收合區塊');
                }
            });
        }

        function toggleNavBar() {
            const navBar = document.getElementById('navBar');
            const toggleBtn = document.getElementById('navToggleBtn');
            const tooltip = toggleBtn.querySelector('.floating-tooltip');
            
            navBarOpen = !navBarOpen;
            
            if (navBarOpen) {
                navBar.classList.remove('animating-out');
                navBar.classList.add('show', 'animating-in');
                toggleBtn.classList.add('active');
                
                // 隱藏 tooltip
                if (tooltip) {
                    tooltip.style.opacity = '0';
                    tooltip.style.visibility = 'hidden';
                }
                
                // 移除動畫類別
                setTimeout(() => {
                    navBar.classList.remove('animating-in');
                }, 300);
            } else {
                navBar.classList.remove('animating-in');
                navBar.classList.add('animating-out');
                toggleBtn.classList.remove('active');
                
                // 延遲顯示 tooltip（等菜單完全關閉後）
                setTimeout(() => {
                    if (tooltip) {
                        tooltip.style.opacity = '';
                        tooltip.style.visibility = '';
                    }
                }, 300);
                
                // 延遲移除 show 類別
                setTimeout(() => {
                    navBar.classList.remove('show', 'animating-out');
                }, 300);
            }
        }
        
        async function fetchPathSuggestions(path) {
            const autocompleteDiv = document.getElementById('pathAutocomplete');
            
            // Show loading
            autocompleteDiv.innerHTML = '<div class="path-loading">載入中...</div>';
            showAutocomplete();
            
            try {
                const response = await fetch('/suggest-path', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ path: path })
                });
                
                const data = await response.json();
                currentSuggestions = data.suggestions || [];
                selectedSuggestionIndex = -1;
                
                if (currentSuggestions.length > 0) {
                    displaySuggestions(currentSuggestions);
                } else {
                    autocompleteDiv.innerHTML = '<div class="path-loading">沒有找到符合的路徑</div>';
                }
            } catch (error) {
                console.error('Error fetching suggestions:', error);
                hideAutocomplete();
            }
        }
        
        function displaySuggestions(suggestions) {
            const autocompleteDiv = document.getElementById('pathAutocomplete');
            autocompleteDiv.innerHTML = '';
            
            suggestions.forEach((suggestion, index) => {
                const div = document.createElement('div');
                div.className = 'path-suggestion';
                div.dataset.index = index;
                
                // Check if this is a marked folder
                if (suggestion.includes(' ⭐')) {
                    const cleanPath = suggestion.replace(' ⭐', '');
                    div.innerHTML = `${escapeHtml(cleanPath)}<span class="star">⭐</span>`;
                    div.dataset.path = cleanPath;
                } else if (suggestion.includes(' 📁')) {
                    const cleanPath = suggestion.replace(' 📁', '');
                    div.innerHTML = `${escapeHtml(cleanPath)}<span class="star">📁</span>`;
                    div.dataset.path = cleanPath;
                } else {
                    div.textContent = suggestion;
                    div.dataset.path = suggestion;
                }
                
                div.addEventListener('click', function() {
                    applySuggestion(div.dataset.path);
                });
                
                div.addEventListener('mouseenter', function() {
                    selectSuggestion(index);
                });
                
                autocompleteDiv.appendChild(div);
            });
            
            showAutocomplete();
        }
        
        function selectSuggestion(index) {
            const suggestions = document.querySelectorAll('.path-suggestion');
            
            // Remove previous selection
            suggestions.forEach(s => s.classList.remove('selected'));
            
            // Validate index
            if (index < 0) index = suggestions.length - 1;
            if (index >= suggestions.length) index = 0;
            
            selectedSuggestionIndex = index;
            
            // Add selection
            if (index >= 0 && index < suggestions.length) {
                suggestions[index].classList.add('selected');
                suggestions[index].scrollIntoView({ block: 'nearest' });
            }
        }
        
        function applySuggestion(suggestion) {
            const pathInput = document.getElementById('pathInput');
            // Use the clean path from dataset
            pathInput.value = suggestion;
            hideAutocomplete();

            // 選擇路徑後立即檢查是否有分析結果
            checkExistingAnalysis(suggestion);

            // Trigger another suggestion fetch if path ends with separator
            if (suggestion.endsWith('/') || suggestion.endsWith('\\')) {
                fetchPathSuggestions(suggestion);
            }
        }
        
        function showAutocomplete() {
            document.getElementById('pathAutocomplete').style.display = 'block';
            document.getElementById('pathInput').classList.add('autocomplete-open');
        }
        
        function hideAutocomplete() {
            document.getElementById('pathAutocomplete').style.display = 'none';
            document.getElementById('pathInput').classList.remove('autocomplete-open');
            selectedSuggestionIndex = -1;
            currentSuggestions = [];
        }
        
        function escapeHtml(text) {
            if (!text) return '';
            const div = document.createElement('div');
            text = String(text);
            div.textContent = text;
            return div.innerHTML;
        }
        
        function scrollToTop() {
            window.scrollTo({
                top: 0,
                behavior: 'smooth'
            });
        }

        // 新增 AI 結果匯出函數
        async function exportAIResults() {
            const path = document.getElementById('pathInput').value;
            if (!path || !window.vpAnalyzeOutputPath) {
                showMessage('無法匯出：找不到分析結果', 'error');
                return;
            }
            
            const exportBtn = document.getElementById('exportExcelBtn');
            if (!exportBtn) return;
            
            exportBtn.disabled = true;
            exportBtn.textContent = '匯出中...';
            
            try {
                const response = await fetch('/export-ai-excel', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        path: path,
                        analysis_output_path: window.vpAnalyzeOutputPath,
                        logs: allLogs
                    })
                });
                
                if (response.ok) {
                    // 下載檔案
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    
                    // 從 header 獲取檔名
                    const contentDisposition = response.headers.get('content-disposition');
                    let filename = 'anr_tombstone_result.xlsx';
                    if (contentDisposition) {
                        const filenameMatch = contentDisposition.match(/filename="?(.+?)"?$/);
                        if (filenameMatch) {
                            filename = filenameMatch[1];
                        }
                    }
                    
                    a.download = filename;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    
                    // 檢查是否創建了新的 all_excel 或檔案已存在
                    const allExcelCreated = response.headers.get('X-All-Excel-Created') === 'true';
                    const allExcelExists = response.headers.get('X-All-Excel-Exists') === 'true';
                    const allExcelPath = response.headers.get('X-All-Excel-Path');
                    
                    if (allExcelCreated) {
                        showMessage('Excel 匯出成功，並已創建 all_anr_tombstone_result.xlsx', 'success');
                    } else {
                        showMessage('Excel 匯出成功', 'success');
                    }
                    
                    // 標記當前分析已匯出
                    window.currentAnalysisExported = true;
                    
                } else {
                    const error = await response.text();
                    try {
                        const errorData = JSON.parse(error);
                        showMessage('匯出失敗: ' + (errorData.error || '未知錯誤'), 'error');
                    } catch {
                        showMessage('匯出失敗: ' + error, 'error');
                    }
                }
            } catch (error) {
                showMessage('匯出失敗: ' + error.message, 'error');
            } finally {
                exportBtn.disabled = false;
                exportBtn.textContent = '匯出 Excel';
            }
        }

        async function analyzeLogs() {
            const path = document.getElementById('pathInput').value;
            if (!path) {
                showMessage('請輸入路徑', 'error');
                return;
            }

            // === 新增：先檢查是否被鎖定 ===
            try {
                const lockResponse = await fetch('/check-analysis-lock', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ path: path })
                });
                
                const lockData = await lockResponse.json();
                
                if (lockData.locked) {
                    const remainingTime = lockData.remaining_time || 0;
                    
                    // 使用美化的對話框
                    const userChoice = await showLockDialog(remainingTime);
                    
                    if (userChoice) {
                        // 使用者選擇等待
                        try {
                            await waitForAnalysisUnlock(path);
                            // 等待成功，繼續執行分析
                        } catch (error) {
                            // 檢查是否為使用者取消
                            if (error.message === '使用者取消等待') {
                                // 不顯示錯誤，直接返回
                                return;
                            }
                            console.error('等待失敗:', error);
                            showMessage('等待過程中發生錯誤', 'error');
                            return;
                        }
                    } else {
                        // 使用者選擇不等待，直接返回，不顯示錯誤
                        return;
                    }
                }
            } catch (error) {
                console.error('檢查鎖狀態失敗:', error);
                // 如果檢查失敗，詢問是否繼續
                const continueAnyway = await showErrorDialog(
                    '無法檢查路徑狀態',
                    '無法確認此路徑是否正在被其他使用者分析。',
                    '是否要繼續執行分析？'
                );
                if (!continueAnyway) {
                    return;
                }
            }
            
            // 重置匯出狀態
            window.currentAnalysisExported = false;
            window.hasCurrentAnalysis = false;

            // === 新增：隱藏歷史區塊 ===
            document.getElementById('historySection').style.display = 'none';

            // === 新增：隱藏所有浮動按鈕 ===
            const floatingButtons = [
                'backToTop',
                'navToggleBtn', 
                'globalToggleBtn',
                'analysisResultBtn'
            ];
            
            floatingButtons.forEach(id => {
                const btn = document.getElementById(id);
                if (btn) {
                    btn.classList.remove('show');
                }
            });
            
            // === 新增：隱藏 header 區域的所有按鈕 ===
            const headerButtons = [
                'exportExcelBtn',
                'exportExcelReportBtn',
                'mergeExcelBtn',
                'exportHtmlBtn',
                'downloadCurrentZipBtn'
            ];
            
            headerButtons.forEach(id => {
                const btn = document.getElementById(id);
                if (btn) {
                    btn.style.display = 'none';
                }
            });
                        
            // 確保導覽列也關閉
            const navBar = document.getElementById('navBar');
            if (navBar && navBar.classList.contains('show')) {
                toggleNavBar();
            }

            analysisIndexPath = null;
            
            // Disable analyze button
            document.getElementById('analyzeBtn').disabled = true;
            document.getElementById('loading').style.display = 'block';
            document.getElementById('results').style.display = 'none';
            
            // 隱藏當次分析的匯出按鈕
            const exportExcelBtn = document.getElementById('exportExcelBtn');
            if (exportExcelBtn) exportExcelBtn.style.display = 'none';
            
            const allExcelPathInfo = document.getElementById('allExcelPathInfo');
            if (allExcelPathInfo) allExcelPathInfo.style.display = 'none';
            
            clearMessage();
            
            try {
                const response = await fetch('/analyze', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ path: path })
                });
                
                const data = await response.json();
                
                // === 新增：處理鎖定狀態 ===
                if (response.status === 423) {  // Locked status
                    // 如果還是被鎖定（可能是新的鎖），詢問是否重試
                    const retry = confirm(
                        data.error + '\n\n' +
                        '是否要自動等待並重試？'
                    );
                    
                    if (retry) {
                        try {
                            await waitForAnalysisUnlock(path);
                            // 遞迴呼叫自己重新分析
                            analyzeLogs();
                            return;
                        } catch (error) {
                            console.error('等待失敗:', error);
                        }
                    }
                    
                    document.getElementById('analyzeBtn').disabled = false;
                    document.getElementById('loading').style.display = 'none';
                    return;
                }
                
                if (!response.ok) {
                    throw new Error(data.error || 'Analysis failed');
                }
                
                if (data.files_with_cmdline === 0) {
                    showMessage('警告: 在 anr/ 和 tombstones/ 資料夾中沒有找到包含 Cmd line: 或 Cmdline: 的檔案', 'error');
                    document.getElementById('analyzeBtn').disabled = false;
                    document.getElementById('loading').style.display = 'none';
                    
                    // 確保結果區域隱藏
                    document.getElementById('results').style.display = 'none';
                    
                    // 確保所有匯出按鈕都隱藏
                    const exportButtons = [
                        'exportExcelBtn',
                        'exportExcelReportBtn', 
                        'mergeExcelBtn',
                        'exportHtmlBtn',
                        'downloadCurrentZipBtn'
                    ];
                    
                    exportButtons.forEach(id => {
                        const btn = document.getElementById(id);
                        if (btn) btn.style.display = 'none';
                    });
                    
                    // 確保導覽相關按鈕也隱藏
                    const navButtons = ['backToTop', 'navToggleBtn', 'globalToggleBtn', 'analysisResultBtn'];
                    navButtons.forEach(id => {
                        const btn = document.getElementById(id);
                        if (btn) btn.classList.remove('show');
                    });
                    
                    return;
                }
                
                currentAnalysisId = data.analysis_id;
                
                // Sort logs by timestamp
                const sortedLogs = data.logs.sort((a, b) => {
                    if (!a.timestamp && !b.timestamp) return 0;
                    if (!a.timestamp) return 1;
                    if (!b.timestamp) return -1;
                    return a.timestamp.localeCompare(b.timestamp);
                });
                
                allLogs = sortedLogs;
                allSummary = data.statistics.type_process_summary || [];
                allFileStats = data.file_statistics || [];
                
                // Reset filters and pagination
                resetFiltersAndPagination();
                
                // 保存分析輸出路徑和狀態
                window.vpAnalyzeOutputPath = data.vp_analyze_output_path;
                window.vpAnalyzeSuccess = data.vp_analyze_success;
                window.hasCurrentAnalysis = true; // 標記有當前分析結果
                
                // 設定分析結果按鈕
                if (data.vp_analyze_success && data.vp_analyze_output_path) {
                    analysisIndexPath = '/view-analysis-report?path=' + encodeURIComponent(data.vp_analyze_output_path);
                    const analysisBtn = document.getElementById('analysisResultBtn');
                    analysisBtn.href = analysisIndexPath;
                    
                    // 顯示 Excel 匯出按鈕
                    if (exportExcelBtn) {
                        exportExcelBtn.style.display = 'block';
                    }

                    // 顯示 Excel 報表按鈕
                    const exportExcelReportBtn = document.getElementById('exportExcelReportBtn');
                    if (exportExcelReportBtn) {
                        exportExcelReportBtn.style.display = 'block';
                    }

                    // 顯示合併 Excel 按鈕
                    const mergeExcelBtn = document.getElementById('mergeExcelBtn');
                    if (mergeExcelBtn) {
                        mergeExcelBtn.style.display = 'block';
                    }

                    // 自動產生並儲存 Excel 到分析資料夾
                    try {
                        await autoExportExcel(data.vp_analyze_output_path);
                        console.log('已自動產生 Excel 檔案');
                    } catch (error) {
                        console.error('自動產生 Excel 失敗:', error);
                    }
                    
                    // 自動產生並儲存 HTML 到分析資料夾
                    try {
                        await autoExportHTML(data.vp_analyze_output_path);
                        console.log('已自動產生 HTML 檔案');
                    } catch (error) {
                        console.error('自動產生 HTML 失敗:', error);
                    }
                    
                    // 自動產生並儲存 Excel 報表到分析資料夾
                    try {
                        await autoExportExcelReport(data.vp_analyze_output_path);
                        console.log('已自動產生 Excel 報表');
                    } catch (error) {
                        console.error('自動產生 Excel 報表失敗:', error);
                    }
                                        
                    // 顯示分析結果打包按鈕
                    const downloadCurrentZipBtn = document.getElementById('downloadCurrentZipBtn');
                    if (downloadCurrentZipBtn) {
                        downloadCurrentZipBtn.style.display = 'block';
                    }
                }
                
                console.log('vp_analyze 執行結果:', {
                    success: data.vp_analyze_success,
                    outputPath: data.vp_analyze_output_path,
                    error: data.vp_analyze_error
                });
                
                // Update UI
                updateResults(data);
                
                // 只有在有資料時才顯示匯出按鈕
                if (data.files_with_cmdline > 0) {
                    document.getElementById('exportHtmlBtn').style.display = 'block';
                    
                    // 其他按鈕的顯示邏輯也要加入條件判斷
                    if (data.vp_analyze_success && data.vp_analyze_output_path) {
                        const exportExcelBtn = document.getElementById('exportExcelBtn');
                        if (exportExcelBtn) {
                            exportExcelBtn.style.display = 'block';
                        }
                        
                        const exportExcelReportBtn = document.getElementById('exportExcelReportBtn');
                        if (exportExcelReportBtn) {
                            exportExcelReportBtn.style.display = 'block';
                        }
                        
                        const mergeExcelBtn = document.getElementById('mergeExcelBtn');
                        if (mergeExcelBtn) {
                            mergeExcelBtn.style.display = 'block';
                        }
                        
                        const downloadCurrentZipBtn = document.getElementById('downloadCurrentZipBtn');
                        if (downloadCurrentZipBtn) {
                            downloadCurrentZipBtn.style.display = 'block';
                        }
                    }
                }
                
                let message = `分析完成！共掃描 ${data.total_files} 個檔案，找到 ${data.anr_subject_count} 個包含 ANR 的檔案，找到 ${data.files_with_cmdline - data.anr_subject_count} 個包含 Tombstone 的檔案`;
                message += `<br>分析耗時: ${data.analysis_time} 秒`;
                if (data.used_grep) {
                    message += '<span class="grep-badge">使用 grep 加速</span>';
                } else {
                    message += '<span class="grep-badge no-grep-badge">未使用 grep</span>';
                }
                
                // 新增 vp_analyze 狀態訊息
                if (data.vp_analyze_success) {
                    message += '<br><span style="color: #28a745;">✓ 詳細分析報告已生成</span>';
                } else if (data.vp_analyze_error) {
                    message += `<br><span style="color: #dc3545;">✗ 詳細分析失敗: ${data.vp_analyze_error}</span>`;
                }
                
                showMessage(message, 'success');
                
            } catch (error) {
                showMessage('錯誤: ' + error.message, 'error');
                window.hasCurrentAnalysis = false;
            } finally {
                document.getElementById('loading').style.display = 'none';
                document.getElementById('analyzeBtn').disabled = false;
            }
        }

        // 修改 CSS，確保全局按鈕預設隱藏
        const globalToggleStyle = `
        <style>
        .global-toggle-btn {
            position: fixed !important;
            width: 50px;
            height: 50px;
            border-radius: 50%;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            font-size: 24px;
            cursor: pointer;
            transition: all 0.3s;
            z-index: 9999 !important;
            right: 30px !important;
            left: auto !important;
            bottom: 150px !important;
            background: #667eea;
            color: white;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
            text-align: center;
            line-height: 1;
            opacity: 0;
            visibility: hidden;
        }

        .global-toggle-btn.show {
            opacity: 1;
            visibility: visible;
        }

        .global-toggle-btn:hover {
            background: #5a67d8;
            transform: translateY(-5px) !important;
            box-shadow: 0 6px 16px rgba(102, 126, 234, 0.6);
        }
        </style>`;

        // 在 DOMContentLoaded 時注入樣式
        document.addEventListener('DOMContentLoaded', function() {
            // 注入全局按鈕樣式
            const styleElement = document.createElement('div');
            styleElement.innerHTML = globalToggleStyle;
            const style = styleElement.querySelector('style');
            if (style) {
                document.head.appendChild(style);
            }
            
            // 確保全局按鈕初始狀態是隱藏的
            const globalToggleBtn = document.getElementById('globalToggleBtn');
            if (globalToggleBtn) {
                globalToggleBtn.classList.remove('show');
            }
        });
        
        

        function resetFiltersAndPagination() {
            summaryPage = 1;
            logsPage = 1;
            filesPage = 1;
            processSummaryPage = 1;
            filteredSummary = [...allSummary];
            filteredLogs = [...allLogs];
            filteredFiles = [...allFileStats];
            filteredProcessSummary = [...allProcessSummary];
            document.getElementById('summarySearchInput').value = '';
            document.getElementById('logsSearchInput').value = '';
            document.getElementById('filesSearchInput').value = '';
            document.getElementById('processSummarySearchInput').value = '';
            // 重置 regex toggles
            document.getElementById('summaryRegexToggle').checked = false;
            document.getElementById('logsRegexToggle').checked = false;
            document.getElementById('filesRegexToggle').checked = false;
            document.getElementById('processSummaryRegexToggle').checked = false;
        }

        // 加入進度條顯示函數
        function createWaitingProgressBar() {
            const progressHtml = `
                <div id="waitingProgress" style="
                    margin-top: 10px;
                    background: #f0f0f0;
                    border-radius: 8px;
                    overflow: hidden;
                    height: 20px;
                    position: relative;
                ">
                    <div id="waitingProgressBar" style="
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        height: 100%;
                        width: 0%;
                        transition: width 1s linear;
                        position: relative;
                    ">
                        <span id="waitingProgressText" style="
                            position: absolute;
                            right: 10px;
                            top: 50%;
                            transform: translateY(-50%);
                            color: white;
                            font-size: 12px;
                            font-weight: bold;
                        "></span>
                    </div>
                </div>
            `;
            
            return progressHtml;
        }

        // 修改 cancelWaiting 函數
        function cancelWaiting() {
            waitingForUnlock = false;
            document.getElementById('analyzeBtn').disabled = false;
            document.getElementById('loading').style.display = 'none';
            // 顯示取消訊息
            showMessage('已取消等待', 'info');
        }

        function updateResults(data) {

            // 新增：檢查是否有資料
            if (!data || data.files_with_cmdline === 0) {
                console.log('No data to display');
                document.getElementById('results').style.display = 'none';
                
                // 確保所有按鈕都隱藏
                const allButtons = [
                    'exportExcelBtn',
                    'exportExcelReportBtn',
                    'mergeExcelBtn',
                    'exportHtmlBtn',
                    'downloadCurrentZipBtn'
                ];
                
                allButtons.forEach(id => {
                    const btn = document.getElementById(id);
                    if (btn) btn.style.display = 'none';
                });
                
                return;
            }
            
            // === 確保 basePath 被設置 ===
            if (!window.basePath && data.path) {
                window.basePath = data.path;
                console.log('在 updateResults 中設置 basePath:', window.basePath);
            }
            
            // 生成程序統計（不分類型）
            const processOnlyData = {};
            data.statistics.type_process_summary.forEach(item => {
                if (!processOnlyData[item.process]) {
                    processOnlyData[item.process] = {
                        count: 0,
                        problem_sets: new Set()
                    };
                }
                processOnlyData[item.process].count += item.count;
                
                // 收集問題 sets
                if (item.problem_sets && Array.isArray(item.problem_sets)) {
                    item.problem_sets.forEach(set => {
                        processOnlyData[item.process].problem_sets.add(set);
                    });
                }
            });
            
            // 轉換為陣列格式
            allProcessSummary = Object.entries(processOnlyData)
                .map(([process, data]) => ({ 
                    process, 
                    count: data.count,
                    problem_sets: Array.from(data.problem_sets).sort()
                }))
                .sort((a, b) => b.count - a.count);
            
            filteredProcessSummary = [...allProcessSummary];

            // Update summary statistics
            const uniqueProcesses = data.statistics.total_unique_processes || 0;
            const summaryHtml = `
                <div class="stat-card">
                    <h3>${data.total_files}</h3>
                    <p>總掃描檔案數</p>
                </div>
                <div class="stat-card highlight">
                    <h3>${data.anr_subject_count || 0}</h3>
                    <p>ANR (Grep Subject)</p>
                </div>
                <div class="stat-card highlight">
                    <h3>${data.files_with_cmdline - data.anr_subject_count || 0}</h3>
                    <p>Tombstone (Grep Cmdline)</p>
                </div>              
                <div class="stat-card">
                    <h3>${uniqueProcesses}</h3>
                    <p>不同的程序</p>
                </div>
                <div class="stat-card">
                    <h3>${data.anr_folders + data.tombstone_folders}</h3>
                    <p>資料夾總數</p>
                </div>
            `;
            document.getElementById('statsSummary').innerHTML = summaryHtml;
            
            // Update charts
            updateProcessChart(data.statistics.type_process_summary);
            updateTypeChart(data.statistics.by_type);
            updateDailyChart(data.statistics.by_date, data.statistics.by_date_type);
            updateHourlyChart(data.statistics.by_hour, data.statistics.by_hour_type);
            
            // Update tables with pagination
            updateSummaryTable();
            updateFilesTable();
            updateLogsTable();
            
            // Show results
            document.getElementById('results').style.display = 'block';

            // 修正：強制圖表重新調整大小
            setTimeout(() => {
                // 確保所有圖表都正確顯示
                Object.values(charts).forEach(chart => {
                    if (chart) {
                        chart.resize();
                        chart.update();
                    }
                });
                
                // 確保圖表區域可見
                const chartSection = document.getElementById('charts-section-container');
                if (chartSection && !chartSection.classList.contains('collapsed')) {
                    // 觸發窗口調整事件
                    window.dispatchEvent(new Event('resize'));
                }
            }, 100);
                        
            // Setup search handlers
            setupSearchHandlers();

            // 更新程序統計表格
            updateProcessSummaryTable();
            
            // 只有在有結果時才顯示全局按鈕
            if (data.files_with_cmdline > 0) {
                showGlobalToggleButton();
            }           
        }

        // 顯示全局展開/收合按鈕
        function showGlobalToggleButton() {
            const globalToggleBtn = document.getElementById('globalToggleBtn');
            if (globalToggleBtn) {
                globalToggleBtn.classList.add('show');
            }
        }

        // 更新程序統計表格
        function updateProcessSummaryTable() {
            const tbody = document.getElementById('processSummaryTableBody');
            tbody.innerHTML = '';
            
            // 更新排序指示器
            document.querySelectorAll('#process-summary-section .sort-indicator').forEach(span => {
                const col = span.dataset.column;
                if (col === processSummarySort.column) {
                    span.textContent = processSummarySort.order === 'asc' ? '▲' : '▼';
                } else {
                    span.textContent = '';
                }
            });
            
            const totalPages = Math.max(1, Math.ceil(filteredProcessSummary.length / itemsPerPage));
            const startIndex = (processSummaryPage - 1) * itemsPerPage;
            const endIndex = Math.min(startIndex + itemsPerPage, filteredProcessSummary.length);
            const pageData = filteredProcessSummary.slice(startIndex, endIndex);
            
            if (pageData.length === 0) {
                const row = tbody.insertRow();
                row.innerHTML = `<td colspan="4" style="text-align: center; padding: 20px; color: #666;">
                    ${filteredProcessSummary.length === 0 && document.getElementById('processSummarySearchInput').value ? 
                      '沒有找到符合搜尋條件的資料' : '沒有資料'}
                </td>`;
            } else {
                const searchTerm = document.getElementById('processSummarySearchInput').value;
                const useRegex = document.getElementById('processSummaryRegexToggle').checked;
                
                pageData.forEach((item, index) => {
                    const row = tbody.insertRow();
                    const globalIndex = startIndex + index + 1;

                    // 處理問題 sets 顯示
                    const problemSetsHtml = item.problem_sets && item.problem_sets.length > 0 ? 
                        item.problem_sets.join(', ') : '-';

                    row.innerHTML = `
                        <td class="rank-number" style="text-align: center;">${globalIndex}</td>
                        <td class="process-name">${highlightText(item.process, searchTerm, useRegex)}</td>
                        <td style="color: #667eea; font-weight: 600;">${highlightText(problemSetsHtml, searchTerm, useRegex)}</td>
                        <td style="text-align: center; font-weight: bold; color: #e53e3e;">${item.count}</td>
                    `;
                });
            }
            
            // 更新分頁資訊
            document.getElementById('processSummaryPageInfo').textContent = 
                `第 ${processSummaryPage} 頁 / 共 ${totalPages} 頁 (總計 ${filteredProcessSummary.length} 筆)`;
            
            updatePaginationButtons('processSummary', processSummaryPage, totalPages);
        }
        
        // 排序函數
        function sortProcessSummaryTable(column) {
            if (processSummarySort.column === column) {
                processSummarySort.order = processSummarySort.order === 'asc' ? 'desc' : 'asc';
            } else {
                processSummarySort.column = column;
                processSummarySort.order = column === 'count' ? 'desc' : 'asc';
            }
            
            filteredProcessSummary.sort((a, b) => {
                let aVal, bVal;
                
                switch (column) {
                    case 'rank':
                        aVal = allProcessSummary.indexOf(a);
                        bVal = allProcessSummary.indexOf(b);
                        break;
                    case 'process':
                        aVal = a.process;
                        bVal = b.process;
                        break;
                    case 'count':
                        aVal = a.count;
                        bVal = b.count;
                        break;
                    case 'set':
                        aVal = a.problem_sets ? a.problem_sets.join(', ') : '';
                        bVal = b.problem_sets ? b.problem_sets.join(', ') : '';
                        break;                   
                }
                
                if (typeof aVal === 'string') {
                    return processSummarySort.order === 'asc' ? 
                        aVal.localeCompare(bVal) : 
                        bVal.localeCompare(aVal);
                } else {
                    return processSummarySort.order === 'asc' ? 
                        aVal - bVal : 
                        bVal - aVal;
                }
            });
            
            processSummaryPage = 1;
            updateProcessSummaryTable();
        }

        // 分頁函數
        function changeProcessSummaryPage(direction) {
            const totalPages = Math.ceil(filteredProcessSummary.length / itemsPerPage);
            
            if (direction === 'first') {
                processSummaryPage = 1;
            } else if (direction === 'last') {
                processSummaryPage = totalPages;
            } else {
                processSummaryPage = Math.max(1, Math.min(processSummaryPage + direction, totalPages));
            }
            
            updateProcessSummaryTable();
        }
        
        function setupSearchHandlers() {
            // Summary search with regex support
            const summarySearchHandler = (e) => {
                const searchTerm = e.target.value;
                const countElement = document.getElementById('summarySearchCount');
                const useRegex = document.getElementById('summaryRegexToggle').checked;
                
                if (searchTerm === '') {
                    // Reset to show all data
                    filteredSummary = [...allSummary];
                    countElement.style.display = 'none';
                } else {
                    try {
                        let searchFunction;
                        if (useRegex) {
                            const regex = new RegExp(searchTerm, 'i');
                            searchFunction = item => 
                                regex.test(item.type) || regex.test(item.process) ||
                                (item.problem_sets && (
                                    regex.test(item.problem_sets.join(', ')) ||
                                    item.problem_sets.some(ps => regex.test(ps))
                                ));
                        } else {
                            const lowerSearchTerm = searchTerm.toLowerCase();
                            searchFunction = item => 
                                item.type.toLowerCase().includes(lowerSearchTerm) ||
                                item.process.toLowerCase().includes(lowerSearchTerm) ||
                                (item.problem_sets && (
                                    item.problem_sets.join(', ').toLowerCase().includes(lowerSearchTerm) ||
                                    item.problem_sets.some(ps => ps.toLowerCase().includes(lowerSearchTerm))
                                ));
                        }
                        
                        filteredSummary = allSummary.filter(searchFunction);
                        
                        // 計算總次數
                        const totalCount = filteredSummary.reduce((sum, item) => sum + item.count, 0);
                        
                        // 顯示搜尋結果數量和總次數
                        countElement.innerHTML = `找到 <span style="color: #e53e3e;">${filteredSummary.length}</span> 筆項目，總次數: <span style="color: #e53e3e;">${totalCount}</span>`;
                        countElement.style.display = 'inline';
                    } catch (error) {
                        // Invalid regex
                        countElement.innerHTML = `<span style="color: #e53e3e;">無效的正則表達式</span>`;
                        countElement.style.display = 'inline';
                        filteredSummary = [];
                    }
                }
                summaryPage = 1; // Reset to first page
                updateSummaryTable();
            };
            
            document.getElementById('summarySearchInput').addEventListener('input', summarySearchHandler);
            document.getElementById('summaryRegexToggle').addEventListener('change', function() {
                // 強制觸發搜尋處理器
                summarySearchHandler({ target: document.getElementById('summarySearchInput') });
            });
            
            // Logs search with regex support
            const logsSearchHandler = (e) => {
                const searchTerm = e.target.value;
                const countElement = document.getElementById('logsSearchCount');
                const useRegex = document.getElementById('logsRegexToggle').checked;
                
                if (searchTerm === '') {
                    // Reset to show all data
                    filteredLogs = [...allLogs];
                    countElement.style.display = 'none';
                } else {
                    try {
                        let searchFunction;
                        if (useRegex) {
                            const regex = new RegExp(searchTerm, 'i');
                            searchFunction = log => 
                                (log.process && regex.test(log.process)) ||
                                (log.type && regex.test(log.type)) ||
                                (log.filename && regex.test(log.filename)) ||
                                (log.folder_path && regex.test(log.folder_path)) ||
                                (log.timestamp && regex.test(log.timestamp)) ||
                                (log.line_number && regex.test(String(log.line_number))) ||
                                (log.problem_set && regex.test(log.problem_set));
                        } else {
                            const lowerSearchTerm = searchTerm.toLowerCase();
                            searchFunction = log => 
                                (log.process && log.process.toLowerCase().includes(lowerSearchTerm)) ||
                                (log.type && log.type.toLowerCase().includes(lowerSearchTerm)) ||
                                (log.filename && log.filename.toLowerCase().includes(lowerSearchTerm)) ||
                                (log.folder_path && log.folder_path.toLowerCase().includes(lowerSearchTerm)) ||
                                (log.timestamp && log.timestamp.toLowerCase().includes(lowerSearchTerm)) ||
                                (log.line_number && String(log.line_number).includes(lowerSearchTerm)) ||
                                (log.problem_set && log.problem_set.toLowerCase().includes(lowerSearchTerm));
                        }
                        
                        filteredLogs = allLogs.filter(searchFunction);
                        
                        // 對於 logs，每一筆就是一個記錄，所以總次數等於筆數
                        countElement.innerHTML = `找到 <span style="color: #e53e3e;">${filteredLogs.length}</span> 筆記錄`;
                        countElement.style.display = 'inline';
                    } catch (error) {
                        // Invalid regex
                        countElement.innerHTML = `<span style="color: #e53e3e;">無效的正則表達式</span>`;
                        countElement.style.display = 'inline';
                        filteredLogs = [];
                    }
                }
                logsPage = 1; // Reset to first page
                updateLogsTable();
            };
            
            document.getElementById('logsSearchInput').addEventListener('input', logsSearchHandler);
            document.getElementById('logsRegexToggle').addEventListener('change', function() {
                // 強制觸發搜尋處理器
                logsSearchHandler({ target: document.getElementById('logsSearchInput') });
            });
            
            // Files search with regex support
            const filesSearchHandler = (e) => {
                const searchTerm = e.target.value;
                const countElement = document.getElementById('filesSearchCount');
                const useRegex = document.getElementById('filesRegexToggle').checked;
                
                if (searchTerm === '') {
                    // Reset to show all data
                    filteredFiles = [...allFileStats];
                    countElement.style.display = 'none';
                } else {
                    try {
                        let searchFunction;
                        if (useRegex) {
                            const regex = new RegExp(searchTerm, 'i');
                            searchFunction = file => 
                                regex.test(file.filename) ||
                                regex.test(file.type) ||
                                regex.test(file.folder_path || '') ||
                                regex.test(file.timestamp || '') ||
                                regex.test(String(file.count)) ||
                                (file.problem_set && regex.test(file.problem_set)) ||
                                file.processes.some(proc => regex.test(proc));
                        } else {
                            const lowerSearchTerm = searchTerm.toLowerCase();
                            searchFunction = file => 
                                file.filename.toLowerCase().includes(lowerSearchTerm) ||
                                file.type.toLowerCase().includes(lowerSearchTerm) ||
                                (file.folder_path && file.folder_path.toLowerCase().includes(lowerSearchTerm)) ||
                                (file.timestamp && file.timestamp.toLowerCase().includes(lowerSearchTerm)) ||
                                String(file.count).includes(lowerSearchTerm) ||
                                (file.problem_set && file.problem_set.toLowerCase().includes(lowerSearchTerm)) ||
                                file.processes.some(proc => proc.toLowerCase().includes(lowerSearchTerm));
                        }
                        
                        filteredFiles = allFileStats.filter(searchFunction);
                        
                        // 計算總次數
                        let totalCount = 0;
                        filteredFiles.forEach(file => {
                            // 解析每個程序的次數
                            file.processes.forEach(procStr => {
                                // procStr 格式: "程序名稱 (次數)"
                                const match = procStr.match(/^(.+)\s+\((\d+)\)$/);
                                if (match) {
                                    const processName = match[1];
                                    const count = parseInt(match[2]);
                                    // 檢查程序名稱是否符合搜尋條件
                                    if (useRegex) {
                                        const regex = new RegExp(searchTerm, 'i');
                                        if (regex.test(processName)) {
                                            totalCount += count;
                                        }
                                    } else {
                                        if (processName.toLowerCase().includes(searchTerm.toLowerCase())) {
                                            totalCount += count;
                                        }
                                    }
                                }
                            });
                        });
                        
                        // 顯示搜尋結果數量和總次數
                        countElement.innerHTML = `找到 <span style="color: #e53e3e;">${filteredFiles.length}</span> 筆檔案，總次數: <span style="color: #e53e3e;">${totalCount}</span>`;
                        countElement.style.display = 'inline';
                    } catch (error) {
                        // Invalid regex
                        countElement.innerHTML = `<span style="color: #e53e3e;">無效的正則表達式</span>`;
                        countElement.style.display = 'inline';
                        filteredFiles = [];
                    }
                }
                filesPage = 1; // Reset to first page
                updateFilesTable();
            };
            
            document.getElementById('filesSearchInput').addEventListener('input', filesSearchHandler);
            document.getElementById('filesRegexToggle').addEventListener('change', function() {
                // 強制觸發搜尋處理器
                filesSearchHandler({ target: document.getElementById('filesSearchInput') });
            });
            
            // Process Summary search with regex support
            const processSummarySearchHandler = (e) => {
                const searchTerm = e.target.value;
                const countElement = document.getElementById('processSummarySearchCount');
                const useRegex = document.getElementById('processSummaryRegexToggle').checked;
                
                if (searchTerm === '') {
                    filteredProcessSummary = [...allProcessSummary];
                    countElement.style.display = 'none';
                } else {
                    try {
                        let searchFunction;
                        if (useRegex) {
                            const regex = new RegExp(searchTerm, 'i');
                            searchFunction = item => 
                                regex.test(item.process) ||
                                regex.test(String(item.count)) ||
                                (item.problem_sets && (
                                    regex.test(item.problem_sets.join(', ')) ||
                                    item.problem_sets.some(ps => regex.test(ps))
                                ));
                        } else {
                            const lowerSearchTerm = searchTerm.toLowerCase();
                            searchFunction = item => 
                                item.process.toLowerCase().includes(lowerSearchTerm) ||
                                String(item.count).includes(lowerSearchTerm) ||
                                (item.problem_sets && (
                                    item.problem_sets.join(', ').toLowerCase().includes(lowerSearchTerm) ||
                                    item.problem_sets.some(ps => ps.toLowerCase().includes(lowerSearchTerm))
                                ));
                        }
                        
                        filteredProcessSummary = allProcessSummary.filter(searchFunction);
                        
                        const totalCount = filteredProcessSummary.reduce((sum, item) => sum + item.count, 0);
                        
                        countElement.innerHTML = `找到 <span style="color: #e53e3e;">${filteredProcessSummary.length}</span> 筆項目，總次數: <span style="color: #e53e3e;">${totalCount}</span>`;
                        countElement.style.display = 'inline';
                    } catch (error) {
                        countElement.innerHTML = `<span style="color: #e53e3e;">無效的正則表達式</span>`;
                        countElement.style.display = 'inline';
                        filteredProcessSummary = [];
                    }
                }
                processSummaryPage = 1;
                updateProcessSummaryTable();
            };

            document.getElementById('processSummarySearchInput').addEventListener('input', processSummarySearchHandler);
            document.getElementById('processSummaryRegexToggle').addEventListener('change', function() {
                // 強制觸發搜尋處理器
                processSummarySearchHandler({ target: document.getElementById('processSummarySearchInput') });
            });
        }

        function sortSummaryTable(column) {
            // 切換排序順序
            if (summarySort.column === column) {
                summarySort.order = summarySort.order === 'asc' ? 'desc' : 'asc';
            } else {
                summarySort.column = column;
                summarySort.order = column === 'count' ? 'desc' : 'asc'; // 次數預設降序，其他升序
            }
            
            // 排序資料
            filteredSummary.sort((a, b) => {
                let aVal, bVal;
                
                switch (column) {
                    case 'rank':
                        // 按原始順序
                        aVal = allSummary.indexOf(a);
                        bVal = allSummary.indexOf(b);
                        break;
                    case 'type':
                        aVal = a.type;
                        bVal = b.type;
                        break;
                    case 'process':
                        aVal = a.process;
                        bVal = b.process;
                        break;
                    case 'count':
                        aVal = a.count;
                        bVal = b.count;
                        break;
                    case 'set':
                        aVal = a.problem_sets ? a.problem_sets.join(', ') : '';
                        bVal = b.problem_sets ? b.problem_sets.join(', ') : '';
                        break;                        
                }
                
                if (typeof aVal === 'string') {
                    return summarySort.order === 'asc' ? 
                        aVal.localeCompare(bVal) : 
                        bVal.localeCompare(aVal);
                } else {
                    return summarySort.order === 'asc' ? 
                        aVal - bVal : 
                        bVal - aVal;
                }
            });
            
            // 重置到第一頁
            summaryPage = 1;
            updateSummaryTable();
        }

        function sortFilesTable(column) {
            if (filesSort.column === column) {
                filesSort.order = filesSort.order === 'asc' ? 'desc' : 'asc';
            } else {
                filesSort.column = column;
                filesSort.order = column === 'count' ? 'desc' : 'asc';
            }
            
            filteredFiles.sort((a, b) => {
                let aVal, bVal;
                
                switch (column) {
                    case 'index':
                        aVal = allFileStats.indexOf(a);
                        bVal = allFileStats.indexOf(b);
                        break;
                    case 'type':
                        aVal = a.type || '';
                        bVal = b.type || '';
                        break;
                    case 'processes':
                        // 使用第一個程序名稱排序
                        aVal = a.processes[0] || '';
                        bVal = b.processes[0] || '';
                        break;
                    case 'folder_path':
                        aVal = a.folder_path || '';
                        bVal = b.folder_path || '';
                        break;
                    case 'filename':
                        aVal = a.filename || '';
                        bVal = b.filename || '';
                        break;
                    case 'count':
                        aVal = a.count || 0;
                        bVal = b.count || 0;
                        break;
                    case 'timestamp':
                        aVal = a.timestamp || '';
                        bVal = b.timestamp || '';
                        break;
                    case 'set':
                        aVal = a.problem_set || '';
                        bVal = b.problem_set || '';
                        break;                        
                }
                
                if (typeof aVal === 'string') {
                    return filesSort.order === 'asc' ? 
                        aVal.localeCompare(bVal) : 
                        bVal.localeCompare(aVal);
                } else {
                    return filesSort.order === 'asc' ? 
                        aVal - bVal : 
                        bVal - aVal;
                }
            });
            
            filesPage = 1;
            updateFilesTable();
        }
        
        function sortLogsTable(column) {
            if (logsSort.column === column) {
                logsSort.order = logsSort.order === 'asc' ? 'desc' : 'asc';
            } else {
                logsSort.column = column;
                logsSort.order = 'asc';
            }
            
            filteredLogs.sort((a, b) => {
                let aVal, bVal;
                
                switch (column) {
                    case 'index':
                        aVal = allLogs.indexOf(a);
                        bVal = allLogs.indexOf(b);
                        break;
                    case 'type':
                        aVal = a.type || '';
                        bVal = b.type || '';
                        break;
                    case 'process':
                        aVal = a.process || '';
                        bVal = b.process || '';
                        break;
                    case 'line_number':
                        aVal = a.line_number || 0;
                        bVal = b.line_number || 0;
                        break;
                    case 'folder_path':
                        aVal = a.folder_path || '';
                        bVal = b.folder_path || '';
                        break;
                    case 'filename':
                        aVal = a.filename || '';
                        bVal = b.filename || '';
                        break;
                    case 'timestamp':
                        aVal = a.timestamp || '';
                        bVal = b.timestamp || '';
                        break;
                    case 'set':
                        aVal = a.problem_set || '';
                        bVal = b.problem_set || '';
                        break;                        
                }
                
                if (typeof aVal === 'string') {
                    return logsSort.order === 'asc' ? 
                        aVal.localeCompare(bVal) : 
                        bVal.localeCompare(aVal);
                } else {
                    return logsSort.order === 'asc' ? 
                        aVal - bVal : 
                        bVal - aVal;
                }
            });
            
            logsPage = 1;
            updateLogsTable();
        }
        
        function updateSummaryTable() {
            const tbody = document.getElementById('summaryTableBody');
            tbody.innerHTML = '';
            
            // 更新排序指示器
            document.querySelectorAll('#summary-section .sort-indicator').forEach(span => {
                const col = span.dataset.column;
                if (col === summarySort.column) {
                    span.textContent = summarySort.order === 'asc' ? '▲' : '▼';
                } else {
                    span.textContent = '';
                }
            });
            
            const totalPages = Math.max(1, Math.ceil(filteredSummary.length / itemsPerPage));
            const startIndex = (summaryPage - 1) * itemsPerPage;
            const endIndex = Math.min(startIndex + itemsPerPage, filteredSummary.length);
            const pageData = filteredSummary.slice(startIndex, endIndex);
            
            if (pageData.length === 0) {
                // Show no data message
                const row = tbody.insertRow();
                row.innerHTML = `<td colspan="5" style="text-align: center; padding: 20px; color: #666;">
                    ${filteredSummary.length === 0 && document.getElementById('summarySearchInput').value ? 
                      '沒有找到符合搜尋條件的資料' : '沒有資料'}
                </td>`;
            } else {
                pageData.forEach((item, index) => {
                    const row = tbody.insertRow();
                    const globalIndex = startIndex + index + 1;
                    const searchTerm = document.getElementById('summarySearchInput').value;
                    const useRegex = document.getElementById('summaryRegexToggle').checked;

                    // 處理問題 sets 顯示
                    const problemSetsHtml = item.problem_sets && item.problem_sets.length > 0 ? 
                        item.problem_sets.join(', ') : '-';

                    row.innerHTML = `
                        <td class="rank-number" style="text-align: center;">${globalIndex}</td>
                        <td>${highlightText(item.type, searchTerm, useRegex)}</td>
                        <td class="process-name">${highlightText(item.process, searchTerm, useRegex)}</td>
                        <td style="color: #667eea; font-weight: 600;">${highlightText(problemSetsHtml, searchTerm, useRegex)}</td>
                        <td style="text-align: center; font-weight: bold; color: #e53e3e;">${item.count}</td>
                    `;
                });
            }
            
            // Update pagination info
            document.getElementById('summaryPageInfo').textContent = 
                `第 ${summaryPage} 頁 / 共 ${totalPages} 頁 (總計 ${filteredSummary.length} 筆)`;
            
            // Update pagination buttons state
            updatePaginationButtons('summary', summaryPage, totalPages);
        }

        function updateLogsTable() {
            const tbody = document.getElementById('logsTableBody');
            tbody.innerHTML = '';

            // 更新排序指示器
            document.querySelectorAll('#logs-section .sort-indicator').forEach(span => {
                const col = span.dataset.column;
                if (col === logsSort.column) {
                    span.textContent = logsSort.order === 'asc' ? '▲' : '▼';
                } else {
                    span.textContent = '';
                }
            });
    
            const totalPages = Math.max(1, Math.ceil(filteredLogs.length / itemsPerPage));
            const startIndex = (logsPage - 1) * itemsPerPage;
            const endIndex = Math.min(startIndex + itemsPerPage, filteredLogs.length);
            const pageData = filteredLogs.slice(startIndex, endIndex);
            
            if (pageData.length === 0) {
                // Show no data message
                const row = tbody.insertRow();
                row.innerHTML = `<td colspan="8" style="text-align: center; padding: 20px; color: #666;">
                    ${filteredLogs.length === 0 && document.getElementById('logsSearchInput').value ? 
                      '沒有找到符合搜尋條件的資料' : '沒有資料'}
                </td>`;
            } else {
                const searchTerm = document.getElementById('logsSearchInput').value;
                const useRegex = document.getElementById('logsRegexToggle').checked;
                
                pageData.forEach((log, index) => {
                    const row = tbody.insertRow();
                    const globalIndex = startIndex + index + 1;
                    
                    // Create clickable file link
                    const fileLink = `/view-file?path=${encodeURIComponent(log.file)}`;
                    
                    // === 新增：建立分析報告連結 ===
                    let analyzeReportLink = '';
                    if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {
                        // 使用全域的 basePath 而不是從 input 取值
                        const basePath = window.basePath || document.getElementById('pathInput').value;
                        const filePath = log.file || '';
                        
                        // 添加除錯訊息
                        console.log('=== Logs Table Analyze Link Debug ===');
                        console.log('Using basePath:', basePath);
                        console.log('Log file path:', filePath);
                        console.log('vpAnalyzeOutputPath:', window.vpAnalyzeOutputPath);
                        
                        // 從檔案路徑中提取相對路徑
                        if (filePath.startsWith(basePath)) {
                            // 取得從基礎路徑之後的相對路徑
                            let relativePath = filePath.substring(basePath.length);
                            // 移除開頭的斜線
                            if (relativePath.startsWith('/')) {
                                relativePath = relativePath.substring(1);
                            }
                            
                            console.log('Relative path:', relativePath);
                            
                            // 建立分析報告路徑
                            const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                            console.log('Analyzed file path:', analyzedFilePath);
                            
                            analyzeReportLink = ` <a href="/view-file?path=${encodeURIComponent(analyzedFilePath)}" target="_blank" class="file-link" style="color: #28a745; font-size: 0.9em;">(分析報告)</a>`;
                        } else {
                            console.log('File path does not start with basePath');
                            console.log('Expected basePath:', basePath);
                            console.log('Actual file path:', filePath);
                        }
                    }
                    
                    // Process display
                    const processDisplay = log.process || '-';
                    
                    // Line number display
                    const lineNumber = log.line_number || '-';
                    
                    // Folder path display
                    const folderPath = log.folder_path || '-';
                    
                    const problemSet = log.problem_set || '-';

                    row.innerHTML = `
                        <td style="text-align: center; color: #666;">${globalIndex}</td>
                        <td>${highlightText(log.type, searchTerm, useRegex)}</td>
                        <td class="process-name">${highlightText(processDisplay, searchTerm, useRegex)}</td>
                        <td style="text-align: center; font-weight: bold; color: #667eea;">${lineNumber}</td>
                        <td style="text-align: center; color: #667eea; font-weight: 600;">${highlightText(problemSet, searchTerm, useRegex)}</td>
                        <td class="tooltip-container">
                            <span class="folder-path-cell">${highlightText(folderPath, searchTerm, useRegex)}</span>
                            <span class="tooltip-text">完整路徑：<br>${escapeHtml(log.file || '')}</span>
                        </td>
                        <td><a href="${fileLink}" target="_blank" class="file-link">${highlightText(log.filename, searchTerm, useRegex)}</a>${analyzeReportLink}</td>
                        <td>${log.timestamp || '-'}</td>
                    `;
                });
            }
            
            // Update pagination info
            document.getElementById('logsPageInfo').textContent = 
                `第 ${logsPage} 頁 / 共 ${totalPages} 頁 (總計 ${filteredLogs.length} 筆)`;
            
            // Update pagination buttons state
            updatePaginationButtons('logs', logsPage, totalPages);
        }
        
        function changeSummaryPage(direction) {
            const totalPages = Math.ceil(filteredSummary.length / itemsPerPage);
            
            if (direction === 'first') {
                summaryPage = 1;
            } else if (direction === 'last') {
                summaryPage = totalPages;
            } else {
                summaryPage = Math.max(1, Math.min(summaryPage + direction, totalPages));
            }
            
            updateSummaryTable();
        }
        
        function updatePaginationButtons(type, currentPage, totalPages) {
            const firstBtn = document.getElementById(`${type}FirstBtn`);
            const prevBtn = document.getElementById(`${type}PrevBtn`);
            const nextBtn = document.getElementById(`${type}NextBtn`);
            const lastBtn = document.getElementById(`${type}LastBtn`);
            
            // Disable/enable buttons based on current page
            firstBtn.disabled = currentPage === 1;
            prevBtn.disabled = currentPage === 1;
            nextBtn.disabled = currentPage === totalPages || totalPages === 0;
            lastBtn.disabled = currentPage === totalPages || totalPages === 0;
        }
        
        function updateFilesTable() {
            const tbody = document.getElementById('filesTableBody');
            tbody.innerHTML = '';

            // 更新排序指示器
            document.querySelectorAll('#files-section .sort-indicator').forEach(span => {
                const col = span.dataset.column;
                if (col === filesSort.column) {
                    span.textContent = filesSort.order === 'asc' ? '▲' : '▼';
                } else {
                    span.textContent = '';
                }
            });
    
            const totalPages = Math.max(1, Math.ceil(filteredFiles.length / itemsPerPage));
            const startIndex = (filesPage - 1) * itemsPerPage;
            const endIndex = Math.min(startIndex + itemsPerPage, filteredFiles.length);
            const pageData = filteredFiles.slice(startIndex, endIndex);
            
            if (pageData.length === 0) {
                // Show no data message
                const row = tbody.insertRow();
                row.innerHTML = `<td colspan="8" style="text-align: center; padding: 20px; color: #666;">
                    ${filteredFiles.length === 0 && document.getElementById('filesSearchInput').value ? 
                      '沒有找到符合搜尋條件的資料' : '沒有資料'}
                </td>`;
            } else {
                const searchTerm = document.getElementById('filesSearchInput').value;
                const useRegex = document.getElementById('filesRegexToggle').checked;
                
                pageData.forEach((file, index) => {
                    const row = tbody.insertRow();
                    const globalIndex = startIndex + index + 1;
                    
                    // 直接使用 file.filepath
                    const fileLink = `/view-file?path=${encodeURIComponent(file.filepath || '')}`;
                    const folderPath = file.folder_path || '-';
                    const problemSet = file.problem_set || '-';

                    // === 新增：建立分析報告連結 ===
                    let analyzeReportLink = '';
                    if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {
                        // 使用全域的 basePath 而不是從 input 取值
                        const basePath = window.basePath || document.getElementById('pathInput').value;
                        const filePath = file.filepath || '';
                        
                        console.log('=== Files Table Analyze Link Debug ===');
                        console.log('Using basePath:', basePath);
                        console.log('File path:', filePath);
                        console.log('vpAnalyzeOutputPath:', window.vpAnalyzeOutputPath);
                        
                        // 從檔案路徑中提取相對路徑
                        if (filePath.startsWith(basePath)) {
                            // 取得從基礎路徑之後的相對路徑
                            let relativePath = filePath.substring(basePath.length);
                            // 移除開頭的斜線
                            if (relativePath.startsWith('/')) {
                                relativePath = relativePath.substring(1);
                            }
                            
                            console.log('Relative path:', relativePath);
                            
                            // 建立分析報告路徑
                            const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                            console.log('Analyzed file path:', analyzedFilePath);
                            
                            analyzeReportLink = ` <a href="/view-file?path=${encodeURIComponent(analyzedFilePath)}" target="_blank" class="file-link" style="color: #28a745; font-size: 0.9em;">(分析報告)</a>`;
                        } else {
                            console.log('File path does not start with basePath');
                        }
                    }
                    
                    // 處理 processes 高亮
                    const processesHtml = file.processes.length > 0 ? 
                        file.processes.map(p => {
                            // 解析程序名稱和次數
                            const match = p.match(/^(.+)\s+\((\d+)\)$/);
                            if (match) {
                                const processName = match[1];
                                const count = match[2];
                                const highlightedName = highlightText(processName, searchTerm, useRegex);
                                return `${highlightedName} <span style="color: #e53e3e; font-weight: bold;">(${count})</span>`;
                            }
                            return highlightText(p, searchTerm, useRegex);
                        }).join(', ') : '-';
                    
                    const timestamp = file.timestamp || '-';
                    
                    row.innerHTML = `
                        <td style="text-align: center; color: #666;">${globalIndex}</td>
                        <td>${highlightText(file.type, searchTerm, useRegex)}</td>
                        <td class="process-name">${processesHtml}</td>
                        <td style="text-align: center; color: #667eea; font-weight: 600;">${highlightText(problemSet, searchTerm, useRegex)}</td>
                        <td class="tooltip-container">
                            <span class="folder-path-cell">${highlightText(folderPath, searchTerm, useRegex)}</span>
                            <span class="tooltip-text">完整路徑：<br>${escapeHtml(file.full_path || file.filepath || '')}</span>
                        </td>
                        <td><a href="${fileLink}" target="_blank" class="file-link">${highlightText(file.filename, searchTerm, useRegex)}</a>${analyzeReportLink}</td>
                        <td style="text-align: center; font-weight: bold; color: #e53e3e;">${file.count}</td>
                        <td>${timestamp}</td>
                    `;
                });
            }
            
            // Update pagination info
            document.getElementById('filesPageInfo').textContent = 
                `第 ${filesPage} 頁 / 共 ${totalPages} 頁 (總計 ${filteredFiles.length} 筆)`;
            
            // Update pagination buttons state
            updatePaginationButtons('files', filesPage, totalPages);
        }
        
        function changeLogsPage(direction) {
            const totalPages = Math.ceil(filteredLogs.length / itemsPerPage);
            
            if (direction === 'first') {
                logsPage = 1;
            } else if (direction === 'last') {
                logsPage = totalPages;
            } else {
                logsPage = Math.max(1, Math.min(logsPage + direction, totalPages));
            }
            
            updateLogsTable();
        }
        
        function changeFilesPage(direction) {
            const totalPages = Math.ceil(filteredFiles.length / itemsPerPage);
            
            if (direction === 'first') {
                filesPage = 1;
            } else if (direction === 'last') {
                filesPage = totalPages;
            } else {
                filesPage = Math.max(1, Math.min(filesPage + direction, totalPages));
            }
            
            updateFilesTable();
        }
        
        function updateProcessChart(typeSummaryData) {
            const ctx = document.getElementById('processChart').getContext('2d');
            
            // 從 type_process_summary 計算每個程序的總數
            const processCount = {};
            typeSummaryData.forEach(item => {
                if (!processCount[item.process]) {
                    processCount[item.process] = 0;
                }
                processCount[item.process] += item.count;
            });
            
            // 排序並取前10
            const sortedProcesses = Object.entries(processCount)
                .sort((a, b) => b[1] - a[1])
                .slice(0, 10);
            
            const labels = sortedProcesses.map(([process, _]) => process);
            const totals = sortedProcesses.map(([_, count]) => count);
            
            // 從 typeSummaryData 獲取每個程序的 ANR 和 Tombstone 數據
            const anrData = [];
            const tombstoneData = [];
            
            labels.forEach(process => {
                let anrCount = 0;
                let tombstoneCount = 0;
                
                typeSummaryData.forEach(item => {
                    if (item.process === process) {
                        if (item.type === 'ANR') {
                            anrCount = item.count;
                        } else if (item.type === 'Tombstone') {
                            tombstoneCount = item.count;
                        }
                    }
                });
                
                anrData.push(anrCount);
                tombstoneData.push(tombstoneCount);
            });
            
            if (charts.process) charts.process.destroy();
            
            charts.process = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: labels,
                    datasets: [
                        {
                            label: 'ANR',
                            data: anrData,
                            backgroundColor: 'rgba(102, 126, 234, 0.8)',
                            borderColor: 'rgba(102, 126, 234, 1)',
                            borderWidth: 1
                        },
                        {
                            label: 'Tombstone',
                            data: tombstoneData,
                            backgroundColor: 'rgba(234, 102, 102, 0.8)',
                            borderColor: 'rgba(234, 102, 102, 1)',
                            borderWidth: 1
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top'
                        },
                        tooltip: {
                            callbacks: {
                                afterLabel: function(context) {
                                    const total = anrData[context.dataIndex] + tombstoneData[context.dataIndex];
                                    return `總計: ${total}`;
                                }
                            }
                        }
                    },
                    scales: {
                        y: {
                            beginAtZero: true,
                            stacked: false,
                            ticks: {
                                stepSize: 1
                            }
                        },
                        x: {
                            ticks: {
                                autoSkip: false,
                                maxRotation: 45,
                                minRotation: 45
                            }
                        }
                    }
                }
            });
        }
        
        function updateTypeChart(typeData) {
            const ctx = document.getElementById('typeChart').getContext('2d');
            
            const labels = Object.keys(typeData);
            const values = Object.values(typeData);
            
            if (charts.type) charts.type.destroy();
            
            charts.type = new Chart(ctx, {
                type: 'doughnut',
                data: {
                    labels: labels,
                    datasets: [{
                        data: values,
                        backgroundColor: [
                            'rgba(102, 126, 234, 0.8)',
                            'rgba(234, 102, 102, 0.8)'
                        ],
                        borderColor: [
                            'rgba(102, 126, 234, 1)',
                            'rgba(234, 102, 102, 1)'
                        ],
                        borderWidth: 1
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            position: 'bottom'
                        }
                    }
                }
            });
        }
        
        function updateDailyChart(dailyData, dailyByType) {
            const ctx = document.getElementById('dailyChart').getContext('2d');
            
            const labels = Object.keys(dailyData);
            const anrData = labels.map(date => dailyByType.ANR[date] || 0);
            const tombstoneData = labels.map(date => dailyByType.Tombstone[date] || 0);
            
            if (charts.daily) charts.daily.destroy();
            
            charts.daily = new Chart(ctx, {
                type: 'line',
                data: {
                    labels: labels,
                    datasets: [
                        {
                            label: 'ANR',
                            data: anrData,
                            borderColor: 'rgba(102, 126, 234, 1)',
                            backgroundColor: 'rgba(102, 126, 234, 0.1)',
                            tension: 0.3,
                            fill: true
                        },
                        {
                            label: 'Tombstone',
                            data: tombstoneData,
                            borderColor: 'rgba(234, 102, 102, 1)',
                            backgroundColor: 'rgba(234, 102, 102, 0.1)',
                            tension: 0.3,
                            fill: true
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top'
                        }
                    },
                    scales: {
                        y: {
                            beginAtZero: true,
                            ticks: {
                                stepSize: 1
                            }
                        }
                    }
                }
            });
        }
        
        function updateHourlyChart(hourlyData, hourlyByType) {
            const ctx = document.getElementById('hourlyChart').getContext('2d');
            
            // Create all 24 hours
            const allHours = [];
            for (let i = 0; i < 24; i++) {
                allHours.push(`${i.toString().padStart(2, '0')}:00`);
            }
            
            const anrData = allHours.map(hour => hourlyByType.ANR[hour] || 0);
            const tombstoneData = allHours.map(hour => hourlyByType.Tombstone[hour] || 0);
            
            if (charts.hourly) charts.hourly.destroy();
            
            charts.hourly = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: allHours,
                    datasets: [
                        {
                            label: 'ANR',
                            data: anrData,
                            backgroundColor: 'rgba(102, 126, 234, 0.8)',
                            borderColor: 'rgba(102, 126, 234, 1)',
                            borderWidth: 1
                        },
                        {
                            label: 'Tombstone',
                            data: tombstoneData,
                            backgroundColor: 'rgba(234, 102, 102, 0.8)',
                            borderColor: 'rgba(234, 102, 102, 1)',
                            borderWidth: 1
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top'
                        }
                    },
                    scales: {
                        y: {
                            beginAtZero: true,
                            stacked: true,  // 使用堆疊
                            ticks: {
                                stepSize: 1
                            }
                        },
                        x: {
                            stacked: true   // 使用堆疊
                        }
                    }
                }
            });
        }
        
        async function exportResults(format) {
            if (!currentAnalysisId) {
                showMessage('請先執行分析', 'error');
                return;
            }
            
            if (format === 'html') {
                // 獲取當前服務器資訊並通過 URL 參數傳遞
                try {
                    const serverResponse = await fetch('/server-info');
                    const serverInfo = await serverResponse.json();
                    
                    // 使用 URL 參數傳遞 base_url
                    const encodedBaseUrl = encodeURIComponent(serverInfo.base_url);
                    window.location.href = `/export/${format}/${currentAnalysisId}?base_url=${encodedBaseUrl}`;
                } catch (error) {
                    // 如果無法獲取服務器資訊，使用原有方式
                    window.location.href = `/export/${format}/${currentAnalysisId}`;
                }
            } else {
                window.location.href = `/export/${format}/${currentAnalysisId}`;
            }
        }
        
        function showMessage(message, type) {
            const messageDiv = document.getElementById('message');
            messageDiv.className = type;
            messageDiv.innerHTML = message;
        }
        
        function clearMessage() {
            document.getElementById('message').innerHTML = '';
            document.getElementById('message').className = '';
        }
        
        function highlightText(text, searchTerm, useRegex = false) {
            if (!searchTerm || !text) return escapeHtml(text);
            
            const escapedText = escapeHtml(text);
            
            try {
                let regex;
                if (useRegex) {
                    regex = new RegExp(`(${searchTerm})`, 'gi');
                } else {
                    const escapedSearchTerm = escapeHtml(searchTerm).replace(/[.*+?^${}()|[\\]\\\\]/g, '\\\\$&');
                    regex = new RegExp(`(${escapedSearchTerm})`, 'gi');
                }
                
                return escapedText.replace(regex, '<span class="table-highlight">$1</span>');
            } catch (error) {
                // If regex is invalid, return text without highlighting
                return escapedText;
            }
        }

        function escapeHtml(text) {
            if (!text) return '';
            const div = document.createElement('div');
            text = String(text);
            div.textContent = text;
            return div.innerHTML;
        }
        
    </script>
    <script>
        // 合併 Excel 相關變數
        let mergeSelectedSuggestionIndex = -1;
        let mergeCurrentSuggestions = [];
        let mergeAutocompleteTimeout = null;
        let selectedMergeFiles = [];  // 改為陣列
        let selectedMergeFilePaths = [];  // 改為陣列

        // 打開合併對話框
        function openMergeDialog() {
            // 重置對話框狀態
            loadExcelMode = false;  // 確保不在載入模式
            
            const dialog = document.getElementById('mergeDialogOverlay');
            const dialogHeader = dialog.querySelector('.merge-dialog-header h3');
            const executeBtn = document.getElementById('mergeExecuteBtn');
            
            // 確保設置為合併 Excel 的預設值
            dialogHeader.innerHTML = '💹 合併 Excel 檔案';
            executeBtn.textContent = '匯出';
            executeBtn.onclick = executeMerge;  // 確保綁定正確的函數
            
            // === 修改：確保檔案輸入的 change 事件正確綁定 ===
            const mergeFileInput = document.getElementById('mergeFileInput');
            if (mergeFileInput) {
                // 先移除舊的事件監聽器
                const newMergeFileInput = mergeFileInput.cloneNode(true);
                mergeFileInput.parentNode.replaceChild(newMergeFileInput, mergeFileInput);
                
                // 重新綁定 change 事件
                document.getElementById('mergeFileInput').addEventListener('change', function(e) {
                    if (e.target.files && e.target.files.length > 0) {
                        handleMergeFileSelect(e.target.files);
                    }
                });
            }
            
            // === 重要：重新綁定檔案選擇按鈕事件 ===
            const selectFileBtn = document.getElementById('selectFileBtn');
            if (selectFileBtn) {
                // 不要替換元素，直接重新綁定事件
                selectFileBtn.onclick = function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('mergeFileInput').click();
                };
            }
            
            // 防止背景滾動
            document.body.style.overflow = 'hidden';
            document.body.style.position = 'fixed';
            document.body.style.width = '100%';
            
            document.getElementById('mergeDialogOverlay').style.display = 'flex';
            clearMergeSelection();
            
            // 設置初始路徑（使用主介面的路徑）
            const mainPath = document.getElementById('pathInput').value;
            if (mainPath) {
                document.getElementById('mergePathInput').value = mainPath;
            }
        }

        // 關閉合併對話框
        function closeMergeDialog() {
            // 清除對話框中的訊息
            clearDialogMessage();
            
            // 恢復背景滾動
            document.body.style.overflow = '';
            document.body.style.position = '';
            document.body.style.width = '';
            
            document.getElementById('mergeDialogOverlay').style.display = 'none';
            hideMergeAutocomplete();
        }

        // 清除選擇
        function clearMergeSelection() {
            selectedMergeFiles = [];
            selectedMergeFilePaths = [];
            document.getElementById('mergePathInput').value = '';
            document.getElementById('mergeFileInfo').style.display = 'none';
            document.getElementById('mergeFileInput').value = '';
            document.getElementById('selectedMergeFiles').innerHTML = '';
            hideMergeAutocomplete();
        }

        // 隱藏自動完成
        function hideMergeAutocomplete() {
            document.getElementById('mergePathAutocomplete').style.display = 'none';
            document.getElementById('mergePathInput').classList.remove('autocomplete-open');
            mergeSelectedSuggestionIndex = -1;
            mergeCurrentSuggestions = [];
        }

        // 顯示自動完成
        function showMergeAutocomplete() {
            document.getElementById('mergePathAutocomplete').style.display = 'block';
            document.getElementById('mergePathInput').classList.add('autocomplete-open');
        }

        // 獲取路徑建議（專門用於 Excel 檔案）
        async function fetchMergePathSuggestions(path) {
            const autocompleteDiv = document.getElementById('mergePathAutocomplete');
            
            autocompleteDiv.innerHTML = '<div class="path-loading">載入中...</div>';
            showMergeAutocomplete();
            
            try {
                const response = await fetch('/suggest-excel-path', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ path: path })
                });
                
                const data = await response.json();
                mergeCurrentSuggestions = data.suggestions || [];
                mergeSelectedSuggestionIndex = -1;
                
                if (mergeCurrentSuggestions.length > 0) {
                    displayMergeSuggestions(mergeCurrentSuggestions);
                } else {
                    autocompleteDiv.innerHTML = '<div class="path-loading">沒有找到 Excel 檔案</div>';
                }
            } catch (error) {
                console.error('Error fetching suggestions:', error);
                hideMergeAutocomplete();
            }
        }

        // 顯示建議
        function displayMergeSuggestions(suggestions) {
            const autocompleteDiv = document.getElementById('mergePathAutocomplete');
            autocompleteDiv.innerHTML = '';
            
            // 創建內部容器
            const innerDiv = document.createElement('div');
            innerDiv.className = 'merge-path-autocomplete-inner';
            
            suggestions.forEach((suggestion, index) => {
                const div = document.createElement('div');
                div.className = 'path-suggestion';
                div.dataset.index = index;
                
                // 處理資料夾標記
                let actualPath = suggestion;
                let displayPath = suggestion;
                
                if (suggestion.endsWith(' 📁')) {
                    actualPath = suggestion.replace(' 📁', '');
                    displayPath = actualPath;
                    // 在路徑後添加圖示
                    div.innerHTML = `<span>${escapeHtml(displayPath)}</span> <span style="color: #ffc107;">📁</span>`;
                } else if (suggestion.endsWith('.xlsx')) {
                    div.innerHTML = `<span>${escapeHtml(suggestion)}</span> <span style="color: #28a745;">📊</span>`;
                } else {
                    div.textContent = suggestion;
                }
                
                div.dataset.path = actualPath;
                
                // 添加 title 屬性但不顯示瀏覽器預設 tooltip
                div.setAttribute('data-full-path', actualPath);
                
                div.addEventListener('click', function(e) {
                    e.stopPropagation();
                    applyMergeSuggestion(actualPath);
                });
                
                div.addEventListener('mouseenter', function() {
                    selectMergeSuggestion(index);
                });
                
                innerDiv.appendChild(div);
            });
            
            autocompleteDiv.appendChild(innerDiv);
            showMergeAutocomplete();
        }

        // 選擇建議
        function selectMergeSuggestion(index) {
            const suggestions = document.querySelectorAll('#mergePathAutocomplete .path-suggestion');
            
            suggestions.forEach(s => s.classList.remove('selected'));
            
            if (index < 0) index = suggestions.length - 1;
            if (index >= suggestions.length) index = 0;
            
            mergeSelectedSuggestionIndex = index;
            
            if (index >= 0 && index < suggestions.length) {
                const selectedElement = suggestions[index];
                selectedElement.classList.add('selected');
                
                // 確保選中的項目在視窗中可見（只處理垂直捲動）
                const container = document.getElementById('mergePathAutocomplete');
                const elementTop = selectedElement.offsetTop;
                const elementBottom = elementTop + selectedElement.offsetHeight;
                const containerTop = container.scrollTop;
                const containerBottom = containerTop + container.clientHeight;
                
                if (elementTop < containerTop) {
                    container.scrollTop = elementTop;
                } else if (elementBottom > containerBottom) {
                    container.scrollTop = elementBottom - container.clientHeight;
                }
            }
        }

        // 修改應用建議函數，支援添加多個伺服器路徑
        function applyMergeSuggestion(suggestion) {
            const pathInput = document.getElementById('mergePathInput');
            pathInput.value = suggestion;
            hideMergeAutocomplete();
            
            // 如果是 xlsx 檔案，添加到路徑列表
            if (suggestion.endsWith('.xlsx')) {
                if (!selectedMergeFilePaths.includes(suggestion)) {
                    selectedMergeFilePaths.push(suggestion);
                    updateSelectedFilesDisplay();
                }
            }
        }

        // 處理檔案選擇（支援多檔）- 用於合併對話框
        function handleMergeFileSelect(files) {
            if (!files || files.length === 0) {
                showDialogMessage('請選擇 .xlsx 格式的 Excel 檔案', 'error');
                return;
            }
            
            // 清除之前的訊息
            clearDialogMessage();
            
            // 驗證所有檔案
            for (let file of files) {
                if (!file.name.endsWith('.xlsx')) {
                    showDialogMessage(`檔案 ${file.name} 不是 .xlsx 格式`, 'error');
                    return;
                }
                selectedMergeFiles.push(file);
            }
            
            // 顯示成功訊息
            showDialogMessage(`成功添加 ${files.length} 個檔案`, 'success');
            
            // 顯示檔案資訊
            updateSelectedFilesDisplay();
        }

        // 更新檔案顯示
        function updateSelectedFilesDisplay() {
            const filesDiv = document.getElementById('selectedMergeFiles');
            filesDiv.innerHTML = '';
            
            // 建立檔案列表容器
            const listContainer = document.createElement('div');
            listContainer.style.cssText = 'max-height: 300px; overflow-y: auto;';
            
            // 顯示本地檔案
            selectedMergeFiles.forEach((file, index) => {
                const fileItem = document.createElement('div');
                fileItem.className = 'selected-item';
                fileItem.innerHTML = `
                    <label style="display: flex; align-items: center; flex: 1; cursor: pointer;">
                        <input type="checkbox" checked style="margin-right: 10px;" onchange="updateSelectAllButtonText()">
                        <span>📄 ${file.name}</span>
                    </label>
                    <button class="btn-clear" style="padding: 2px 8px; font-size: 12px;" onclick="removeFile(${index}, 'local')">移除</button>
                `;
                listContainer.appendChild(fileItem);
            });
            
            // 顯示伺服器路徑檔案
            selectedMergeFilePaths.forEach((path, index) => {
                const fileItem = document.createElement('div');
                fileItem.className = 'selected-item';
                const fileName = path.split('/').pop().split('\\').pop();
                fileItem.innerHTML = `
                    <label style="display: flex; align-items: center; flex: 1; cursor: pointer;">
                        <input type="checkbox" checked style="margin-right: 10px;" onchange="updateSelectAllButtonText()">
                        <span>📊 ${fileName}</span>
                    </label>
                    <button class="btn-clear" style="padding: 2px 8px; font-size: 12px;" onclick="removeFile(${index}, 'path')">移除</button>
                `;
                listContainer.appendChild(fileItem);
            });
            
            filesDiv.appendChild(listContainer);
            
            // 顯示或隱藏檔案資訊區域
            document.getElementById('mergeFileInfo').style.display = 
                (selectedMergeFiles.length > 0 || selectedMergeFilePaths.length > 0) ? 'block' : 'none';
            
            // 更新全選按鈕文字
            updateSelectAllButtonText();
            
            // 清空路徑輸入
            if (selectedMergeFiles.length > 0) {
                document.getElementById('mergePathInput').value = '';
                hideMergeAutocomplete();
            }
        }

        // 移除單個檔案
        function removeFile(index, type) {
            if (type === 'local') {
                selectedMergeFiles.splice(index, 1);
            } else {
                selectedMergeFilePaths.splice(index, 1);
            }
            updateSelectedFilesDisplay();
        }

        // 執行合併（支援多檔）
        async function executeMerge() {
            if (selectedMergeFiles.length === 0 && selectedMergeFilePaths.length === 0) {
                showDialogMessage('請選擇要合併的 Excel 檔案', 'error');
                return;
            }
            
            // 檢查是否有當前分析結果
            const hasCurrentAnalysis = window.vpAnalyzeOutputPath && allLogs && allLogs.length > 0;
            const currentPath = document.getElementById('pathInput').value;
            
            // 禁用按鈕
            const executeBtn = document.getElementById('mergeExecuteBtn');
            executeBtn.disabled = true;
            executeBtn.textContent = '合併中...';
            
            try {
                // 準備 FormData
                const formData = new FormData();
                
                // 如果有分析結果，加入分析相關資料
                if (hasCurrentAnalysis) {
                    formData.append('current_path', currentPath);
                    formData.append('analysis_output_path', window.vpAnalyzeOutputPath);
                    formData.append('logs', JSON.stringify(allLogs));
                    formData.append('has_analysis', 'true');
                } else {
                    formData.append('has_analysis', 'false');
                }
                
                // 添加所有本地檔案
                selectedMergeFiles.forEach((file, index) => {
                    formData.append(`files`, file);
                });
                
                // 添加所有伺服器路徑
                formData.append('file_paths', JSON.stringify(selectedMergeFilePaths));
                
                const response = await fetch('/merge-multiple-excel', {
                    method: 'POST',
                    body: formData
                });
                
                if (response.ok) {
                    // 下載合併後的檔案
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    
                    // 從 header 獲取檔名
                    const contentDisposition = response.headers.get('content-disposition');
                    let filename = 'merged_anr_tombstone_result.xlsx';
                    if (contentDisposition) {
                        const filenameMatch = contentDisposition.match(/filename="?(.+?)"?$/);
                        if (filenameMatch) {
                            filename = filenameMatch[1];
                        }
                    }
                    
                    a.download = filename;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    
                    const fileCount = selectedMergeFiles.length + selectedMergeFilePaths.length;
                    const message = hasCurrentAnalysis ? 
                        `成功合併 ${fileCount} 個 Excel 檔案與當前分析結果` : 
                        `成功合併 ${fileCount} 個 Excel 檔案`;
                    showDialogMessage(message, 'success');
                    // 延遲關閉對話框，讓用戶看到成功訊息
                    setTimeout(() => {
                        closeMergeDialog();
                    }, 1000);
                } else {
                    const error = await response.text();
                    try {
                        const errorData = JSON.parse(error);
                        showDialogMessage('合併失敗: ' + (errorData.error || '未知錯誤'), 'error');
                    } catch {
                        showDialogMessage('合併失敗: ' + error, 'error');
                    }
                }
            } catch (error) {
                showDialogMessage('合併失敗: ' + error.message, 'error');
            } finally {
                executeBtn.disabled = false;
                executeBtn.textContent = '匯出';
            }
        }

        // 在 DOMContentLoaded 事件中添加事件監聽器
        document.addEventListener('DOMContentLoaded', function() {

            const mergeAutocomplete = document.getElementById('mergePathAutocomplete');
            if (mergeAutocomplete) {
                // 防止點擊自動完成框時關閉
                mergeAutocomplete.addEventListener('mousedown', function(e) {
                    e.stopPropagation();
                });
                
                // 處理捲軸事件
                mergeAutocomplete.addEventListener('scroll', function(e) {
                    e.stopPropagation();
                });
            }

            document.addEventListener('keydown', function(e) {
                if (e.key === 'Escape') {
                    const mergeDialog = document.getElementById('mergeDialogOverlay');
                    if (mergeDialog && mergeDialog.style.display === 'flex') {
                        closeMergeDialog();
                    }
                }
            });

            const mergeDialogOverlay = document.getElementById('mergeDialogOverlay');
            if (mergeDialogOverlay) {
                // 點擊遮罩層關閉對話框
                // mergeDialogOverlay.addEventListener('click', function(e) {
                //    if (e.target === this) {
                //        closeMergeDialog();
                //    }
                //});
                
                // 防止對話框內的點擊事件冒泡
                const mergeDialog = mergeDialogOverlay.querySelector('.merge-dialog');
                if (mergeDialog) {
                    mergeDialog.addEventListener('click', function(e) {
                        e.stopPropagation();
                    });
                }
            }
                    
            // 合併 Excel 輸入框事件
            const mergePathInput = document.getElementById('mergePathInput');
            if (mergePathInput) {
                mergePathInput.addEventListener('input', function(e) {
                    clearTimeout(mergeAutocompleteTimeout);
                    const value = e.target.value;
                    
                    mergeAutocompleteTimeout = setTimeout(() => {
                        fetchMergePathSuggestions(value);
                    }, 300);
                });
                
                mergePathInput.addEventListener('keydown', function(e) {
                    switch(e.key) {
                        case 'ArrowDown':
                            if (mergeCurrentSuggestions.length > 0) {
                                e.preventDefault();
                                selectMergeSuggestion(mergeSelectedSuggestionIndex + 1);
                            }
                            break;
                        case 'ArrowUp':
                            if (mergeCurrentSuggestions.length > 0) {
                                e.preventDefault();
                                selectMergeSuggestion(mergeSelectedSuggestionIndex - 1);
                            }
                            break;
                        case 'Enter':
                            if (mergeCurrentSuggestions.length > 0 && mergeSelectedSuggestionIndex >= 0) {
                                e.preventDefault();
                                applyMergeSuggestion(mergeCurrentSuggestions[mergeSelectedSuggestionIndex]);
                            } else if (mergeCurrentSuggestions.length === 0 || mergeSelectedSuggestionIndex === -1) {
                                hideMergeAutocomplete();
                            }
                            break;
                        case 'Tab':
                            if (mergeCurrentSuggestions.length > 0) {
                                e.preventDefault();
                                const index = mergeSelectedSuggestionIndex >= 0 ? mergeSelectedSuggestionIndex : 0;
                                applyMergeSuggestion(mergeCurrentSuggestions[index]);
                            }
                            break;
                        case 'Escape':
                            hideMergeAutocomplete();
                            break;
                    }
                });
                
                mergePathInput.addEventListener('focus', function() {
                    if (mergeCurrentSuggestions.length > 0) {
                        showMergeAutocomplete();
                    } else if (this.value) {
                        fetchMergePathSuggestions(this.value);
                    }
                });
                
                mergePathInput.addEventListener('blur', function(e) {
                    setTimeout(() => {
                        if (!document.activeElement.closest('#mergePathAutocomplete')) {
                            hideMergeAutocomplete();
                        }
                    }, 200);
                });
            }
            
            // 檔案輸入事件
            const mergeFileInput = document.getElementById('mergeFileInput');
            if (mergeFileInput) {
                // 移除舊的事件監聽器
                const newMergeFileInput = mergeFileInput.cloneNode(true);
                mergeFileInput.parentNode.replaceChild(newMergeFileInput, mergeFileInput);
                
                // 綁定新的事件
                document.getElementById('mergeFileInput').addEventListener('change', function(e) {
                    if (e.target.files && e.target.files.length > 0) {
                        handleMergeFileSelect(e.target.files);
                    }
                });
            }
            
            // 拖曳功能
            const dropZone = document.getElementById('mergeDropZone');
            if (dropZone) {
                // 不要替換 dropZone，直接綁定事件
                
                // 拖曳相關事件
                dropZone.addEventListener('dragover', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    this.classList.add('drag-over');
                });
                
                dropZone.addEventListener('dragleave', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    this.classList.remove('drag-over');
                });
                
                dropZone.addEventListener('drop', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    this.classList.remove('drag-over');
                    
                    const files = e.dataTransfer.files;
                    if (files.length > 0) {
                        handleMergeFileSelect(files);
                    }
                });
            }
            
            // 點擊對話框外部關閉
            //document.getElementById('mergeDialogOverlay').addEventListener('click', function(e) {
            //    if (e.target === this) {
            //        closeMergeDialog();
            //    }
            //});
        });

    </script>
    <script>
        // 載入 Excel 相關變數
        let loadExcelMode = false;
        let selectedLoadExcelPath = null;

        // 打開載入 Excel 對話框
        function openLoadExcelDialog() {
            loadExcelMode = true;
            
            // 清除之前的選擇
            clearMergeSelection();
            
            // 使用現有的合併對話框，但修改標題和按鈕
            const dialog = document.getElementById('mergeDialogOverlay');
            const dialogHeader = dialog.querySelector('.merge-dialog-header h3');
            const executeBtn = document.getElementById('mergeExecuteBtn');
            
            // 修改對話框內容
            dialogHeader.innerHTML = '📊 載入 Excel 檔案';
            executeBtn.textContent = '分析 Report';
            executeBtn.onclick = executeLoadExcel;  // 改變按鈕功能
            
            // 修改提示文字以反映多檔支援
            const dropZoneText = dialog.querySelector('.drop-zone-content p');
            if (dropZoneText) {
                dropZoneText.textContent = '拖曳 Excel 檔案到這裡（支援多檔）';
            }
            
            // === 修改：確保檔案輸入的 change 事件正確綁定 ===
            const mergeFileInput = document.getElementById('mergeFileInput');
            if (mergeFileInput) {
                // 先移除舊的事件監聽器
                const newMergeFileInput = mergeFileInput.cloneNode(true);
                mergeFileInput.parentNode.replaceChild(newMergeFileInput, mergeFileInput);
                
                // 重新綁定 change 事件
                document.getElementById('mergeFileInput').addEventListener('change', function(e) {
                    if (e.target.files && e.target.files.length > 0) {
                        handleMergeFileSelect(e.target.files);
                    }
                });
            }
            
            // === 重要：重新綁定檔案選擇按鈕事件 ===
            const selectFileBtn = document.getElementById('selectFileBtn');
            if (selectFileBtn) {
                // 不要替換元素，直接重新綁定事件
                selectFileBtn.onclick = function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('mergeFileInput').click();
                };
            }
            
            // 隱藏匯出相關按鈕
            const exportBtns = document.querySelectorAll('.export-html-btn, .export-excel-btn, .merge-excel-btn');
            exportBtns.forEach(btn => {
                if (btn) btn.style.display = 'none';
            });
            
            // 開啟對話框
            document.body.style.overflow = 'hidden';
            document.body.style.position = 'fixed';
            document.body.style.width = '100%';
            
            document.getElementById('mergeDialogOverlay').style.display = 'flex';
            
            // 設置初始路徑（使用主介面的路徑）
            const mainPath = document.getElementById('pathInput').value;
            if (mainPath) {
                document.getElementById('mergePathInput').value = mainPath;
            }
        }

        // 執行載入 Excel
        async function executeLoadExcel() {
            if (selectedMergeFiles.length === 0 && selectedMergeFilePaths.length === 0) {
                showDialogMessage('請選擇要載入的 Excel 檔案', 'error');
                return;
            }

            // 添加調試信息
            console.log('=== 檔案選擇情況 ===');
            console.log('selectedMergeFiles:', selectedMergeFiles);
            console.log('selectedMergeFilePaths:', selectedMergeFilePaths);
                        
            const executeBtn = document.getElementById('mergeExecuteBtn');
            executeBtn.disabled = true;
            executeBtn.textContent = '分析中...';
            
            try {
                let formData = new FormData();
                
                // 收集所有檔案名稱和路徑
                let allFilenames = [];
                let allPaths = [];

                // 收集本地檔案資訊
                selectedMergeFiles.forEach(file => {
                    allFilenames.push(file.name);
                    // 本地檔案確實沒有路徑，只能顯示「本地上傳」
                    allPaths.push("本地上傳");
                });

                // 收集伺服器檔案資訊
                selectedMergeFilePaths.forEach(path => {
                    const filename = path.split('/').pop().split('\\').pop();
                    allFilenames.push(filename);
                    // 這裡應該要顯示完整路徑
                    allPaths.push(path);  // 確保這裡是完整路徑
                });
                
                // 添加調試信息
                console.log('檔案名稱:', allFilenames);
                console.log('檔案路徑:', allPaths);
                
                // 處理多個檔案
                if (selectedMergeFiles.length + selectedMergeFilePaths.length > 1) {
                    // 需要先合併
                    const mergeFormData = new FormData();
                    mergeFormData.append('has_analysis', 'false');
                    
                    // 添加所有本地檔案
                    selectedMergeFiles.forEach(file => {
                        mergeFormData.append('files', file);
                    });
                    
                    // 添加伺服器路徑檔案
                    mergeFormData.append('file_paths', JSON.stringify(selectedMergeFilePaths));
                    
                    // 呼叫合併 API
                    const mergeResponse = await fetch('/merge-multiple-excel', {
                        method: 'POST',
                        body: mergeFormData
                    });
                    
                    if (!mergeResponse.ok) {
                        throw new Error('合併檔案失敗');
                    }
                    
                    // 獲取合併後的檔案
                    const mergedBlob = await mergeResponse.blob();
                    const mergedFile = new File([mergedBlob], 'merged_result.xlsx', { 
                        type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' 
                    });
                    
                    // 使用合併後的檔案進行分析
                    formData.append('file', mergedFile);
                    formData.append('is_merged', 'true');
                    formData.append('file_count', allFilenames.length);
                    
                    // 傳遞原始檔案資訊
                    formData.append('merged_file_info', JSON.stringify({
                        filenames: allFilenames,
                        paths: allPaths
                    }));
                    
                } else {
                    // 只有一個檔案
                    if (selectedMergeFiles.length > 0) {
                        formData.append('file', selectedMergeFiles[0]);
                    } else {
                        formData.append('file_path', selectedMergeFilePaths[0]);
                    }
                    formData.append('is_merged', 'false');
                    formData.append('file_count', '1');
                }
                
                // 發送到報告路由
                const response = await fetch('/load-excel-report', {
                    method: 'POST',
                    body: formData
                });
                
                if (response.ok) {
                    const data = await response.json();
                    if (data.report_url) {
                        // 開啟新視窗顯示報告
                        window.open(data.report_url, '_blank');
                        closeMergeDialog();
                        
                        const fileCount = allFilenames.length;
                        const message = fileCount > 1 ? 
                            `已合併 ${fileCount} 個 Excel 檔案並生成報告` : 
                            'Excel 載入成功，報告已在新視窗開啟';
                        showDialogMessage(message, 'success');
                        // 延遲關閉對話框
                        setTimeout(() => {
                            closeMergeDialog();
                        }, 1000);
                    }
                } else {
                    const error = await response.json();
                    showDialogMessage('載入失敗: ' + (error.error || '未知錯誤'), 'error');
                }
            } catch (error) {
                showDialogMessage('載入失敗: ' + error.message, 'error');
            } finally {
                executeBtn.disabled = false;
                executeBtn.textContent = '分析 Report';
                loadExcelMode = false;
            }
        }

        // 修改 closeMergeDialog 函數，重置對話框
        const originalCloseMergeDialog = closeMergeDialog;
        closeMergeDialog = function() {
            // 先重置對話框內容（無論是否在 loadExcelMode）
            const dialogHeader = document.querySelector('.merge-dialog-header h3');
            const executeBtn = document.getElementById('mergeExecuteBtn');
            
            // 總是重置為預設值
            if (dialogHeader) {
                dialogHeader.innerHTML = '💹 合併 Excel 檔案';
            }
            if (executeBtn) {
                executeBtn.textContent = '匯出';
                executeBtn.onclick = executeMerge;
            }
            
            // 呼叫原始的關閉函數
            originalCloseMergeDialog();
            
            // 確保 loadExcelMode 被重置
            loadExcelMode = false;
            
            // 恢復匯出按鈕顯示
            const exportBtns = document.querySelectorAll('.export-html-btn, .export-excel-btn, .merge-excel-btn');
            exportBtns.forEach(btn => {
                if (btn && window.hasCurrentAnalysis) {
                    btn.style.display = 'block';
                }
            });
        };  
    </script>
    <script>
        async function exportExcelReport() {
            if (!currentAnalysisId || !allLogs || allLogs.length === 0) {
                showMessage('請先執行分析', 'error');
                return;
            }
            
            const exportBtn = document.getElementById('exportExcelReportBtn');
            if (!exportBtn) return;
            
            exportBtn.disabled = true;
            exportBtn.textContent = '生成報表中...';
            
            try {
                const response = await fetch('/export-excel-report', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        path: document.getElementById('pathInput').value,
                        analysis_id: currentAnalysisId
                    })
                });
                
                if (response.ok) {
                    const data = await response.json();
                    if (data.success && data.report_url) {
                        // 在新視窗開啟報表
                        window.open(data.report_url, '_blank');
                        showMessage('Excel 報表已生成', 'success');
                    }
                } else {
                    const error = await response.text();
                    showMessage('生成報表失敗: ' + error, 'error');
                }
            } catch (error) {
                showMessage('生成報表失敗: ' + error.message, 'error');
            } finally {
                exportBtn.disabled = false;
                exportBtn.textContent = '匯出 Excel 報表';
            }
        }    
    </script>
    <script>
        // 儲存歷史分析路徑資訊
        let historyAnalysisInfo = null;

        // 檢查是否有已存在的分析結果
        async function checkExistingAnalysis(path) {
            if (!path) return;
            
            try {
                const response = await fetch('/check-existing-analysis', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ path: path })
                });
                
                const data = await response.json();
                
                // 保存歷史分析資訊
                historyAnalysisInfo = data;
                const historySection = document.getElementById('historySection');
                const viewIndexBtn = document.getElementById('viewIndexBtn');
                const downloadExcelBtn = document.getElementById('downloadExcelBtn');
                const viewHTMLBtn = document.getElementById('viewHTMLBtn');
                const viewExcelReportBtn = document.getElementById('viewExcelReportBtn');
                const downloadZipBtn = document.getElementById('downloadZipBtn');
                
                console.log('檢查結果:', data); // 調試用
                
                if (data.exists && data.has_folder) {
                    historySection.style.display = 'block';
                    
                    // 根據檔案存在狀況顯示按鈕
                    viewIndexBtn.style.display = data.has_index ? 'inline-flex' : 'none';
                    downloadExcelBtn.style.display = data.has_excel ? 'inline-flex' : 'none';
                    viewHTMLBtn.style.display = data.has_html ? 'inline-flex' : 'none';
                    
                    // 特別檢查 Excel 報表
                    if (data.has_excel_report) {
                        viewExcelReportBtn.style.display = 'inline-flex';
                        console.log('找到 Excel 報表:', data.excel_report_path);
                    } else {
                        viewExcelReportBtn.style.display = 'none';
                    }
                    
                    downloadZipBtn.style.display = 'inline-flex';
                } else {
                    historySection.style.display = 'none';
                }
                
            } catch (error) {
                console.error('檢查已有分析結果失敗:', error);
                document.getElementById('historySection').style.display = 'none';
            }
        }

        // 查看已存在的 HTML 統計
        function viewExistingHTML() {
            if (historyAnalysisInfo && historyAnalysisInfo.html_path) {
                window.open('/view-analysis-report?path=' + encodeURIComponent(historyAnalysisInfo.html_path), '_blank');
            }
        }

        // 下載已存在的 Excel
        function downloadExistingExcel() {
            if (historyAnalysisInfo && historyAnalysisInfo.excel_path) {
                window.location.href = '/download-file?path=' + encodeURIComponent(historyAnalysisInfo.excel_path);
            }
        }

        // 下載歷史分析 zip
        async function downloadAnalysisZip() {
            if (!historyAnalysisInfo || !historyAnalysisInfo.analysis_path) return;
            
            // 獲取按鈕並添加轉場效果
            const downloadBtn = document.getElementById('downloadZipBtn');
            if (!downloadBtn) return;
            
            const originalText = downloadBtn.innerHTML;
            downloadBtn.disabled = true;
            downloadBtn.innerHTML = '📦 打包中...';
            
            try {
                const response = await fetch('/download-analysis-zip', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        analysis_path: historyAnalysisInfo.analysis_path
                    })
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    
                    const contentDisposition = response.headers.get('content-disposition');
                    let filename = 'analysis.zip';
                    if (contentDisposition) {
                        const filenameMatch = contentDisposition.match(/filename="?(.+?)"?$/);
                        if (filenameMatch) {
                            filename = filenameMatch[1];
                        }
                    }
                    
                    a.download = filename;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    
                    showMessage('分析結果打包完成', 'success');
                } else {
                    throw new Error('下載失敗');
                }
            } catch (error) {
                showMessage('下載失敗: ' + error.message, 'error');
            } finally {
                // 恢復按鈕狀態
                downloadBtn.disabled = false;
                downloadBtn.innerHTML = originalText;
            }
        }

        // 下載當前分析結果 zip
        async function downloadCurrentAnalysisZip() {
            if (!window.vpAnalyzeOutputPath) {
                showMessage('請先執行分析', 'error');
                return;
            }
            
            // 獲取按鈕並添加轉場效果
            const downloadBtn = document.getElementById('downloadCurrentZipBtn');
            if (!downloadBtn) return;
            
            const originalText = downloadBtn.innerHTML;
            downloadBtn.disabled = true;
            downloadBtn.innerHTML = '📦 打包中...';
            
            try {
                const response = await fetch('/download-analysis-zip', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        analysis_path: window.vpAnalyzeOutputPath
                    })
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    
                    const contentDisposition = response.headers.get('content-disposition');
                    let filename = 'analysis.zip';
                    if (contentDisposition) {
                        const filenameMatch = contentDisposition.match(/filename="?(.+?)"?$/);
                        if (filenameMatch) {
                            filename = filenameMatch[1];
                        }
                    }
                    
                    a.download = filename;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    
                    showMessage('分析結果打包完成', 'success');
                } else {
                    throw new Error('下載失敗');
                }
            } catch (error) {
                showMessage('下載失敗: ' + error.message, 'error');
            } finally {
                // 恢復按鈕狀態
                downloadBtn.disabled = false;
                downloadBtn.innerHTML = originalText;
            }
        }

        // 查看歷史分析結果（從歷史文件區塊）
        function viewHistoryIndex() {
            if (historyAnalysisInfo && historyAnalysisInfo.analysis_path) {
                window.open('/view-analysis-report?path=' + encodeURIComponent(historyAnalysisInfo.analysis_path), '_blank');
            }
        }

        // 自動匯出 HTML 到分析資料夾
        async function autoExportHTML(outputPath) {
            if (!currentAnalysisId || !outputPath) {
                throw new Error('缺少必要參數');
            }
            
            try {
                // 獲取當前服務器資訊
                const serverResponse = await fetch('/server-info');
                const serverInfo = await serverResponse.json();
                
                // 呼叫匯出 API，並指定儲存路徑
                const response = await fetch(`/export-html-to-folder/${currentAnalysisId}`, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        output_path: outputPath,
                        base_url: serverInfo.base_url
                    })
                });
                
                if (!response.ok) {
                    throw new Error('匯出 HTML 失敗');
                }
                
                const result = await response.json();
                console.log('HTML 已儲存到:', result.saved_path);
                
            } catch (error) {
                console.error('自動匯出 HTML 錯誤:', error);
                throw error;
            }
        }

        // 自動匯出 Excel 到分析資料夾
        async function autoExportExcel(outputPath) {
            if (!outputPath || !window.vpAnalyzeOutputPath) {
                throw new Error('缺少必要參數');
            }
            
            try {
                const response = await fetch('/export-excel-to-folder', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        path: document.getElementById('pathInput').value,
                        analysis_output_path: window.vpAnalyzeOutputPath,
                        output_folder: outputPath,
                        logs: allLogs
                    })
                });
                
                if (!response.ok) {
                    throw new Error('匯出 Excel 失敗');
                }
                
                const result = await response.json();
                console.log('Excel 已儲存到:', result.saved_path);
                
            } catch (error) {
                console.error('自動匯出 Excel 錯誤:', error);
                throw error;
            }
        }

        // 自動匯出 Excel 報表到分析資料夾
        async function autoExportExcelReport(outputPath) {
            if (!currentAnalysisId || !outputPath) {
                throw new Error('缺少必要參數');
            }
            
            try {
                const response = await fetch('/export-excel-report-to-folder', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        path: document.getElementById('pathInput').value,
                        analysis_id: currentAnalysisId,
                        output_folder: outputPath
                    })
                });
                
                if (!response.ok) {
                    throw new Error('匯出 Excel 報表失敗');
                }
                
                const result = await response.json();
                console.log('Excel 報表已儲存到:', result.saved_path);
                
            } catch (error) {
                console.error('自動匯出 Excel 報表錯誤:', error);
                throw error;
            }
        }

    </script>
    <script>
        // 檔案/資料夾選擇相關變數
        let selectedFiles = [];
        let selectedFolders = [];

        // 開啟檔案選擇對話框
        function openFileSelectDialog() {
            document.getElementById('fileSelectDialogOverlay').style.display = 'flex';
            document.body.style.overflow = 'hidden';
            document.body.style.position = 'fixed';
            document.body.style.width = '100%';
            clearFileSelection();
        }

        // 關閉檔案選擇對話框
        function closeFileSelectDialog() {
            document.getElementById('fileSelectDialogOverlay').style.display = 'none';
            document.body.style.overflow = '';
            document.body.style.position = '';
            document.body.style.width = '';
        }

        // 清除選擇
        function clearFileSelection() {
            selectedFiles = [];
            selectedFolders = [];
            document.getElementById('selectedItemsList').innerHTML = '';
            document.getElementById('selectedItemsSection').style.display = 'none';
            document.getElementById('fileSelectInput').value = '';
            document.getElementById('folderSelectInput').value = '';
        }

        // 更新已選擇項目顯示
        function updateSelectedItemsDisplay() {
            const listDiv = document.getElementById('selectedItemsList');
            listDiv.innerHTML = '';
            
            let itemCount = 0;
            
            // 顯示選擇的檔案
            selectedFiles.forEach((file, index) => {
                const item = createItemDisplay(file.name, 'file', () => removeItem('file', index));
                listDiv.appendChild(item);
                itemCount++;
            });
            
            // 顯示選擇的資料夾（透過檔案列表）
            if (selectedFolders.length > 0) {
                // 計算資料夾數量
                const folderPaths = new Set();
                selectedFolders.forEach(file => {
                    const pathParts = file.webkitRelativePath.split('/');
                    if (pathParts.length > 1) {
                        folderPaths.add(pathParts[0]);
                    }
                });
                
                folderPaths.forEach(folderName => {
                    const item = createItemDisplay(folderName, 'folder', () => removeFolderByName(folderName));
                    listDiv.appendChild(item);
                    itemCount++;
                });
            }
            
            // 顯示或隱藏區域
            document.getElementById('selectedItemsSection').style.display = itemCount > 0 ? 'block' : 'none';
        }

        // 建立項目顯示元素
        function createItemDisplay(name, type, removeCallback) {
            const div = document.createElement('div');
            div.className = 'selected-item';
            
            const icon = type === 'file' ? '📄' : '📁';
            div.innerHTML = `
                <span>${icon} ${name}</span>
                <button class="btn-clear" style="padding: 2px 8px; font-size: 12px;">移除</button>
            `;
            
            div.querySelector('.btn-clear').onclick = removeCallback;
            return div;
        }

        // 移除項目
        function removeItem(type, index) {
            if (type === 'file') {
                selectedFiles.splice(index, 1);
            }
            updateSelectedItemsDisplay();
        }

        // 移除資料夾
        function removeFolderByName(folderName) {
            selectedFolders = selectedFolders.filter(file => {
                const pathParts = file.webkitRelativePath.split('/');
                return pathParts[0] !== folderName;
            });
            updateSelectedItemsDisplay();
        }

        // 處理檔案選擇
        function handleFileSelect(files) {
            if (!files || files.length === 0) return;
            
            // 轉換 FileList 為 Array
            const fileArray = Array.from(files);
            
            // 添加到選擇列表（避免重複）
            fileArray.forEach(file => {
                const exists = selectedFiles.some(f => 
                    f.name === file.name && f.size === file.size
                );
                if (!exists) {
                    selectedFiles.push(file);
                }
            });
            
            updateSelectedItemsDisplay();
        }

        // 處理資料夾選擇
        function handleFolderSelect(files) {
            if (!files || files.length === 0) return;
            
            // 轉換 FileList 為 Array
            const fileArray = Array.from(files);
            
            // 替換現有的資料夾選擇
            selectedFolders = fileArray;
            
            updateSelectedItemsDisplay();
        }

        // 執行檔案分析
        async function executeFileAnalysis() {
            const totalFiles = selectedFiles.length + selectedFolders.length;
            
            if (totalFiles === 0) {
                showMessage('請選擇要分析的檔案或資料夾', 'error');
                return;
            }
            
            const executeBtn = document.getElementById('fileAnalysisExecuteBtn');
            executeBtn.disabled = true;
            executeBtn.textContent = '準備中...';
            
            try {
                const formData = new FormData();
                
                // 添加單獨的檔案
                selectedFiles.forEach(file => {
                    formData.append('files', file);
                });
                
                // 添加資料夾中的檔案
                selectedFolders.forEach(file => {
                    formData.append('folder_files', file, file.webkitRelativePath);
                });
                
                // 添加選項
                formData.append('auto_group', document.getElementById('autoGroupFiles').checked);
                
                // 發送請求
                const response = await fetch('/analyze-selected-items', {
                    method: 'POST',
                    body: formData
                });
                
                if (response.ok) {
                    const data = await response.json();
                    
                    // 關閉對話框
                    closeFileSelectDialog();
                    
                    // 設定路徑並執行分析
                    if (data.temp_path) {
                        document.getElementById('pathInput').value = data.temp_path;
                        // 自動執行分析
                        analyzeLogs();
                    }
                } else {
                    const error = await response.json();
                    showMessage('準備失敗: ' + (error.error || '未知錯誤'), 'error');
                }
            } catch (error) {
                showMessage('準備失敗: ' + error.message, 'error');
            } finally {
                executeBtn.disabled = false;
                executeBtn.textContent = '開始分析';
            }
        }

        // 在 DOMContentLoaded 事件中設置事件監聽器
        document.addEventListener('DOMContentLoaded', function() {
            // 檔案選擇按鈕
            const selectFilesBtn = document.getElementById('selectLocalFilesBtn');
            if (selectFilesBtn) {
                selectFilesBtn.addEventListener('click', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('fileSelectInput').click();
                });
            }
            
            // 資料夾選擇按鈕
            const selectFolderBtn = document.getElementById('selectLocalFolderBtn');
            if (selectFolderBtn) {
                selectFolderBtn.addEventListener('click', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('folderSelectInput').click();
                });
            }
            
            // 檔案輸入變化事件
            const fileInput = document.getElementById('fileSelectInput');
            if (fileInput) {
                fileInput.addEventListener('change', function(e) {
                    handleFileSelect(e.target.files);
                });
            }
            
            // 資料夾輸入變化事件
            const folderInput = document.getElementById('folderSelectInput');
            if (folderInput) {
                folderInput.addEventListener('change', function(e) {
                    handleFolderSelect(e.target.files);
                });
            }
            
            // 拖曳功能
            const dropZone = document.getElementById('fileSelectDropZone');
            if (dropZone) {
                // 防止瀏覽器預設行為
                ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
                    dropZone.addEventListener(eventName, preventDefaults, false);
                    document.body.addEventListener(eventName, preventDefaults, false);
                });
                
                // 拖曳進入和離開的視覺效果
                ['dragenter', 'dragover'].forEach(eventName => {
                    dropZone.addEventListener(eventName, highlight, false);
                });
                
                ['dragleave', 'drop'].forEach(eventName => {
                    dropZone.addEventListener(eventName, unhighlight, false);
                });
                
                // 處理拖放
                dropZone.addEventListener('drop', handleDrop, false);
            }
        });

        // 防止預設行為
        function preventDefaults(e) {
            e.preventDefault();
            e.stopPropagation();
        }

        // 高亮拖放區域
        function highlight(e) {
            document.getElementById('fileSelectDropZone').classList.add('drag-over');
        }

        // 取消高亮
        function unhighlight(e) {
            document.getElementById('fileSelectDropZone').classList.remove('drag-over');
        }

        // 處理拖放
        async function handleDrop(e) {
            const dt = e.dataTransfer;
            const items = dt.items;
            
            if (items) {
                // 使用 DataTransferItemList 介面
                const files = [];
                
                for (let i = 0; i < items.length; i++) {
                    const item = items[i];
                    
                    if (item.kind === 'file') {
                        const entry = item.webkitGetAsEntry();
                        if (entry) {
                            if (entry.isFile) {
                                // 處理檔案
                                const file = item.getAsFile();
                                if (file) {
                                    files.push(file);
                                }
                            } else if (entry.isDirectory) {
                                // 處理資料夾
                                showMessage('拖放資料夾功能需要使用選擇資料夾按鈕', 'info');
                            }
                        }
                    }
                }
                
                if (files.length > 0) {
                    handleFileSelect(files);
                }
            } else {
                // 舊版瀏覽器
                const files = dt.files;
                handleFileSelect(files);
            }
        }

        // 頁籤切換功能
        function switchAnalysisTab(tab) {
            // 更新頁籤按鈕狀態
            document.querySelectorAll('.tab-button').forEach(btn => {
                btn.classList.remove('active');
            });
            
            // 更新內容顯示
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.remove('active');
            });
            
            if (tab === 'path') {
                document.getElementById('pathTabBtn').classList.add('active');
                document.getElementById('pathTabContent').classList.add('active');
            } else if (tab === 'files') {
                document.getElementById('filesTabBtn').classList.add('active');
                document.getElementById('filesTabContent').classList.add('active');
                
                // 新增：切換到檔案選擇頁籤時，檢查是否應該顯示按鈕
                if (!window.hasCurrentAnalysis) {
                    // 如果沒有當前分析結果，確保按鈕隱藏
                    const buttons = [
                        'exportExcelBtn',
                        'exportExcelReportBtn',
                        'mergeExcelBtn',
                        'exportHtmlBtn',
                        'downloadCurrentZipBtn'
                    ];
                    
                    buttons.forEach(id => {
                        const btn = document.getElementById(id);
                        if (btn) btn.style.display = 'none';
                    });
                }
            }
        }

        // 主頁面的檔案選擇相關變數
        let mainSelectedFiles = [];
        let mainSelectedFolders = [];

        // 清除主頁面選擇
        function clearMainFileSelection() {
            mainSelectedFiles = [];
            mainSelectedFolders = [];
            document.getElementById('mainSelectedItemsList').innerHTML = '';
            document.getElementById('mainSelectedItemsSection').style.display = 'none';
            document.getElementById('mainFileSelectInput').value = '';
            document.getElementById('mainFolderSelectInput').value = '';
        }

        // 更新主頁面已選擇項目顯示
        function updateMainSelectedItemsDisplay() {
            const listDiv = document.getElementById('mainSelectedItemsList');
            listDiv.innerHTML = '';
            
            let itemCount = 0;
            
            // 顯示選擇的檔案
            mainSelectedFiles.forEach((file, index) => {
                const item = createMainItemDisplay(file.name, 'file', index);
                listDiv.appendChild(item);
                itemCount++;
            });
            
            // 顯示選擇的資料夾
            if (mainSelectedFolders.length > 0) {
                const folderPaths = new Set();
                mainSelectedFolders.forEach(file => {
                    const pathParts = file.webkitRelativePath.split('/');
                    if (pathParts.length > 1) {
                        folderPaths.add(pathParts[0]);
                    }
                });
                
                let folderIndex = 0;
                folderPaths.forEach(folderName => {
                    const item = createMainItemDisplay(folderName, 'folder', folderIndex);
                    listDiv.appendChild(item);
                    itemCount++;
                    folderIndex++;
                });
            }
            
            document.getElementById('mainSelectedItemsSection').style.display = itemCount > 0 ? 'block' : 'none';
            
            // 更新全選按鈕文字
            updateMainSelectAllButtonText();
        }

        // 建立主頁面項目顯示元素
        function createMainItemDisplay(name, type, index) {
            const div = document.createElement('div');
            div.className = 'selected-item';
            
            const icon = type === 'file' ? '📄' : '📁';
            div.innerHTML = `
                <label style="display: flex; align-items: center; flex: 1; cursor: pointer;">
                    <input type="checkbox" checked style="margin-right: 10px;" onchange="updateMainSelectAllButtonText()">
                    <span>${icon} ${name}</span>
                </label>
                <button class="btn-clear" style="padding: 2px 8px; font-size: 12px;" 
                        onclick="removeMainItem('${type}', ${index})">移除</button>
            `;
            
            return div;
        }

        // 全選/取消全選主頁面所有檔案
        function toggleAllMainFiles() {
            const checkboxes = document.querySelectorAll('#mainSelectedItemsList input[type="checkbox"]');
            
            if (checkboxes.length === 0) return;
            
            // 檢查是否全部已選中
            const allChecked = Array.from(checkboxes).every(cb => cb.checked);
            
            // 切換選擇狀態
            checkboxes.forEach(cb => {
                cb.checked = !allChecked;
            });
            
            // 更新按鈕文字
            updateMainSelectAllButtonText();
        }

        // 更新主頁面全選按鈕的文字
        function updateMainSelectAllButtonText() {
            const checkboxes = document.querySelectorAll('#mainSelectedItemsList input[type="checkbox"]');
            const checkedCount = Array.from(checkboxes).filter(cb => cb.checked).length;
            const button = document.querySelector('#mainSelectedItemsSection .btn-select-all');
            
            if (button) {
                if (checkedCount === 0) {
                    button.textContent = '全選';
                } else if (checkedCount === checkboxes.length) {
                    button.textContent = '取消全選';
                } else {
                    button.textContent = `全選 (${checkedCount}/${checkboxes.length})`;
                }
            }
        }

        // 批量移除選中的主頁面檔案
        function removeSelectedMainFiles() {
            const items = document.querySelectorAll('#mainSelectedItemsList .selected-item');
            const toRemove = {
                files: [],
                folders: []
            };
            
            // 收集要移除的項目
            items.forEach((item, index) => {
                const checkbox = item.querySelector('input[type="checkbox"]');
                if (checkbox && checkbox.checked) {
                    const icon = item.querySelector('span').textContent.trim();
                    if (icon.startsWith('📄')) {
                        toRemove.files.push(index);
                    } else if (icon.startsWith('📁')) {
                        const folderName = icon.substring(2).trim(); // 移除圖示
                        toRemove.folders.push(folderName);
                    }
                }
            });
            
            // 從後往前刪除檔案，避免索引錯位
            toRemove.files.sort((a, b) => b - a).forEach(index => {
                if (index < mainSelectedFiles.length) {
                    mainSelectedFiles.splice(index, 1);
                }
            });
            
            // 移除資料夾
            if (toRemove.folders.length > 0) {
                mainSelectedFolders = mainSelectedFolders.filter(file => {
                    const pathParts = file.webkitRelativePath.split('/');
                    const folderName = pathParts[0];
                    return !toRemove.folders.includes(folderName);
                });
            }
            
            // 更新顯示
            updateMainSelectedItemsDisplay();
        }

        // 清除主頁面所有選擇
        function clearMainFileSelection() {
            mainSelectedFiles = [];
            mainSelectedFolders = [];
            document.getElementById('mainSelectedItemsList').innerHTML = '';
            document.getElementById('mainSelectedItemsSection').style.display = 'none';
            document.getElementById('mainFileSelectInput').value = '';
            document.getElementById('mainFolderSelectInput').value = '';
        }

        // 移除主頁面項目
        function removeMainItem(type, index) {
            if (type === 'file') {
                mainSelectedFiles.splice(index, 1);
            }
            updateMainSelectedItemsDisplay();
        }

        // 移除主頁面資料夾
        function removeMainFolderByName(folderName) {
            mainSelectedFolders = mainSelectedFolders.filter(file => {
                const pathParts = file.webkitRelativePath.split('/');
                return pathParts[0] !== folderName;
            });
            updateMainSelectedItemsDisplay();
        }

        // 處理主頁面檔案選擇
        function handleMainFileSelect(files) {
            if (!files || files.length === 0) return;
            
            const fileArray = Array.from(files);
            fileArray.forEach(file => {
                const exists = mainSelectedFiles.some(f => 
                    f.name === file.name && f.size === file.size
                );
                if (!exists) {
                    mainSelectedFiles.push(file);
                }
            });
            
            updateMainSelectedItemsDisplay();
        }

        // 處理主頁面資料夾選擇
        function handleMainFolderSelect(files) {
            if (!files || files.length === 0) return;
            mainSelectedFolders = Array.from(files);
            updateMainSelectedItemsDisplay();
        }

        // 執行主頁面檔案分析
        async function executeMainFileAnalysis() {
            const totalFiles = mainSelectedFiles.length + mainSelectedFolders.length;
            
            if (totalFiles === 0) {
                showMessage('請選擇要分析的檔案或資料夾', 'error');
                return;
            }
            
            // === 新增：重置狀態和隱藏按鈕 ===
            window.hasCurrentAnalysis = false;
            window.currentAnalysisExported = false;

            // 隱藏所有匯出按鈕
            const headerButtons = [
                'exportExcelBtn',
                'exportExcelReportBtn',
                'mergeExcelBtn',
                'exportHtmlBtn',
                'downloadCurrentZipBtn'
            ];

            headerButtons.forEach(id => {
                const btn = document.getElementById(id);
                if (btn) {
                    btn.style.display = 'none';
                }
            });

            // 隱藏結果區域
            document.getElementById('results').style.display = 'none';
            
            // 隱藏歷史區塊
            document.getElementById('historySection').style.display = 'none';

            // 隱藏所有浮動按鈕
            const floatingButtons = [
                'backToTop',
                'navToggleBtn', 
                'globalToggleBtn',
                'analysisResultBtn'
            ];
            
            floatingButtons.forEach(id => {
                const btn = document.getElementById(id);
                if (btn) {
                    btn.classList.remove('show');
                }
            });
            
            // 確保導覽列也關閉
            const navBar = document.getElementById('navBar');
            if (navBar && navBar.classList.contains('show')) {
                toggleNavBar();
            }
            
            const analyzeBtn = document.getElementById('mainFileAnalysisBtn');
            analyzeBtn.disabled = true;
            analyzeBtn.textContent = '準備中...';
            document.getElementById('loading').style.display = 'block';
            
            try {
                const formData = new FormData();
                
                // 添加單獨的檔案
                mainSelectedFiles.forEach(file => {
                    formData.append('files', file);
                });
                
                // 添加資料夾中的檔案
                mainSelectedFolders.forEach(file => {
                    formData.append('folder_files', file, file.webkitRelativePath);
                });
                
                // 添加選項
                formData.append('auto_group', document.getElementById('mainAutoGroupFiles').checked);
                
                // 發送請求
                const response = await fetch('/analyze-selected-items', {
                    method: 'POST',
                    body: formData
                });
                
                if (response.ok) {
                    const data = await response.json();
                    
                    // 設定路徑並執行分析
                    if (data.temp_path) {
                        // === 重要：儲存原始臨時路徑 ===
                        const originalTempPath = data.temp_path;
                        
                        // === 重要：先設置 basePath ===
                        window.basePath = data.temp_path;
                        console.log('設置 basePath:', window.basePath);
                        
                        // 使用臨時路徑直接呼叫 analyze 端點
                        const analyzeResponse = await fetch('/analyze', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({ path: data.temp_path })
                        });
                        
                        const analyzeData = await analyzeResponse.json();
                        
                        if (analyzeResponse.ok) {
                            // === 新增：檢查是否有找到資料 ===
                            if (analyzeData.files_with_cmdline === 0) {
                                // 隱藏所有匯出按鈕
                                headerButtons.forEach(id => {
                                    const btn = document.getElementById(id);
                                    if (btn) {
                                        btn.style.display = 'none';
                                    }
                                });
                                
                                // 隱藏導覽相關按鈕
                                floatingButtons.forEach(id => {
                                    const btn = document.getElementById(id);
                                    if (btn) {
                                        btn.classList.remove('show');
                                    }
                                });
                                
                                // 隱藏結果區域
                                document.getElementById('results').style.display = 'none';
                                
                                // 顯示錯誤訊息
                                showMessage('警告: 在選擇的檔案中沒有找到包含 Cmd line: 或 Cmdline: 的檔案', 'error');
                                
                                // 清除選擇
                                clearMainFileSelection();
                                
                                // 重置狀態
                                window.hasCurrentAnalysis = false;
                                window.currentAnalysisExported = false;
                                
                                return; // 提早結束函數
                            }
                            
                            // 只有在有資料時才繼續處理
                            currentAnalysisId = analyzeData.analysis_id;
                            
                            // === 重要：處理檔案路徑映射 ===
                            const sortedLogs = analyzeData.logs.sort((a, b) => {
                                if (!a.timestamp && !b.timestamp) return 0;
                                if (!a.timestamp) return 1;
                                if (!b.timestamp) return -1;
                                return a.timestamp.localeCompare(b.timestamp);
                            });
                            
                            // === 修正：確保檔案路徑正確 ===
                            allLogs = sortedLogs.map(log => {
                                // 添加除錯資訊
                                console.log('=== Log Path Debug ===');
                                console.log('Original log.file:', log.file);
                                console.log('basePath:', window.basePath);
                                
                                // 確保 file 路徑存在
                                if (!log.file && log.filepath) {
                                    log.file = log.filepath;
                                }
                                
                                // 如果是臨時目錄的檔案，保持原樣
                                if (log.file && log.file.startsWith(originalTempPath)) {
                                    console.log('File is in temp directory, keeping original path');
                                }
                                
                                return log;
                            });
                            
                            allSummary = analyzeData.statistics.type_process_summary || [];
                            
                            // === 修正：處理檔案統計 ===
                            allFileStats = analyzeData.file_statistics || [];
                            allFileStats = allFileStats.map(stat => {
                                // 確保 filepath 存在
                                if (!stat.filepath && stat.file) {
                                    stat.filepath = stat.file;
                                }
                                
                                console.log('=== FileStat Path Debug ===');
                                console.log('Original stat.filepath:', stat.filepath);
                                
                                return stat;
                            });
                            
                            // Reset filters and pagination
                            resetFiltersAndPagination();
                            
                            // === 重要：確保所有路徑相關變數都被正確設置 ===
                            window.basePath = originalTempPath;  // 使用原始臨時路徑
                            window.vpAnalyzeOutputPath = analyzeData.vp_analyze_output_path;
                            window.vpAnalyzeSuccess = analyzeData.vp_analyze_success;
                            window.hasCurrentAnalysis = true;
                            
                            console.log('=== Final Path Configuration ===');
                            console.log('basePath:', window.basePath);
                            console.log('vpAnalyzeOutputPath:', window.vpAnalyzeOutputPath);
                            console.log('vpAnalyzeSuccess:', window.vpAnalyzeSuccess);
                            console.log('Sample log path:', allLogs[0]?.file);
                            console.log('Sample file stat path:', allFileStats[0]?.filepath);
                            
                            // Update UI
                            updateResults(analyzeData);
                            
                            // === 修改：只有在有資料且分析成功時才顯示按鈕 ===
                            if (analyzeData.files_with_cmdline > 0) {
                                // HTML 匯出按鈕只需要有資料就顯示
                                document.getElementById('exportHtmlBtn').style.display = 'block';
                                
                                // 其他按鈕需要 vp_analyze 成功
                                if (analyzeData.vp_analyze_success && analyzeData.vp_analyze_output_path) {
                                    analysisIndexPath = '/view-analysis-report?path=' + encodeURIComponent(analyzeData.vp_analyze_output_path);
                                    const analysisBtn = document.getElementById('analysisResultBtn');
                                    analysisBtn.href = analysisIndexPath;
                                    
                                    // 顯示匯出按鈕
                                    document.getElementById('exportExcelBtn').style.display = 'block';
                                    document.getElementById('exportExcelReportBtn').style.display = 'block';
                                    document.getElementById('mergeExcelBtn').style.display = 'block';
                                    document.getElementById('downloadCurrentZipBtn').style.display = 'block';
                                    
                                    // 自動產生各種格式的檔案
                                    try {
                                        await autoExportExcel(analyzeData.vp_analyze_output_path);
                                        console.log('已自動產生 Excel 檔案');
                                    } catch (error) {
                                        console.error('自動產生 Excel 失敗:', error);
                                    }
                                    
                                    try {
                                        await autoExportHTML(analyzeData.vp_analyze_output_path);
                                        console.log('已自動產生 HTML 檔案');
                                    } catch (error) {
                                        console.error('自動產生 HTML 失敗:', error);
                                    }
                                    
                                    try {
                                        await autoExportExcelReport(analyzeData.vp_analyze_output_path);
                                        console.log('已自動產生 Excel 報表');
                                    } catch (error) {
                                        console.error('自動產生 Excel 報表失敗:', error);
                                    }
                                    
                                    // === 新增：強制重新渲染表格以確保分析報告連結顯示 ===
                                    // 第一次渲染
                                    setTimeout(() => {
                                        console.log('第一次重新渲染表格...');
                                        updateFilesTable();
                                        updateLogsTable();
                                        
                                        // 檢查是否成功添加了分析報告連結
                                        const analyzeLinks = document.querySelectorAll('.analyze-report-link');
                                        console.log('找到的分析報告連結數量:', analyzeLinks.length);
                                        
                                        // 如果沒有找到連結，再次嘗試
                                        if (analyzeLinks.length === 0) {
                                            setTimeout(() => {
                                                console.log('第二次重新渲染表格...');
                                                
                                                // 確保路徑變數仍然正確
                                                window.basePath = originalTempPath;
                                                window.vpAnalyzeOutputPath = analyzeData.vp_analyze_output_path;
                                                
                                                updateFilesTable();
                                                updateLogsTable();
                                                
                                                // 再次檢查
                                                const retryLinks = document.querySelectorAll('.analyze-report-link');
                                                console.log('重試後找到的分析報告連結數量:', retryLinks.length);
                                            }, 1000);
                                        }
                                    }, 500);
                                    
                                    // 額外的除錯：檢查路徑計算
                                    setTimeout(() => {
                                        console.log('=== 路徑計算除錯 ===');
                                        if (allLogs.length > 0) {
                                            const testLog = allLogs[0];
                                            const testPath = testLog.file;
                                            console.log('測試檔案路徑:', testPath);
                                            
                                            // 模擬路徑計算
                                            let relativePath = testPath;
                                            const normalizedBasePath = window.basePath.replace(/\/+$/, '');
                                            const normalizedFilePath = testPath.replace(/\/+$/, '');
                                            
                                            console.log('normalizedBasePath:', normalizedBasePath);
                                            console.log('normalizedFilePath:', normalizedFilePath);
                                            
                                            // 如果檔案路徑包含基礎路徑，則提取相對路徑
                                            if (normalizedFilePath.includes(normalizedBasePath)) {
                                                // 找到基礎路徑的位置
                                                const basePathIndex = normalizedFilePath.indexOf(normalizedBasePath);
                                                console.log('basePathIndex:', basePathIndex);
                                                
                                                // 提取基礎路徑之後的部分
                                                relativePath = normalizedFilePath.substring(basePathIndex + normalizedBasePath.length);
                                                console.log('計算出的相對路徑:', relativePath);
                                            }
                                            
                                            relativePath = relativePath.replace(/^\/+/, '');
                                            const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                                            console.log('預期的分析檔案路徑:', analyzedFilePath);
                                        }
                                    }, 2000);
                                }
                            }
                            
                            let message = `分析完成！共掃描 ${analyzeData.total_files} 個檔案，找到 ${analyzeData.anr_subject_count || 0} 個包含 ANR 的檔案，找到 ${analyzeData.files_with_cmdline - (analyzeData.anr_subject_count || 0)} 個包含 Tombstone 的檔案`;
                            message += `<br>分析耗時: ${analyzeData.analysis_time} 秒`;
                            if (analyzeData.used_grep) {
                                message += '<span class="grep-badge">使用 grep 加速</span>';
                            } else {
                                message += '<span class="grep-badge no-grep-badge">未使用 grep</span>';
                            }
                            
                            // 新增 vp_analyze 狀態訊息
                            if (analyzeData.vp_analyze_success) {
                                message += '<br><span style="color: #28a745;">✓ 詳細分析報告已生成</span>';
                            } else if (analyzeData.vp_analyze_error) {
                                message += `<br><span style="color: #dc3545;">✗ 詳細分析失敗: ${analyzeData.vp_analyze_error}</span>`;
                            }
                            
                            showMessage(message, 'success');
                            
                            // 清除選擇
                            clearMainFileSelection();
                            
                        } else {
                            throw new Error(analyzeData.error || 'Analysis failed');
                        }
                    }
                } else {
                    const error = await response.json();
                    showMessage('準備失敗: ' + (error.error || '未知錯誤'), 'error');
                }
            } catch (error) {
                showMessage('分析失敗: ' + error.message, 'error');
                window.hasCurrentAnalysis = false;
            } finally {
                analyzeBtn.disabled = false;
                analyzeBtn.textContent = '開始分析';
                document.getElementById('loading').style.display = 'none';
            }
        }

        // 在 DOMContentLoaded 中添加主頁面的事件監聽器
        document.addEventListener('DOMContentLoaded', function() {
            // 主頁面檔案選擇按鈕
            const mainSelectFilesBtn = document.getElementById('mainSelectLocalFilesBtn');
            if (mainSelectFilesBtn) {
                mainSelectFilesBtn.addEventListener('click', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('mainFileSelectInput').click();
                });
            }
            
            // 主頁面資料夾選擇按鈕
            const mainSelectFolderBtn = document.getElementById('mainSelectLocalFolderBtn');
            if (mainSelectFolderBtn) {
                mainSelectFolderBtn.addEventListener('click', function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    document.getElementById('mainFolderSelectInput').click();
                });
            }
            
            // 主頁面檔案輸入變化事件
            const mainFileInput = document.getElementById('mainFileSelectInput');
            if (mainFileInput) {
                mainFileInput.addEventListener('change', function(e) {
                    handleMainFileSelect(e.target.files);
                });
            }
            
            // 主頁面資料夾輸入變化事件
            const mainFolderInput = document.getElementById('mainFolderSelectInput');
            if (mainFolderInput) {
                mainFolderInput.addEventListener('change', function(e) {
                    handleMainFolderSelect(e.target.files);
                });
            }
            
            // 主頁面拖曳功能
            const mainDropZone = document.getElementById('mainFileSelectDropZone');
            if (mainDropZone) {
                ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
                    mainDropZone.addEventListener(eventName, preventDefaults, false);
                });
                
                ['dragenter', 'dragover'].forEach(eventName => {
                    mainDropZone.addEventListener(eventName, function() {
                        mainDropZone.classList.add('drag-over');
                    }, false);
                });
                
                ['dragleave', 'drop'].forEach(eventName => {
                    mainDropZone.addEventListener(eventName, function() {
                        mainDropZone.classList.remove('drag-over');
                    }, false);
                });
                
                mainDropZone.addEventListener('drop', function(e) {
                    const dt = e.dataTransfer;
                    const files = dt.files;
                    if (files.length > 0) {
                        handleMainFileSelect(files);
                    }
                }, false);
            }
        });

    </script>
</body>
</html>
'''

@main_page_bp.route('/suggest-path', methods=['POST'])
def suggest_path():
    """Suggest paths based on user input"""
    try:
        data = request.json
        if not data:
            return jsonify({'suggestions': []})
        
        current_path = data.get('path', '')
        
        # Handle empty path - suggest root directories and common paths
        if not current_path:
            suggestions = []
            # Windows drives
            if os.name == 'nt':
                import string
                for drive in string.ascii_uppercase:
                    drive_path = f"{drive}:\\"
                    if os.path.exists(drive_path):
                        suggestions.append(drive_path)
            # Unix/Linux common paths
            else:
                suggestions.append('/')
                suggestions.append('~/')
                if os.path.exists('/home'):
                    suggestions.append('/home/')
                if os.path.exists('/data'):
                    suggestions.append('/data/')
                if os.path.exists('/sdcard'):
                    suggestions.append('/sdcard/')
                if os.path.exists('/storage'):
                    suggestions.append('/storage/')
            
            # Add current directory
            suggestions.append('./')
            suggestions.append('../')
            
            return jsonify({'suggestions': suggestions[:10]})
        
        # Expand user path
        if current_path.startswith('~'):
            current_path = os.path.expanduser(current_path)
        
        # Determine the directory to list and the prefix to match
        if current_path.endswith(os.sep) or current_path.endswith('/'):
            # If path ends with separator, list contents of this directory
            list_dir = current_path
            prefix = ''
        else:
            # Otherwise, list contents of parent directory
            list_dir = os.path.dirname(current_path)
            prefix = os.path.basename(current_path).lower()
        
        suggestions = []
        
        try:
            # Special handling for Windows network paths
            if current_path.startswith('\\\\'):
                # This is a UNC path
                if current_path.count('\\') == 2:
                    # Just \\server, can't list network computers
                    return jsonify({'suggestions': []})
                
            # Check if directory exists
            if os.path.exists(list_dir) and os.path.isdir(list_dir):
                items = os.listdir(list_dir)
                
                for item in items:
                    # Skip hidden files unless user is specifically looking for them
                    if item.startswith('.') and not prefix.startswith('.'):
                        continue
                    
                    # Check if item matches prefix
                    if item.lower().startswith(prefix):
                        full_path = os.path.join(list_dir, item)
                        
                        # Only suggest directories
                        if os.path.isdir(full_path):
                            # Add path separator for directories
                            if not full_path.endswith(os.sep):
                                full_path += os.sep
                            
                            # Check for anr/tombstones folders
                            item_lower = item.lower()
                            if item_lower in ['anr', 'tombstone', 'tombstones']:
                                # Add a marker to indicate this is a target folder
                                suggestions.append(f"{full_path} ⭐")
                            else:
                                # Check if this directory contains anr/tombstones subdirectories
                                try:
                                    subdirs = os.listdir(full_path.rstrip(os.sep))
                                    has_target = any(sub.lower() in ['anr', 'tombstone', 'tombstones'] 
                                                   for sub in subdirs 
                                                   if os.path.isdir(os.path.join(full_path.rstrip(os.sep), sub)))
                                    if has_target:
                                        suggestions.append(f"{full_path} 📁")
                                    else:
                                        suggestions.append(full_path)
                                except:
                                    suggestions.append(full_path)
                
                # Sort suggestions
                suggestions.sort()
                
                # Limit to 20 suggestions
                suggestions = suggestions[:20]
                
        except PermissionError:
            # Can't read directory
            pass
        except Exception as e:
            print(f"Error listing directory: {e}")
        
        return jsonify({'suggestions': suggestions})
        
    except Exception as e:
        print(f"Error in suggest_path: {e}")
        return jsonify({'suggestions': []})
        
@main_page_bp.route('/server-info')
def server_info():
    """Get current server information"""
    # 獲取請求的主機資訊
    host = request.host  # 這會包含 IP:port
    protocol = 'https' if request.is_secure else 'http'
    base_url = f"{protocol}://{host}"
    
    return jsonify({
        'base_url': base_url,
        'host': request.host_url.rstrip('/')
    })
    
@main_page_bp.route('/analyze', methods=['POST'])
def analyze():
    """Analyze logs endpoint"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No JSON data received'}), 400
        
        path = data.get('path', '')
        
        if not path:
            return jsonify({'error': 'Path is required'}), 400

        # === 新增：檢查並獲取分析鎖 ===
        # 使用 session ID 或 IP 作為 owner_id
        owner_id = request.remote_addr  # 或使用 session.get('id') 如果有 session
        
        # 嘗試獲取鎖
        lock_acquired, error_message = analysis_lock_manager.acquire_lock(path, owner_id)
        
        if not lock_acquired:
            return jsonify({
                'error': error_message,
                'locked': True,
                'lock_info': analysis_lock_manager.get_lock_info(path)
            }), 423  # 423 Locked status code
        
        try:
            # Handle Windows UNC paths and regular paths
            original_path = path
            
            # Convert Windows backslashes to forward slashes for consistency
            # But preserve UNC paths starting with \\
            if path.startswith('\\\\'):
                # This is a UNC path, keep it as is
                print(f"UNC path detected: {path}")
            else:
                # Expand user path and convert to absolute
                path = os.path.expanduser(path)
                path = os.path.abspath(path)
            
            # print(f"Analyzing path: {path}")
            # print(f"Original input: {original_path}")
            
            if not os.path.exists(path):
                return jsonify({
                    'error': f'Path does not exist: {path}',
                    'original_path': original_path,
                    'resolved_path': path,
                    'hint': 'For network paths, ensure the share is accessible and mounted'
                }), 400
            
            if not os.path.isdir(path):
                return jsonify({'error': f'Path is not a directory: {path}'}), 400
            
            # 生成包含時間戳和 UUID 的唯一 ID
            analysis_id = datetime.now().strftime('%Y%m%d%H%M%S') + '_' + str(uuid.uuid4())[:8]
                        
            # === 新增：執行 vp_analyze_logs.py ===
            # 取得路徑的最後一個資料夾名稱
            last_folder_name = os.path.basename(path.rstrip(os.sep))
            # 建立輸出目錄名稱
            output_dir_name = f".{last_folder_name}_anr_tombstones_analyze"
            output_path = os.path.join(path, output_dir_name)

            # 修改：備份 all_anr_tombstone_result.xlsx 到上一層目錄
            all_excel_exists = False
            all_excel_path_in_output = os.path.join(output_path, 'all_anr_tombstone_result.xlsx')
            all_excel_backup_path = os.path.join(path, 'all_anr_tombstone_result.xlsx.backup')
            
            # 如果輸出目錄存在
            if os.path.exists(output_path):
                # print(f"發現已存在的輸出目錄: {output_path}")
                
                # 備份 all_anr_tombstone_result.xlsx 到上一層目錄
                if os.path.exists(all_excel_path_in_output):
                    try:
                        # 備份到上一層目錄（使用 .backup 後綴避免衝突）
                        shutil.copy2(all_excel_path_in_output, all_excel_backup_path)
                        all_excel_exists = True
                        # print(f"已備份 all_anr_tombstone_result.xlsx 到: {all_excel_backup_path}")
                    except Exception as e:
                        print(f"備份 all_anr_tombstone_result.xlsx 失敗: {e}")
                        all_excel_exists = False
                
                # 刪除輸出目錄
                try:
                    shutil.rmtree(output_path)
                    # print(f"已刪除舊的輸出目錄: {output_path}")
                except Exception as e:
                    print(f"刪除輸出目錄失敗: {e}")
                    # 如果刪除失敗，清理備份
                    if all_excel_exists and os.path.exists(all_excel_backup_path):
                        try:
                            os.unlink(all_excel_backup_path)
                        except:
                            pass
                    return jsonify({'error': f'無法刪除現有的輸出目錄: {output_path}, 錯誤: {str(e)}'}), 500
            else:
                print(f"輸出目錄不存在，將建立新的: {output_path}")

            # 執行分析
            results = analyzer.analyze_logs(path)

            # 檢查是否有找到資料
            if results.get('files_with_cmdline', 0) == 0:
                # 立即釋放鎖
                analysis_lock_manager.release_lock(path, owner_id)
                
                return jsonify({
                    'analysis_id': analysis_id,
                    'total_files': results['total_files'],
                    'files_with_cmdline': 0,
                    'anr_folders': 0,
                    'tombstone_folders': 0,
                    'statistics': {'type_process_summary': [], 'by_type': {}, 'by_date': {}, 'by_hour': {}},
                    'file_statistics': [],
                    'logs': [],
                    'analysis_time': results.get('analysis_time', 0),
                    'used_grep': results.get('used_grep', False),
                    'anr_subject_count': 0,
                    'message': '沒有找到包含 Cmd line: 或 Cmdline: 的檔案'
                })
                
            # 重要：保存基礎路徑到結果中
            results['path'] = path  # 確保保存原始輸入路徑
            results['base_path'] = path  # 也保存為 base_path
                        
            # 執行 vp_analyze_logs.py
            vp_analyze_success = False
            vp_analyze_error = None
            
            try:
                # 確保 vp_analyze_logs.py 在同一目錄
                vp_script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'vp_analyze_logs.py')
                
                if os.path.exists(vp_script_path):
                    # print(f"找到 vp_analyze_logs.py: {vp_script_path}")
                    # print(f"執行命令: python3.12 {vp_script_path} {path} {output_path}")
                    
                    # 使用 python3.12 執行 vp_analyze_logs.py
                    cmd = ['python3.12', vp_script_path, path, output_path]
                    
                    # 執行命令
                    result = subprocess.run(
                        cmd, 
                        capture_output=True, 
                        text=True, 
                        timeout=1000,
                        cwd=os.path.dirname(vp_script_path)  # 設定工作目錄
                    )
                    
                    # print(f"Return code: {result.returncode}")
                    # print(f"STDOUT: {result.stdout}")
                    # print(f"STDERR: {result.stderr}")
                    
                    if result.returncode == 0:
                        vp_analyze_success = True
                        # print("vp_analyze_logs.py 執行成功")
                        # print(f"分析結果輸出到: {output_path}")
                        
                        # 確保輸出目錄存在
                        if os.path.exists(output_path):
                            # print(f"確認輸出目錄已建立: {output_path}")
                            
                            # 還原備份的 all_anr_tombstone_result.xlsx
                            if all_excel_exists and os.path.exists(all_excel_backup_path):
                                try:
                                    # 將備份檔案移回輸出目錄
                                    shutil.move(all_excel_backup_path, all_excel_path_in_output)
                                    # print(f"已還原 all_anr_tombstone_result.xlsx 到: {all_excel_path_in_output}")
                                except Exception as e:
                                    print(f"還原 all_anr_tombstone_result.xlsx 失敗: {e}")
                                    # 如果移動失敗，嘗試複製
                                    try:
                                        shutil.copy2(all_excel_backup_path, all_excel_path_in_output)
                                        os.unlink(all_excel_backup_path)
                                        # print(f"已複製並還原 all_anr_tombstone_result.xlsx")
                                    except Exception as e2:
                                        print(f"複製 all_anr_tombstone_result.xlsx 也失敗: {e2}")
                            
                            # 列出目錄內容
                            try:
                                files = os.listdir(output_path)
                                print(f"輸出目錄包含 {len(files)} 個檔案")
                                if 'all_anr_tombstone_result.xlsx' in files:
                                    print("確認 all_anr_tombstone_result.xlsx 已在輸出目錄中")
                            except:
                                pass
                        else:
                            print("警告：輸出目錄不存在")
                            vp_analyze_success = False
                            vp_analyze_error = "輸出目錄未建立"
                    else:
                        vp_analyze_error = f"vp_analyze_logs.py 執行失敗 (return code: {result.returncode})\nSTDERR: {result.stderr}\nSTDOUT: {result.stdout}"
                        print(vp_analyze_error)
                else:
                    vp_analyze_error = f"找不到 vp_analyze_logs.py: {vp_script_path}"
                    print(vp_analyze_error)
                    
            except subprocess.TimeoutExpired:
                vp_analyze_error = "vp_analyze_logs.py 執行超時 (超過 300 秒)"
                print(vp_analyze_error)
            except FileNotFoundError:
                vp_analyze_error = "找不到 python3.12 命令，請確認已安裝 Python 3.12"
                print(vp_analyze_error)
            except Exception as e:
                vp_analyze_error = f"執行 vp_analyze_logs.py 時發生錯誤: {str(e)}"
                print(vp_analyze_error)
                import traceback
                traceback.print_exc()
            finally:
                # 清理備份檔案（如果還存在）
                if os.path.exists(all_excel_backup_path):
                    try:
                        os.unlink(all_excel_backup_path)
                        print(f"已清理備份檔案: {all_excel_backup_path}")
                    except Exception as e:
                        print(f"清理備份檔案失敗: {e}")

            # 在每個 log 中添加完整路徑
            for log in results['logs']:
                log['full_path'] = log.get('file', '')
                
                # 從檔案路徑中提取 problem_set
                # 假設路徑格式為 /base/path/問題集名稱/anr或tombstones/檔案名稱
                if 'problem_set' not in log and log.get('file'):
                    try:
                        file_path = log['file']
                        # 移除基礎路徑
                        relative_path = file_path.replace(path, '').lstrip('/')
                        # 取得第一層資料夾名稱作為 problem_set
                        parts = relative_path.split('/')
                        if len(parts) > 1:
                            log['problem_set'] = parts[0]
                        else:
                            log['problem_set'] = '未分類'
                    except:
                        log['problem_set'] = '未分類'
        
            # 在每個 file_stat 中添加完整路徑  
            for file_stat in results['file_statistics']:
                file_stat['full_path'] = file_stat.get('filepath', '')
                
                # 同樣處理 file_statistics 的 problem_set
                if 'problem_set' not in file_stat and file_stat.get('filepath'):
                    try:
                        file_path = file_stat['filepath']
                        relative_path = file_path.replace(path, '').lstrip('/')
                        parts = relative_path.split('/')
                        if len(parts) > 1:
                            file_stat['problem_set'] = parts[0]
                        else:
                            file_stat['problem_set'] = '未分類'
                    except:
                        file_stat['problem_set'] = '未分類'

            # 將分析輸出路徑加入結果中
            results['vp_analyze_output_path'] = output_path if vp_analyze_success else None
            results['vp_analyze_success'] = vp_analyze_success
            results['vp_analyze_error'] = vp_analyze_error
            
            # 檢查 all_anr_tombstone_result.xlsx 是否存在
            results['has_all_excel'] = os.path.exists(all_excel_path_in_output)
            results['all_excel_path'] = all_excel_path_in_output if results['has_all_excel'] else None
            
            # print(f"最終檢查 all_anr_tombstone_result.xlsx: {results['has_all_excel']}")
            
            # 使用新的 cache
            analysis_cache.set(analysis_id, results)
            
            # Return results
            return jsonify({
                'analysis_id': analysis_id,
                'total_files': results['total_files'],
                'files_with_cmdline': results['files_with_cmdline'],
                'anr_folders': results['anr_folders'],
                'tombstone_folders': results['tombstone_folders'],
                'statistics': results['statistics'],
                'file_statistics': results['file_statistics'],
                'logs': results['logs'],
                'analysis_time': results['analysis_time'],
                'used_grep': results['used_grep'],
                'zip_files_extracted': results.get('zip_files_extracted', 0),
                'anr_subject_count': results.get('anr_subject_count', 0),
                'vp_analyze_output_path': results.get('vp_analyze_output_path'),
                'vp_analyze_success': results.get('vp_analyze_success', False),
                'vp_analyze_error': results.get('vp_analyze_error'),
                'has_all_excel': results.get('has_all_excel', False),
                'all_excel_path': results.get('all_excel_path')
            })
        finally:
            # === 重要：確保釋放鎖 ===
            analysis_lock_manager.release_lock(path, owner_id)

    except Exception as e:
        # === 發生異常時也要釋放鎖 ===
        if 'path' in locals():
            analysis_lock_manager.release_lock(path, owner_id)

        print(f"Error in analyze endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def _extract_javascript_functions(self):
    """從 HTML_TEMPLATE 中提取所有 JavaScript 函數"""
    # 這裡應該返回所有必要的 JavaScript 函數
    # 為了簡化，我們可以從 HTML_TEMPLATE 中提取 <script> 標籤內的內容
    scripts = []
    import re
    
    # 找出所有的 script 標籤內容（排除外部引用）
    script_pattern = r'<script>(.+?)</script>'
    matches = re.findall(script_pattern, HTML_TEMPLATE, re.DOTALL)
    
    for match in matches:
        # 排除只有變數定義的部分
        if 'function' in match or 'addEventListener' in match:
            scripts.append(match)
    
    return '\n'.join(scripts)

@main_page_bp.route('/view-analysis-report')
def view_analysis_report():
    """查看 vp_analyze 生成的分析報告"""
    file_path = request.args.get('path')

    if not file_path:
        return """
        <html>
        <body style="font-family: Arial; padding: 20px;">
            <h2>錯誤：未提供檔案路徑</h2>
            <p>請從分析結果頁面點擊「查看詳細分析報告」按鈕。</p>
            <button onclick="window.history.back()">返回</button>
        </body>
        </html>
        """, 400
    
    if not file_path:
        return "No file path provided", 400
    
    # Security check - prevent directory traversal
    if '..' in file_path:
        return "Invalid file path", 403
    
    # Check if file exists
    if not os.path.exists(file_path):
        return f"File not found: {file_path}", 404
    
    # Check if it's a file
    if not os.path.isfile(file_path):
        # 如果是目錄，尋找 index.html
        index_path = os.path.join(file_path, 'index.html')
        if os.path.exists(index_path) and os.path.isfile(index_path):
            file_path = index_path
        else:
            return "Not a file", 400
    
    try:
        # 根據檔案類型返回適當的內容
        if file_path.endswith('.html'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # 修改 HTML 中的相對路徑，使其通過我們的路由
            base_dir = os.path.dirname(file_path)
            
            # 替換相對路徑的連結
            # content = re.sub(
            #    r'(href|src)="(?!http|https|//|#)([^"]+)"',
            #    lambda m: f'{m.group(1)}="/view-file?path={quote(os.path.join(base_dir, m.group(2)))}"',
            #    content
            #)
            
            return Response(content, mimetype='text/html; charset=utf-8')
        
        elif file_path.endswith('.css'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            return Response(content, mimetype='text/css; charset=utf-8')
        
        elif file_path.endswith('.js'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            return Response(content, mimetype='application/javascript; charset=utf-8')
        
        else:
            # 其他檔案類型
            return send_file(file_path)
            
    except Exception as e:
        return f"Error reading file: {str(e)}", 500
        
@main_page_bp.route('/')
def index():
    """Main page"""
    return HTML_TEMPLATE

@main_page_bp.route('/export-ai-csv', methods=['POST'])
def export_ai_csv():
    """Export AI analysis results to CSV with proper encoding"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        base_path = data.get('path')
        analysis_output_path = data.get('analysis_output_path')
        logs = data.get('logs', [])
        
        if not base_path or not analysis_output_path:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # 準備 CSV 資料
        csv_data = []
        sn = 1
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        
        for log in logs:
            # 讀取對應的 AI 分析結果
            ai_result = ""
            if log.get('file'):
                try:
                    # 計算分析檔案路徑
                    file_path = log['file']
                    if file_path.startswith(base_path):
                        relative_path = os.path.relpath(file_path, base_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        # 使用 UTF-8 編碼讀取
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            # 提取可能原因和關鍵堆疊
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            csv_data.append({
                'SN': sn,
                'Date': current_time,
                'Problem Set': log.get('problem_set', '-'),  # 新增問題 set 欄位
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': ai_result,
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('file', '')
            })
            sn += 1
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_anr_tombstone_result.csv"
        
        # 建立 CSV 內容（使用 UTF-8 with BOM）
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['SN', 'Date', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path'])
        writer.writeheader()
        writer.writerows(csv_data)
        
        # 處理 all_anr_tombstone_result.csv
        all_csv_path = os.path.join(os.path.dirname(base_path), 'all_anr_tombstone_result.csv')
        all_csv_updated = False
        
        try:
            if os.path.exists(all_csv_path):
                # 讀取現有的 all_anr_tombstone_result.csv
                existing_data = []
                with open(all_csv_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                    reader = csv.DictReader(f)
                    existing_data = list(reader)
                
                # 更新 SN
                max_sn = max([int(row.get('SN', 0)) for row in existing_data] + [0])
                for row in csv_data:
                    row['SN'] = max_sn + row['SN']
                
                # 合併資料
                all_data = existing_data + csv_data
                
                # 寫入更新後的檔案
                with open(all_csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['SN', 'Date', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path'])
                    writer.writeheader()
                    writer.writerows(all_data)
                
                all_csv_updated = True
            else:
                # 建立新的 all_anr_tombstone_result.csv
                with open(all_csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['SN', 'Date', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path'])
                    writer.writeheader()
                    writer.writerows(csv_data)
                
                all_csv_updated = True
        except Exception as e:
            print(f"處理 all_csv 時發生錯誤: {str(e)}")
        
        # 準備下載（加入 BOM 以確保 Excel 正確識別 UTF-8）
        output_bytes = io.BytesIO()
        # 加入 UTF-8 BOM
        output_bytes.write(b'\xef\xbb\xbf')
        output_bytes.write(output.getvalue().encode('utf-8'))
        output_bytes.seek(0)
        
        response = send_file(
            output_bytes,
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=filename
        )
        
        # 在 response header 中加入更新狀態
        response.headers['X-All-CSV-Updated'] = str(all_csv_updated).lower()
        
        return response
        
    except Exception as e:
        print(f"Error in export_ai_csv: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def extract_ai_summary(content):
    """從 AI 分析內容中提取摘要（可能原因和關鍵堆疊）"""
    if not content:
        return "無分析內容"
    
    summary_parts = []
    
    # 提取可能原因
    patterns = ['可能原因', '可能的原因', '問題原因', 'Possible cause']
    for pattern in patterns:
        if pattern in content:
            start_idx = content.find(pattern)
            if start_idx != -1:
                # 找到下一個段落或結束
                end_markers = ['\n\n', '\n建議', '\n詳細分析', '\n堆疊分析', '\n##']
                end_idx = len(content)
                for marker in end_markers:
                    idx = content.find(marker, start_idx)
                    if idx != -1 and idx < end_idx:
                        end_idx = idx
                
                reason = content[start_idx:end_idx].strip()
                # 清理並簡化內容
                lines = reason.split('\n')
                clean_lines = []
                for line in lines[:4]:  # 只取前4行
                    line = line.strip()
                    if line and not line.startswith('#'):
                        clean_lines.append(line)
                if clean_lines:
                    summary_parts.append(' '.join(clean_lines))
                break
    
    # 提取關鍵堆疊
    patterns = ['關鍵堆疊', '問題堆疊', '重要堆疊', 'Key stack']
    for pattern in patterns:
        if pattern in content:
            start_idx = content.find(pattern)
            if start_idx != -1:
                # 找到下一個段落或結束
                end_markers = ['\n\n', '\n建議', '\n其他', '\n##']
                end_idx = len(content)
                for marker in end_markers:
                    idx = content.find(marker, start_idx)
                    if idx != -1 and idx < end_idx:
                        end_idx = idx
                
                stack = content[start_idx:end_idx].strip()
                # 清理並簡化內容
                lines = stack.split('\n')
                clean_lines = []
                for line in lines[:3]:  # 只取前3行
                    line = line.strip()
                    if line and not line.startswith('#'):
                        clean_lines.append(line)
                if clean_lines:
                    summary_parts.append(' '.join(clean_lines))
                break
    
    # 如果沒有找到特定段落，嘗試提取前幾行有意義的內容
    if not summary_parts:
        lines = content.strip().split('\n')
        meaningful_lines = []
        for line in lines[:10]:  # 檢查前10行
            line = line.strip()
            if line and not line.startswith('#') and len(line) > 10:
                meaningful_lines.append(line)
            if len(meaningful_lines) >= 3:
                break
        if meaningful_lines:
            return ' '.join(meaningful_lines)[:500]
    
    result = ' | '.join(summary_parts)
    # 移除多餘的空白和換行
    result = ' '.join(result.split())
    return result[:500] if result else "無法提取摘要"  # 限制總長度

@main_page_bp.route('/export-ai-excel', methods=['POST'])
def export_ai_excel():
    """Export current AI analysis results to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        base_path = data.get('path')
        analysis_output_path = data.get('analysis_output_path')
        logs = data.get('logs', [])
        
        if not base_path or not analysis_output_path:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # 準備 Excel 資料
        excel_data = []
        sn = 1
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        
        for log in logs:
            # 讀取對應的 AI 分析結果
            ai_result = ""
            if log.get('file'):
                try:
                    file_path = log['file']
                    if file_path.startswith(base_path):
                        relative_path = os.path.relpath(file_path, base_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            excel_data.append({
                'SN': sn,
                'Date': current_time,
                'Problem set': log.get('problem_set', '-'),  # 新增問題 set 欄位
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': ai_result,
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('file', '')
            })
            sn += 1
        
        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "ANR Tombstone Analysis"
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題（加入問題 set）
        headers = ['SN', 'Date', 'Problem set', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # 設定資料樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 寫入資料
        for row_idx, row_data in enumerate(excel_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.font = data_font
                cell.border = data_border
                
                # SN 欄位置中對齊
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                # 根據類型設定背景色（Type 欄位現在是第 4 欄）
                if col_idx == 4:  # Type 欄位
                    if row_data.get('Type') == 'ANR':
                        cell.fill = anr_fill
                    elif row_data.get('Type') == 'Tombstone':
                        cell.fill = tombstone_fill
        
        # 調整欄寬
        column_widths = {
            'A': 8,   # SN
            'B': 20,  # Date
            'C': 20,  # 問題 set
            'D': 12,  # Type
            'E': 30,  # Process
            'F': 60,  # AI result
            'G': 40,  # Filename
            'H': 80   # Folder Path
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 凍結標題列
        ws.freeze_panes = 'A2'
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_anr_tombstone_result.xlsx"
        
        # 儲存到記憶體並回傳
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        # 儲存到分析資料夾
        if analysis_output_path and os.path.exists(analysis_output_path):
            try:
                # 儲存 Excel 檔案到分析資料夾
                excel_save_path = os.path.join(analysis_output_path, 'all_anr_tombstone_result.xlsx')
                wb_copy = Workbook()
                ws_copy = wb_copy.active
                ws_copy.title = "ANR Tombstone Analysis"
                
                # 複製內容
                for row in ws.iter_rows():
                    for cell in row:
                        ws_copy[cell.coordinate].value = cell.value
                        if cell.has_style:
                            ws_copy[cell.coordinate]._style = cell._style
                
                # 複製欄寬
                for col, width in column_widths.items():
                    ws_copy.column_dimensions[col].width = width
                
                # 凍結標題列
                ws_copy.freeze_panes = 'A2'
                
                wb_copy.save(excel_save_path)
                # print(f"已儲存 Excel 檔案到: {excel_save_path}")
            except Exception as e:
                print(f"儲存 Excel 檔案到分析資料夾失敗: {str(e)}")
                        
        return response
        
    except Exception as e:
        print(f"Error in export_ai_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/check-all-excel', methods=['POST'])
def check_all_excel():
    """檢查是否存在 all_anr_tombstone_result.xlsx"""
    try:
        data = request.json
        output_path = data.get('output_path')
        
        if not output_path:
            return jsonify({'exists': False})
        
        all_excel_path = os.path.join(output_path, 'all_anr_tombstone_result.xlsx')
        exists = os.path.exists(all_excel_path)
        
        return jsonify({
            'exists': exists,
            'path': all_excel_path if exists else None
        })
        
    except Exception as e:
        return jsonify({'exists': False, 'error': str(e)})

@main_page_bp.route('/download-file')
def download_file():
    """下載檔案"""
    file_path = request.args.get('path')
    
    if not file_path or not os.path.exists(file_path):
        return "File not found", 404
    
    # 安全檢查
    if '..' in file_path:
        return "Invalid path", 403
    
    return send_file(
        file_path,
        as_attachment=True,
        download_name=os.path.basename(file_path)
    )

@main_page_bp.route('/check-existing-analysis', methods=['POST'])
def check_existing_analysis():
    """檢查是否有已存在的分析結果"""
    try:
        data = request.json
        base_path = data.get('path')
        
        if not base_path or not os.path.exists(base_path):
            return jsonify({'exists': False})
        
        # 尋找分析資料夾
        last_folder_name = os.path.basename(base_path.rstrip(os.sep))
        analysis_folder_name = f".{last_folder_name}_anr_tombstones_analyze"
        analysis_path = os.path.join(base_path, analysis_folder_name)
        
        result = {
            'exists': False,
            'analysis_path': None,
            'has_index': False,
            'has_excel': False,
            'has_html': False,
            'has_excel_report': False,
            'has_folder': False
        }
        
        if os.path.exists(analysis_path) and os.path.isdir(analysis_path):
            result['exists'] = True
            result['analysis_path'] = analysis_path
            result['has_folder'] = True
            
            # 檢查各種檔案
            index_path = os.path.join(analysis_path, 'index.html')
            excel_path = os.path.join(analysis_path, 'all_anr_tombstone_result.xlsx')
            html_path = os.path.join(analysis_path, 'all_anr_tombstone_result.html')
            excel_report_path = os.path.join(analysis_path, 'all_anr_tombstone_excel_result.html')
            
            # 詳細記錄檢查結果
            # print(f"檢查分析資料夾: {analysis_path}")
            # print(f"  index.html 存在: {os.path.exists(index_path)}")
            # print(f"  Excel 存在: {os.path.exists(excel_path)}")
            # print(f"  HTML 存在: {os.path.exists(html_path)}")
            # print(f"  Excel 報表存在: {os.path.exists(excel_report_path)}")
            
            result['has_index'] = os.path.exists(index_path)
            result['has_excel'] = os.path.exists(excel_path)
            result['has_html'] = os.path.exists(html_path)
            result['has_excel_report'] = os.path.exists(excel_report_path)
            
            result['excel_path'] = excel_path if result['has_excel'] else None
            result['html_path'] = html_path if result['has_html'] else None
            result['excel_report_path'] = excel_report_path if result['has_excel_report'] else None
            
            # print(f"回傳結果: {result}")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error checking existing analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'exists': False, 'error': str(e)})

@main_page_bp.route('/export-all-excel-with-current', methods=['POST'])
def export_all_excel_with_current():
    """匯出全部 Excel，可選擇性包含當前分析結果"""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        data = request.json
        all_excel_path = data.get('all_excel_path')
        include_current = data.get('include_current', False)
        
        if not all_excel_path or not os.path.exists(all_excel_path):
            return jsonify({'error': '找不到歷史 Excel 檔案'}), 404
        
        # 讀取現有的 Excel 檔案
        existing_wb = load_workbook(all_excel_path)
        existing_ws = existing_wb.active
        
        # 如果需要包含當前分析結果
        if include_current and data.get('current_data'):
            current_data = data['current_data']
            base_path = current_data.get('path')
            analysis_output_path = current_data.get('analysis_output_path')
            logs = current_data.get('logs', [])
            
            # 準備當前分析的資料
            current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
            new_rows = []
            
            # 獲取現有資料的最大 SN
            max_sn = 0
            for row in existing_ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    try:
                        max_sn = max(max_sn, int(row[0]))
                    except:
                        pass
            
            # 處理當前分析結果
            sn = max_sn + 1
            for log in logs:
                ai_result = ""
                if log.get('file') and analysis_output_path:
                    try:
                        file_path = log['file']
                        if file_path.startswith(base_path):
                            relative_path = os.path.relpath(file_path, base_path)
                        else:
                            relative_path = file_path
                        
                        analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                        
                        if os.path.exists(analyzed_file):
                            with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                                content = f.read()
                                ai_result = extract_ai_summary(content)
                        else:
                            ai_result = "找不到分析結果"
                    except Exception as e:
                        ai_result = f"讀取錯誤: {str(e)}"
                
                new_rows.append([
                    sn,
                    current_time,
                    log.get('problem_set', '-'),  # 新增問題 set 欄位
                    log.get('type', ''),
                    log.get('process', ''),
                    ai_result,
                    log.get('filename', ''),
                    log.get('file', '')
                ])
                sn += 1
            
            # 定義樣式
            data_font = Font(size=11)
            data_alignment = Alignment(vertical="top", wrap_text=True)
            data_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # 將新資料加入工作表
            for row_data in new_rows:
                row_idx = existing_ws.max_row + 1
                for col_idx, value in enumerate(row_data, 1):
                    cell = existing_ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = data_border
                    
                    # SN 欄位置中
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = data_alignment
                    
                    # Type 欄位背景色
                    if col_idx == 4:
                        if value == 'ANR':
                            cell.fill = anr_fill
                        elif value == 'Tombstone':
                            cell.fill = tombstone_fill
            
            # 保存更新後的檔案
            existing_wb.save(all_excel_path)
        
        # 準備下載
        output = io.BytesIO()
        existing_wb.save(output)
        output.seek(0)
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"all_anr_tombstone_result_{date_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in export_all_excel_with_current: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/export-all-history-excel', methods=['POST'])
def export_all_history_excel():
    """匯出歷史 Excel，包含當前新的分析結果，並更新原始檔案"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        data = request.json
        all_excel_path = data.get('all_excel_path')
        include_current = data.get('include_current', False)
        
        if not all_excel_path or not os.path.exists(all_excel_path):
            return jsonify({'error': '找不到歷史 Excel 檔案'}), 404
        
        # 讀取現有的 Excel 檔案
        wb = load_workbook(all_excel_path)
        ws = wb.active
        
        includes_current = False
        records_added = 0  # 新增的記錄數
        
        # 如果需要包含當前分析結果
        if include_current and data.get('current_data'):
            current_data = data['current_data']
            base_path = current_data.get('path')
            analysis_output_path = current_data.get('analysis_output_path')
            logs = current_data.get('logs', [])
            
            # 準備當前分析的資料
            current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
            
            # 獲取現有資料的最大 SN
            max_sn = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    try:
                        max_sn = max(max_sn, int(row[0]))
                    except:
                        pass
            
            # 定義樣式
            data_font = Font(size=11)
            data_alignment = Alignment(vertical="top", wrap_text=True)
            data_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # 處理當前分析結果
            sn = max_sn + 1
            for log in logs:
                ai_result = ""
                if log.get('file') and analysis_output_path:
                    try:
                        file_path = log['file']
                        if file_path.startswith(base_path):
                            relative_path = os.path.relpath(file_path, base_path)
                        else:
                            relative_path = file_path
                        
                        analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                        
                        if os.path.exists(analyzed_file):
                            with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                                content = f.read()
                                ai_result = extract_ai_summary(content)
                        else:
                            ai_result = "找不到分析結果"
                    except Exception as e:
                        ai_result = f"讀取錯誤: {str(e)}"
                
                # 寫入新資料
                row_idx = ws.max_row + 1
                row_data = [
                    sn,
                    current_time,
                    log.get('problem_set', '-'),  # 新增問題 set 欄位
                    log.get('type', ''),
                    log.get('process', ''),
                    ai_result,
                    log.get('filename', ''),
                    log.get('file', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = data_border
                    
                    # SN 欄位置中
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = data_alignment
                    
                    # Type 欄位背景色
                    if col_idx == 4:
                        if value == 'ANR':
                            cell.fill = anr_fill
                        elif value == 'Tombstone':
                            cell.fill = tombstone_fill
                
                sn += 1
                records_added += 1
            
            includes_current = True
        
        # 重要：先保存更新後的原始檔案
        if includes_current:
            try:
                wb.save(all_excel_path)
                # print(f"Updated original all_anr_tombstone_result.xlsx at: {all_excel_path}")
                # print(f"Added {records_added} new records")
            except Exception as e:
                print(f"Failed to update original file: {str(e)}")
        
        # 準備匯出的檔案（帶日期的版本）
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成檔名（加上日期）
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_all_anr_tombstone_result.xlsx"
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        response.headers['X-Includes-Current'] = str(includes_current).lower()
        response.headers['X-Original-Updated'] = str(includes_current).lower()
        response.headers['X-Records-Added'] = str(records_added)
        
        return response
        
    except Exception as e:
        print(f"Error in export_all_history_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/suggest-excel-path', methods=['POST'])
def suggest_excel_path():
    """Suggest Excel file paths based on user input"""
    try:
        data = request.json
        if not data:
            return jsonify({'suggestions': []})
        
        current_path = data.get('path', '')
        
        # 如果沒有輸入，使用當前目錄
        if not current_path:
            current_path = '.'
        
        # 展開用戶路徑
        if current_path.startswith('~'):
            current_path = os.path.expanduser(current_path)
        
        suggestions = []
        
        # 決定要列出的目錄和前綴
        if current_path.endswith(os.sep) or current_path.endswith('/'):
            list_dir = current_path
            prefix = ''
        else:
            list_dir = os.path.dirname(current_path)
            prefix = os.path.basename(current_path).lower()
        
        try:
            # 檢查目錄是否存在
            if os.path.exists(list_dir) and os.path.isdir(list_dir):
                items = os.listdir(list_dir)
                
                for item in items:
                    # 跳過隱藏檔案
                    if item.startswith('.') and not prefix.startswith('.'):
                        continue
                    
                    # 檢查是否符合前綴
                    if item.lower().startswith(prefix):
                        full_path = os.path.join(list_dir, item)
                        
                        # 如果是目錄，添加路徑分隔符並標記
                        if os.path.isdir(full_path):
                            if not full_path.endswith(os.sep):
                                full_path += os.sep
                            suggestions.append(full_path + ' 📁')
                        # 如果是 Excel 檔案
                        elif item.endswith('.xlsx'):
                            suggestions.append(full_path)
            
            # 特別處理：如果是目錄，也搜尋子目錄中的 Excel 檔案
            if os.path.isdir(list_dir) and (not prefix or prefix == ''):
                try:
                    # 限制搜尋深度為 2 層
                    for root, dirs, files in os.walk(list_dir):
                        depth = root[len(list_dir):].count(os.sep)
                        if depth > 2:
                            dirs.clear()  # 不再深入
                            continue
                        
                        for file in files:
                            if file.endswith('.xlsx'):
                                full_path = os.path.join(root, file)
                                suggestions.append(full_path)
                        
                        # 限制結果數量
                        if len(suggestions) > 50:
                            break
                except PermissionError:
                    pass
            
            # 排序：目錄優先，然後是檔案
            dirs = [s for s in suggestions if s.endswith(' 📁')]
            files = [s for s in suggestions if not s.endswith(' 📁')]
            
            suggestions = sorted(dirs) + sorted(files)
            suggestions = suggestions[:30]  # 限制顯示 30 個
            
        except PermissionError:
            pass
        except Exception as e:
            print(f"Error listing directory: {e}")
        
        return jsonify({'suggestions': suggestions})
        
    except Exception as e:
        print(f"Error in suggest_excel_path: {e}")
        return jsonify({'suggestions': []})

@main_page_bp.route('/merge-excel', methods=['POST'])
def merge_excel():
    """Merge current analysis results with another Excel file"""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        data = request.json
        merge_file_path = data.get('merge_file_path')
        current_path = data.get('current_path')
        analysis_output_path = data.get('analysis_output_path')
        logs = data.get('logs', [])
        
        if not merge_file_path or not os.path.exists(merge_file_path):
            return jsonify({'error': '找不到要合併的 Excel 檔案'}), 404
        
        if not merge_file_path.endswith('.xlsx'):
            return jsonify({'error': '只支援 .xlsx 格式的 Excel 檔案'}), 400
        
        # 讀取要合併的 Excel 檔案
        merge_wb = load_workbook(merge_file_path)
        merge_ws = merge_wb.active
        
        # 準備當前分析的資料
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        new_rows = []
        
        # 獲取現有資料的最大 SN
        max_sn = 0
        for row in merge_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    max_sn = max(max_sn, int(row[0]))
                except:
                    pass
        
        # 處理當前分析結果
        sn = max_sn + 1
        for log in logs:
            ai_result = ""
            if log.get('file') and analysis_output_path:
                try:
                    file_path = log['file']
                    if file_path.startswith(current_path):
                        relative_path = os.path.relpath(file_path, current_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            new_rows.append([
                sn,
                current_time,
                log.get('problem_set', '-'),  # 新增問題 set 欄位
                log.get('type', ''),
                log.get('process', ''),
                ai_result,
                log.get('filename', ''),
                log.get('file', '')
            ])
            sn += 1
        
        # 定義樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 將新資料加入工作表
        for row_data in new_rows:
            row_idx = merge_ws.max_row + 1
            for col_idx, value in enumerate(row_data, 1):
                cell = merge_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = data_border
                
                # SN 欄位置中
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                # Type 欄位背景色
                if col_idx == 4:
                    if value == 'ANR':
                        cell.fill = anr_fill
                    elif value == 'Tombstone':
                        cell.fill = tombstone_fill
        
        # 儲存到記憶體並回傳
        output = io.BytesIO()
        merge_wb.save(output)
        output.seek(0)
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_merged_anr_tombstone_result.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in merge_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/merge-excel-upload', methods=['POST'])
def merge_excel_upload():
    """Merge current analysis results with uploaded Excel file"""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 獲取上傳的檔案
        if 'file' not in request.files:
            return jsonify({'error': '沒有上傳檔案'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '沒有選擇檔案'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': '只支援 .xlsx 格式的 Excel 檔案'}), 400
        
        # 獲取其他參數
        current_path = request.form.get('current_path')
        analysis_output_path = request.form.get('analysis_output_path')
        logs = json.loads(request.form.get('logs', '[]'))
        
        # 讀取上傳的 Excel 檔案
        merge_wb = load_workbook(file)
        merge_ws = merge_wb.active
        
        # 準備當前分析的資料
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        
        # 獲取現有資料的最大 SN
        max_sn = 0
        for row in merge_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    max_sn = max(max_sn, int(row[0]))
                except:
                    pass
        
        # 定義樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 處理當前分析結果
        sn = max_sn + 1
        for log in logs:
            ai_result = ""
            if log.get('file') and analysis_output_path:
                try:
                    file_path = log['file']
                    if file_path.startswith(current_path):
                        relative_path = os.path.relpath(file_path, current_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            # 寫入新資料
            row_idx = merge_ws.max_row + 1
            row_data = [
                sn,
                current_time,
                log.get('problem_set', '-'),  # 新增問題 set 欄位
                log.get('type', ''),
                log.get('process', ''),
                ai_result,
                log.get('filename', ''),
                log.get('file', '')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = merge_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = data_border
                
                # SN 欄位置中
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                # Type 欄位背景色
                if col_idx == 4:
                    if value == 'ANR':
                        cell.fill = anr_fill
                    elif value == 'Tombstone':
                        cell.fill = tombstone_fill
            
            sn += 1
        
        # 儲存到記憶體並回傳
        output = io.BytesIO()
        merge_wb.save(output)
        output.seek(0)
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_merged_anr_tombstone_result.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in merge_excel_upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/load-excel-report', methods=['POST'])
def load_excel_report():
    """載入 Excel 並跳轉到報告頁面（支援合併後的檔案）"""
    try:
        import tempfile
        
        # 處理檔案
        excel_path = None
        temp_file = None
        original_filenames = []  # 改為列表儲存所有檔案名稱
        original_paths = []      # 改為列表儲存所有檔案路徑
        is_merged = request.form.get('is_merged') == 'true'
        file_count = request.form.get('file_count', '1')
        
        # 如果是合併的檔案，嘗試從請求中獲取原始檔案資訊
        merged_file_info = request.form.get('merged_file_info')
        if merged_file_info:
            merged_info = json.loads(merged_file_info)
            original_filenames = merged_info.get('filenames', [])
            original_paths = merged_info.get('paths', [])
        
        if 'file' in request.files:
            # 上傳的檔案
            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': '沒有選擇檔案'}), 400
            
            if not file.filename.endswith('.xlsx'):
                return jsonify({'error': '只支援 .xlsx 格式的 Excel 檔案'}), 400
            
            # 根據是否為合併檔案設定檔名
            if is_merged and original_filenames:
                # 使用傳遞過來的原始檔案名稱
                display_filename = f"合併 {len(original_filenames)} 個 Excel 檔案"
                display_path = original_filenames  # 傳遞檔案名稱列表
            else:
                display_filename = file.filename
                display_path = f"本地上傳: {file.filename}"
                original_filenames = [file.filename]
                original_paths = [f"本地上傳: {file.filename}"]
            
            # 儲存到暫存檔案
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            file.save(temp_file.name)
            excel_path = temp_file.name
            
        elif 'file_path' in request.form:
            # 伺服器路徑
            excel_path = request.form.get('file_path')
            if not os.path.exists(excel_path):
                return jsonify({'error': '檔案不存在'}), 404
            
            display_filename = os.path.basename(excel_path)
            display_path = excel_path
            original_filenames = [display_filename]
            original_paths = [excel_path]
        else:
            return jsonify({'error': '未提供檔案'}), 400
        
        # 將檔案路徑存入 session 或生成唯一 ID
        import uuid
        report_id = str(uuid.uuid4())
        
        # 使用 analysis_cache 儲存檔案資訊
        analysis_cache.set(f"excel_report_{report_id}", {
            'excel_path': excel_path,
            'is_temp': temp_file is not None,
            'original_filenames': original_filenames,  # 儲存所有檔案名稱
            'original_paths': original_paths,          # 儲存所有檔案路徑
            'is_merged': is_merged,
            'file_count': len(original_filenames)      # 使用實際檔案數量
        })
        
        # 返回報告 URL
        return jsonify({
            'report_url': f'/excel-report/{report_id}'
        })
        
    except Exception as e:
        print(f"Error in load_excel_report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/export-excel-report', methods=['POST'])
def export_excel_report():
    """將分析結果匯出為 Excel 報表格式"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # 使用者選擇的路徑
        path = data.get('path')
        # 分析 ID
        analysis_id = data.get('analysis_id')
        
        if not path or not analysis_id:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # 從快取中取得分析結果
        analysis_data = analysis_cache.get(analysis_id)
        if not analysis_data:
            return jsonify({'error': 'Analysis data not found'}), 404
        
        # 生成唯一的報表 ID
        report_id = str(uuid.uuid4())[:8]
        
        # 建立臨時 Excel 檔案
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 準備資料
        logs = analysis_data.get('logs', [])
        
        # 生成檔案名稱
        date_str = datetime.now().strftime('%Y%m%d')
        original_filename = f"{date_str}_anr_tombstone_result.xlsx"
        
        # 準備資料
        excel_data = []
        for idx, log in enumerate(logs, 1):
            excel_data.append({
                'SN': idx,
                'Date': log.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                'Problem set': log.get('problem_set', '-'),
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': '-',
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('folder_path', '')
            })
        
        # 轉換為 DataFrame
        df = pd.DataFrame(excel_data)
        
        # 建立臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
            
            # 將 DataFrame 寫入 Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='ANR Tombstone Analysis', index=False)
                
                # 美化 Excel
                workbook = writer.book
                worksheet = writer.sheets['ANR Tombstone Analysis']
                
                # 設定欄寬
                column_widths = {
                    'A': 8,   # SN
                    'B': 20,  # Date
                    'C': 20,  # Problem set
                    'D': 12,  # Type
                    'E': 35,  # Process
                    'F': 30,  # AI result
                    'G': 40,  # Filename
                    'H': 60   # Folder Path
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 凍結標題列
                worksheet.freeze_panes = 'A2'
        
        # 儲存檔案資訊到快取
        file_info = {
            'excel_path': excel_path,
            'is_temp': True,
            'original_filenames': [original_filename],  # 確保是列表格式
            'original_paths': [f"分析結果: {path}"],   # 顯示分析的路徑
            'is_merged': False,
            'file_count': 1
        }
        
        analysis_cache.set(f"excel_report_{report_id}", file_info)
        
        # 返回報表頁面的 URL
        return jsonify({
            'success': True,
            'report_url': f'/excel-report/{report_id}'
        })
        
    except Exception as e:
        print(f"Error in export_excel_report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/download-analysis-zip', methods=['POST'])
def download_analysis_zip():
    """打包並下載分析結果"""
    try:
        import zipfile
        import tempfile
        
        data = request.json
        analysis_path = data.get('analysis_path')
        
        if not analysis_path or not os.path.exists(analysis_path):
            return jsonify({'error': '分析資料夾不存在'}), 404
        
        # 建立暫存 zip 檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp_file:
            zip_path = tmp_file.name
        
        # 打包資料夾
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(analysis_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, os.path.dirname(analysis_path))
                    zipf.write(file_path, arcname)
        
        # 生成檔名
        folder_name = os.path.basename(analysis_path.rstrip(os.sep))
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_all_anr_tombstone_analysis_result.zip"
        
        # 讀取並返回 zip 檔案
        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in download_analysis_zip: {str(e)}")
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/export/<format>/<analysis_id>')
def export(format, analysis_id):
    # 使用 LimitedCache 的 get 方法
    data = analysis_cache.get(analysis_id)

    if data is None:
        return jsonify({'error': 'Analysis not found or expired'}), 404

    if format == 'html':
        # 從 URL 參數獲取 base_url
        base_url = request.args.get('base_url', '')
        if not base_url:
            base_url = f"{request.scheme}://{request.host}"
        
        # 獲取基礎路徑 - 這裡需要確保正確獲取
        # 修正：從分析數據中獲取原始輸入路徑
        base_path = ''
        
        # 嘗試從多個來源獲取基礎路徑
        if data.get('path'):
            base_path = data.get('path')
        elif data.get('base_path'):
            base_path = data.get('base_path')
        elif data.get('logs') and len(data.get('logs', [])) > 0:
            # 從第一個 log 的檔案路徑推斷基礎路徑
            first_log = data['logs'][0]
            if first_log.get('file'):
                file_path = first_log['file']
                # 嘗試從分析輸出路徑推斷
                if data.get('vp_analyze_output_path'):
                    output_path = data['vp_analyze_output_path']
                    # 分析輸出路徑格式：/base/path/.foldername_anr_tombstones_analyze
                    # 提取基礎路徑
                    if '_anr_tombstones_analyze' in output_path:
                        base_path = os.path.dirname(output_path)
        
        # print(f"Export HTML - base_path: {base_path}")  # 調試輸出
        
        # 創建 HTML 報告
        html_report = HTML_TEMPLATE
        
        # 在注入的腳本中修改檔案連結並自動載入資料
        static_script = f'''
<script>
    // 靜態頁面標記
    window.isStaticExport = true;
    window.exportBaseUrl = "{base_url}";
    window.basePath = "{base_path}";
    
    console.log('Initial basePath set to:', window.basePath);  // 調試輸出
    
    // 覆寫檔案連結點擊行為 - 讓連結能正常工作
    window.addEventListener('click', function(e) {{
        if (e.target.classList.contains('file-link') && !e.target.classList.contains('analyze-report-link')) {{
            e.preventDefault();
            const href = e.target.getAttribute('href');
            if (href) {{
                // 如果是相對路徑，加上 base URL
                const fullUrl = href.startsWith('http') ? href : window.exportBaseUrl + href;
                window.open(fullUrl, '_blank');
            }}
            return false;
        }}
    }}, true);
</script>
'''
        
        # Inject the data directly into the HTML
        script_injection = static_script + f'''
<script>
    // 設定 base URL
    window.exportBaseUrl = "{base_url}";
    window.basePath = "{base_path}";
    
    // 保存分析報告相關資訊
    window.vpAnalyzeOutputPath = {json.dumps(data.get('vp_analyze_output_path'))};
    window.vpAnalyzeSuccess = {json.dumps(data.get('vp_analyze_success', False))};
    
    // 修改表格更新函數以使用完整 URL
    const originalUpdateFilesTable = window.updateFilesTable;
    const originalUpdateLogsTable = window.updateLogsTable;
    
    window.updateFilesTable = function() {{
        if (originalUpdateFilesTable) {{
            originalUpdateFilesTable.call(this);
        }}
        
        // 替換所有檔案連結為完整連結
        document.querySelectorAll('#filesTableBody .file-link').forEach(link => {{
            const href = link.getAttribute('href');
            if (href && !href.startsWith('http')) {{
                link.setAttribute('href', window.exportBaseUrl + href);
                link.setAttribute('target', '_blank');
            }}
        }});
        
        // 添加分析報告連結
        if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {{
            // 從資料中取得檔案資訊
            const startIndex = (filesPage - 1) * itemsPerPage;
            const pageData = filteredFiles.slice(startIndex, startIndex + itemsPerPage);
            
            document.querySelectorAll('#filesTableBody tr').forEach((row, index) => {{
                const fileCell = row.cells[5];
                if (fileCell && !fileCell.querySelector('.analyze-report-link') && index < pageData.length) {{
                    const fileData = pageData[index];
                    const filePath = fileData.filepath || fileData.file;
                    
                    if (filePath) {{
                        console.log('=== Files Table Debug Info ===');
                        console.log('Original filePath:', filePath);
                        console.log('basePath:', window.basePath);
                        console.log('vpAnalyzeOutputPath:', window.vpAnalyzeOutputPath);
                        
                        // 修正：確保正確提取相對路徑
                        let relativePath = filePath;
                        
                        // 正規化路徑（移除尾部斜線）
                        const normalizedBasePath = window.basePath.replace(/\\/+$/, '');
                        const normalizedFilePath = filePath.replace(/\\/+$/, '');
                        
                        console.log('normalizedBasePath:', normalizedBasePath);
                        console.log('normalizedFilePath:', normalizedFilePath);
                        
                        // 如果檔案路徑包含基礎路徑，則提取相對路徑
                        if (normalizedFilePath.includes(normalizedBasePath)) {{
                            // 找到基礎路徑的位置
                            const basePathIndex = normalizedFilePath.indexOf(normalizedBasePath);
                            console.log('basePathIndex:', basePathIndex);
                            
                            // 提取基礎路徑之後的部分
                            relativePath = normalizedFilePath.substring(basePathIndex + normalizedBasePath.length);
                            console.log('relativePath after substring:', relativePath);
                        }}
                        
                        // 移除開頭的斜線
                        relativePath = relativePath.replace(/^\\/+/, '');
                        console.log('relativePath after removing leading slash:', relativePath);
                        
                        // 建立分析報告路徑
                        const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                        console.log('Final analyzedFilePath:', analyzedFilePath);
                        
                        // 建立分析報告連結
                        const analyzeReportUrl = window.exportBaseUrl + '/view-file?path=' + encodeURIComponent(analyzedFilePath);
                        console.log('Final analyzeReportUrl:', analyzeReportUrl);
                        console.log('=== End Debug Info ===\\n');
                        
                        const analyzeLink = document.createElement('a');
                        analyzeLink.href = analyzeReportUrl;
                        analyzeLink.target = '_blank';
                        analyzeLink.className = 'file-link analyze-report-link';
                        analyzeLink.textContent = ' (分析報告)';
                        analyzeLink.style.cssText = 'color: #28a745; font-size: 0.9em;';
                        fileCell.appendChild(analyzeLink);
                    }}
                }}
            }});
        }}
    }};
    
    window.updateLogsTable = function() {{
        if (originalUpdateLogsTable) {{
            originalUpdateLogsTable.call(this);
        }}
        
        // 替換所有檔案連結為完整連結
        document.querySelectorAll('#logsTableBody .file-link').forEach(link => {{
            const href = link.getAttribute('href');
            if (href && !href.startsWith('http')) {{
                link.setAttribute('href', window.exportBaseUrl + href);
                link.setAttribute('target', '_blank');
            }}
        }});
        
        // 添加分析報告連結
        if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {{
            // 從資料中取得檔案資訊
            const startIndex = (logsPage - 1) * itemsPerPage;
            const pageData = filteredLogs.slice(startIndex, startIndex + itemsPerPage);
            
            document.querySelectorAll('#logsTableBody tr').forEach((row, index) => {{
                const fileCell = row.cells[6];
                if (fileCell && !fileCell.querySelector('.analyze-report-link') && index < pageData.length) {{
                    const logData = pageData[index];
                    const filePath = logData.file;
                    
                    if (filePath) {{
                        console.log('=== Logs Table Debug Info ===');
                        console.log('Original filePath:', filePath);
                        console.log('basePath:', window.basePath);
                        console.log('vpAnalyzeOutputPath:', window.vpAnalyzeOutputPath);
                        
                        // 修正：確保正確提取相對路徑
                        let relativePath = filePath;
                        
                        // 正規化路徑（移除尾部斜線）
                        const normalizedBasePath = window.basePath.replace(/\\/+$/, '');
                        const normalizedFilePath = filePath.replace(/\\/+$/, '');
                        
                        console.log('normalizedBasePath:', normalizedBasePath);
                        console.log('normalizedFilePath:', normalizedFilePath);
                        
                        // 如果檔案路徑包含基礎路徑，則提取相對路徑
                        if (normalizedFilePath.includes(normalizedBasePath)) {{
                            // 找到基礎路徑的位置
                            const basePathIndex = normalizedFilePath.indexOf(normalizedBasePath);
                            console.log('basePathIndex:', basePathIndex);
                            
                            // 提取基礎路徑之後的部分
                            relativePath = normalizedFilePath.substring(basePathIndex + normalizedBasePath.length);
                            console.log('relativePath after substring:', relativePath);
                        }}
                        
                        // 移除開頭的斜線
                        relativePath = relativePath.replace(/^\\/+/, '');
                        console.log('relativePath after removing leading slash:', relativePath);
                        
                        // 建立分析報告路徑
                        const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                        console.log('Final analyzedFilePath:', analyzedFilePath);
                        
                        // 建立分析報告連結
                        const analyzeReportUrl = window.exportBaseUrl + '/view-file?path=' + encodeURIComponent(analyzedFilePath);
                        console.log('Final analyzeReportUrl:', analyzeReportUrl);
                        console.log('=== End Debug Info ===\\n');
                        
                        const analyzeLink = document.createElement('a');
                        analyzeLink.href = analyzeReportUrl;
                        analyzeLink.target = '_blank';
                        analyzeLink.className = 'file-link analyze-report-link';
                        analyzeLink.textContent = ' (分析報告)';
                        analyzeLink.style.cssText = 'color: #28a745; font-size: 0.9em;';
                        fileCell.appendChild(analyzeLink);
                    }}
                }}
            }});
        }}
    }};
    
    // Injected analysis data
    window.injectedData = {json.dumps({
        'analysis_id': data.get('analysis_id', analysis_id),
        'total_files': data['total_files'],
        'files_with_cmdline': data['files_with_cmdline'],
        'anr_folders': data['anr_folders'],
        'tombstone_folders': data['tombstone_folders'],
        'statistics': data['statistics'],
        'file_statistics': data['file_statistics'],
        'logs': data['logs'],
        'analysis_time': data['analysis_time'],
        'used_grep': data['used_grep'],
        'zip_files_extracted': data.get('zip_files_extracted', 0),
        'anr_subject_count': data.get('anr_subject_count', 0),
        'vp_analyze_output_path': data.get('vp_analyze_output_path'),
        'vp_analyze_success': data.get('vp_analyze_success', False)
    })};
    
    // Auto-load the data when page loads
    window.addEventListener('DOMContentLoaded', function() {{
        // === 重要：立即隱藏控制面板並顯示結果 ===
        document.querySelector('.control-panel').style.display = 'none';
        document.getElementById('results').style.display = 'block';
        
        // 隱藏歷史區塊
        const historySection = document.getElementById('historySection');
        if (historySection) historySection.style.display = 'none';
        
        // 隱藏不需要的按鈕（但保留分析結果按鈕）
        document.getElementById('exportHtmlBtn').style.display = 'none';
        ['exportExcelBtn', 'exportExcelReportBtn', 'mergeExcelBtn', 'downloadCurrentZipBtn'].forEach(id => {{
            const btn = document.getElementById(id);
            if (btn) btn.style.display = 'none';
        }});
        
        // === 重要：設定並顯示分析結果按鈕 ===
        if (window.injectedData.vp_analyze_success && window.injectedData.vp_analyze_output_path) {{
            const analysisBtn = document.getElementById('analysisResultBtn');
            if (analysisBtn) {{
                // 設定連結到完整的分析報告
                const reportUrl = window.exportBaseUrl + '/view-analysis-report?path=' + encodeURIComponent(window.injectedData.vp_analyze_output_path);
                analysisBtn.href = reportUrl;
                analysisBtn.target = '_blank';
                
                // 確保按鈕可點擊
                analysisBtn.onclick = null; // 移除任何 onclick 處理
            }}
        }}
        
        // 載入資料
        currentAnalysisId = window.injectedData.analysis_id;
        allLogs = window.injectedData.logs.sort((a, b) => {{
            if (!a.timestamp && !b.timestamp) return 0;
            if (!a.timestamp) return 1;
            if (!b.timestamp) return -1;
            return a.timestamp.localeCompare(b.timestamp);
        }});
        allSummary = window.injectedData.statistics.type_process_summary || [];
        allFileStats = window.injectedData.file_statistics || [];
        
        // 生成程序統計資料
        const processOnlyData = {{}};
        window.injectedData.statistics.type_process_summary.forEach(item => {{
            if (!processOnlyData[item.process]) {{
                processOnlyData[item.process] = {{
                    count: 0,
                    problem_sets: new Set()
                }};
            }}
            processOnlyData[item.process].count += item.count;
            
            if (item.problem_sets && Array.isArray(item.problem_sets)) {{
                item.problem_sets.forEach(set => {{
                    processOnlyData[item.process].problem_sets.add(set);
                }});
            }}
        }});
        
        allProcessSummary = Object.entries(processOnlyData)
            .map(([process, data]) => ({{ 
                process, 
                count: data.count,
                problem_sets: Array.from(data.problem_sets).sort()
            }}))
            .sort((a, b) => b.count - a.count);
        
        // Reset filters and pagination
        resetFiltersAndPagination();
        
        // Update UI - 這會顯示所有的統計圖表和表格
        updateResults(window.injectedData);
        
        // 顯示導覽相關按鈕
        setTimeout(() => {{
            const navToggleBtn = document.getElementById('navToggleBtn');
            if (navToggleBtn) navToggleBtn.classList.add('show');
            
            const globalToggleBtn = document.getElementById('globalToggleBtn');
            if (globalToggleBtn) globalToggleBtn.classList.add('show');
            
            const backToTopBtn = document.getElementById('backToTop');
            if (backToTopBtn) backToTopBtn.classList.add('show');
            
            // === 重要：確保分析結果按鈕在滾動時也保持顯示 ===
            const analysisResultBtn = document.getElementById('analysisResultBtn');
            if (analysisResultBtn && window.injectedData.vp_analyze_success) {{
                analysisResultBtn.classList.add('show');
            }}
        }}, 300);
        
        // === 新增：監聽滾動事件，確保分析結果按鈕保持顯示 ===
        window.addEventListener('scroll', function() {{
            const analysisResultBtn = document.getElementById('analysisResultBtn');
            if (analysisResultBtn && window.injectedData.vp_analyze_success) {{
                // 只有在滾動超過 300px 時才顯示
                if (window.pageYOffset > 300) {{
                    analysisResultBtn.classList.add('show');
                }} else {{
                    analysisResultBtn.classList.remove('show');
                }}
            }}
        }});
        
        // Show analysis info message
        let message = `分析完成！共掃描 ${{window.injectedData.total_files}} 個檔案，找到 ${{window.injectedData.anr_subject_count || 0}} 個包含 ANR 的檔案，找到 ${{window.injectedData.files_with_cmdline - (window.injectedData.anr_subject_count || 0)}} 個包含 Tombstone 的檔案`;
        message += `<br>分析耗時: ${{window.injectedData.analysis_time}} 秒`;
        if (window.injectedData.used_grep) {{
            message += '<span class="grep-badge">使用 grep 加速</span>';
        }} else {{
            message += '<span class="grep-badge no-grep-badge">未使用 grep</span>';
        }}
        
        // 添加 vp_analyze 狀態訊息
        if (window.injectedData.vp_analyze_success) {{
            message += '<br><span style="color: #28a745;">✓ 詳細分析報告已生成</span>';
        }}
        
        message += `<br><br>報告生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`;
        
        const infoDiv = document.createElement('div');
        infoDiv.className = 'success';
        infoDiv.innerHTML = message;
        document.querySelector('.header').appendChild(infoDiv);
        
        // 確保導覽列顯示
        document.getElementById('navBar').classList.remove('show');
    }});
</script>
'''

        # Insert the script before closing body tag
        html_report = html_report.replace('</body>', script_injection + '</body>')
        
        # Convert to bytes
        output_bytes = io.BytesIO()
        output_bytes.write(html_report.encode('utf-8'))
        output_bytes.seek(0)
        
        # 生成日期字串
        date_str = datetime.now().strftime('%Y%m%d')
        
        # 儲存 HTML 到分析資料夾
        if data.get('vp_analyze_output_path') and os.path.exists(data.get('vp_analyze_output_path')):
            try:
                html_save_path = os.path.join(data.get('vp_analyze_output_path'), 'all_anr_tombstone_result.html')
                with open(html_save_path, 'w', encoding='utf-8') as f:
                    f.write(html_report)
                # print(f"已儲存 HTML 檔案到: {html_save_path}")
            except Exception as e:
                print(f"儲存 HTML 檔案到分析資料夾失敗: {str(e)}")
                        
        return send_file(
            output_bytes,
            mimetype='text/html',
            as_attachment=True,
            download_name=f'{date_str}_anr_tombstone_result.html'
        )
    
    else:
        return jsonify({'error': 'Invalid format'}), 400        

@main_page_bp.route('/export-excel-to-folder', methods=['POST'])
def export_excel_to_folder():
    """匯出 Excel 到指定資料夾（不下載）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        base_path = data.get('path')
        analysis_output_path = data.get('analysis_output_path')
        output_folder = data.get('output_folder')
        logs = data.get('logs', [])
        
        if not base_path or not analysis_output_path or not output_folder:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # 準備 Excel 資料
        excel_data = []
        sn = 1
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        
        for log in logs:
            # 讀取對應的 AI 分析結果
            ai_result = ""
            if log.get('file'):
                try:
                    file_path = log['file']
                    if file_path.startswith(base_path):
                        relative_path = os.path.relpath(file_path, base_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            excel_data.append({
                'SN': sn,
                'Date': current_time,
                'Problem set': log.get('problem_set', '-'),
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': ai_result,
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('file', '')
            })
            sn += 1
        
        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "ANR Tombstone Analysis"
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題
        headers = ['SN', 'Date', 'Problem set', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # 設定資料樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 寫入資料
        for row_idx, row_data in enumerate(excel_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.font = data_font
                cell.border = data_border
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                if col_idx == 4:  # Type 欄位
                    if row_data.get('Type') == 'ANR':
                        cell.fill = anr_fill
                    elif row_data.get('Type') == 'Tombstone':
                        cell.fill = tombstone_fill
        
        # 調整欄寬
        column_widths = {
            'A': 8,   # SN
            'B': 20,  # Date
            'C': 20,  # 問題 set
            'D': 12,  # Type
            'E': 30,  # Process
            'F': 60,  # AI result
            'G': 40,  # Filename
            'H': 80   # Folder Path
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 凍結標題列
        ws.freeze_panes = 'A2'
        
        # 儲存到指定資料夾
        excel_save_path = os.path.join(output_folder, 'all_anr_tombstone_result.xlsx')
        wb.save(excel_save_path)
        
        # print(f"已儲存 Excel 檔案到: {excel_save_path}")
        
        return jsonify({
            'success': True,
            'saved_path': excel_save_path
        })
        
    except Exception as e:
        print(f"Error in export_excel_to_folder: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/view-analysis-html')
def view_analysis_html():
    """查看分析產生的 HTML 報告"""
    file_path = request.args.get('path')
    
    if not file_path:
        return """
        <html>
        <body style="font-family: Arial; padding: 20px;">
            <h2>錯誤：未提供檔案路徑</h2>
            <p>請從分析結果頁面點擊「已統計分析」按鈕。</p>
            <button onclick="window.history.back()">返回</button>
        </body>
        </html>
        """, 400
    
    # Security check - prevent directory traversal
    if '..' in file_path:
        return "Invalid file path", 403
    
    # Check if file exists
    if not os.path.exists(file_path):
        return f"File not found: {file_path}", 404
    
    # Check if it's an HTML file
    if not file_path.endswith('.html'):
        return "Not an HTML file", 400
    
    try:
        # 讀取 HTML 檔案
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # 修改 HTML 中的相對路徑
        base_dir = os.path.dirname(file_path)
        
        # 注入一個腳本來處理檔案連結
        inject_script = """
        <script>
        // 確保所有檔案連結都能正常運作
        window.addEventListener('DOMContentLoaded', function() {
            document.querySelectorAll('.file-link').forEach(link => {
                const href = link.getAttribute('href');
                if (href && href.startsWith('/view-file')) {
                    // 保持原有的連結格式
                    link.setAttribute('target', '_blank');
                }
            });
        });
        </script>
        """
        
        # 在 </body> 前插入腳本
        content = content.replace('</body>', inject_script + '</body>')
        
        return Response(content, mimetype='text/html; charset=utf-8')
        
    except Exception as e:
        return f"Error reading file: {str(e)}", 500

@main_page_bp.route('/export-excel-report-to-folder', methods=['POST'])
def export_excel_report_to_folder():
    """將 Excel 報表儲存到指定資料夾"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        analysis_id = data.get('analysis_id')
        output_folder = data.get('output_folder')
        path = data.get('path')
        
        if not analysis_id or not output_folder:
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # 從快取獲取分析資料
        analysis_data = analysis_cache.get(analysis_id)
        if not analysis_data:
            return jsonify({'error': 'Analysis data not found'}), 404
        
        # 準備資料
        logs = analysis_data.get('logs', [])
        
        # 轉換為 Excel 報表格式的資料
        excel_data = []
        sn = 1
        for log in logs:
            excel_data.append({
                'SN': sn,
                'Date': log.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                'Problem set': log.get('problem_set', '-'),  # 改為英文，與模板一致
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': '-',
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('folder_path', '')
            })
            sn += 1
        
        # 建立 DataFrame
        df = pd.DataFrame(excel_data)
        
        # 如果沒有資料，補充預設值
        if df.empty:
            excel_data = [{
                'SN': 1,
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Problem set': '-',
                'Type': '-',
                'Process': '-',
                'AI result': '-',
                'Filename': '-',
                'Folder Path': '-'
            }]
            df = pd.DataFrame(excel_data)
        
        # 生成唯一的報表 ID
        report_id = str(uuid.uuid4())[:8]
        
        # 建立臨時 Excel 檔案
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
            
            # 將 DataFrame 寫入 Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='ANR Tombstone Analysis', index=False)
                
                # 美化 Excel
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                workbook = writer.book
                worksheet = writer.sheets['ANR Tombstone Analysis']
                
                # 設定欄寬
                column_widths = {
                    'A': 8,   # SN
                    'B': 20,  # Date
                    'C': 20,  # Problem set
                    'D': 12,  # Type
                    'E': 35,  # Process
                    'F': 30,  # AI result
                    'G': 40,  # Filename
                    'H': 60   # Folder Path
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 凍結標題列
                worksheet.freeze_panes = 'A2'
        
        # 將檔案資訊存入快取
        file_info = {
            'excel_path': excel_path,
            'is_temp': True,
            'original_filename': os.path.basename(output_folder),
            'original_path': path
        }
        
        analysis_cache.set(f"excel_report_{report_id}", file_info)
        
        # 從 routes/excel_report.py 複製模板內容
        from routes.excel_report import EXCEL_REPORT_TEMPLATE
        
        # 準備模板資料 - 重要：傳入正確格式的資料
        template_data = {
            'filename': os.path.basename(output_folder),
            'filepath': path,
            'load_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data': json.dumps(excel_data),  # 傳入格式化後的 excel_data
            'excel_data_base64': '',  # Excel 報表不需要內嵌檔案
            # 新增這兩個必要的變數
            'filename_list': [os.path.basename(output_folder)],  # 改為列表格式
            'path_list': [path]  # 改為列表格式
        }
        
        # 生成 HTML 內容
        html_content = EXCEL_REPORT_TEMPLATE
        
        # 替換模板變數
        for key, value in template_data.items():
            if key in ['data', 'filename_list', 'path_list']:
                # 這些需要特殊處理
                if key == 'data':
                    html_content = html_content.replace('{{ data | tojson }}', value)
                elif key == 'filename_list':
                    html_content = html_content.replace('{{ filename_list | tojson }}', json.dumps(value))
                elif key == 'path_list':
                    html_content = html_content.replace('{{ path_list | tojson }}', json.dumps(value))
            else:
                # 替換其他變數
                html_content = html_content.replace(f'{{{{ {key} }}}}', str(value))
        
        # 儲存到檔案
        report_save_path = os.path.join(output_folder, 'all_anr_tombstone_excel_result.html')
        with open(report_save_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        # print(f"已儲存 Excel 報表到: {report_save_path}")
        
        # 清理臨時檔案
        try:
            os.unlink(excel_path)
        except:
            pass
        
        return jsonify({
            'success': True,
            'saved_path': report_save_path
        })
        
    except Exception as e:
        print(f"Error in export_excel_report_to_folder: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/merge-multiple-excel', methods=['POST'])
def merge_multiple_excel():
    """合併多個 Excel 檔案，可選擇性包含當前分析結果"""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 檢查是否有分析結果
        has_analysis = request.form.get('has_analysis') == 'true'
        
        # 獲取檔案相關參數
        file_paths = json.loads(request.form.get('file_paths', '[]'))
        uploaded_files = request.files.getlist('files')
        
        # 如果有分析結果，獲取分析相關參數
        current_path = None
        analysis_output_path = None
        logs = []
        
        if has_analysis:
            current_path = request.form.get('current_path')
            analysis_output_path = request.form.get('analysis_output_path')
            logs = json.loads(request.form.get('logs', '[]'))
        
        # 檢查是否有任何檔案
        if len(uploaded_files) + len(file_paths) == 0:
            return jsonify({'error': '沒有選擇任何檔案'}), 400
        
        # 建立新的工作簿
        merged_wb = Workbook()
        merged_ws = merged_wb.active
        merged_ws.title = "ANR Tombstone Analysis"
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題
        headers = ['SN', 'Date', 'Problem set', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path']
        for col, header in enumerate(headers, 1):
            cell = merged_ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # 準備樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 合併所有檔案的資料
        all_rows = []
        
        # 處理上傳的檔案
        for file in uploaded_files:
            if file.filename.endswith('.xlsx'):
                try:
                    wb = load_workbook(file)
                    ws = wb.active
                    
                    # 檢查標題列是否匹配（處理不同格式的 Excel）
                    headers_row = None
                    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                        if row and 'SN' in str(row[0]):
                            headers_row = row
                            break
                    
                    # 找到標題列後，從下一列開始讀取資料
                    if headers_row:
                        start_row = None
                        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                            if row == headers_row:
                                start_row = idx + 1
                                break
                        
                        if start_row:
                            for row in ws.iter_rows(min_row=start_row, values_only=True):
                                if row[0] is not None:  # 檢查是否有資料
                                    # 確保有 8 個欄位
                                    row_list = list(row)[:8]
                                    while len(row_list) < 8:
                                        row_list.append('')
                                    all_rows.append(row_list)
                    else:
                        # 如果找不到標題，假設從第二列開始是資料
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] is not None:
                                row_list = list(row)[:8]
                                while len(row_list) < 8:
                                    row_list.append('')
                                all_rows.append(row_list)
                except Exception as e:
                    print(f"處理檔案 {file.filename} 時發生錯誤: {str(e)}")
                    continue
        
        # 處理伺服器路徑的檔案
        for file_path in file_paths:
            if os.path.exists(file_path) and file_path.endswith('.xlsx'):
                try:
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # 同樣的標題列檢查邏輯
                    headers_row = None
                    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                        if row and 'SN' in str(row[0]):
                            headers_row = row
                            break
                    
                    if headers_row:
                        start_row = None
                        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                            if row == headers_row:
                                start_row = idx + 1
                                break
                        
                        if start_row:
                            for row in ws.iter_rows(min_row=start_row, values_only=True):
                                if row[0] is not None:
                                    row_list = list(row)[:8]
                                    while len(row_list) < 8:
                                        row_list.append('')
                                    all_rows.append(row_list)
                    else:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] is not None:
                                row_list = list(row)[:8]
                                while len(row_list) < 8:
                                    row_list.append('')
                                all_rows.append(row_list)
                except Exception as e:
                    print(f"處理檔案 {file_path} 時發生錯誤: {str(e)}")
                    continue
        
        # 按日期排序（假設日期在第二欄）
        try:
            all_rows.sort(key=lambda x: str(x[1]) if x[1] else '')
        except:
            pass  # 如果排序失敗，保持原順序
        
        # 重新編號並寫入資料
        for sn, row_data in enumerate(all_rows, 1):
            row_data[0] = sn  # 更新 SN
            row_idx = merged_ws.max_row + 1
            for col_idx, value in enumerate(row_data, 1):
                cell = merged_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = data_border
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                # Type 欄位背景色（第 4 欄）
                if col_idx == 4 and value:
                    if str(value).upper() == 'ANR':
                        cell.fill = anr_fill
                    elif str(value).upper() == 'TOMBSTONE':
                        cell.fill = tombstone_fill
        
        # 如果有分析結果，加入當前分析的資料
        if has_analysis and logs:
            current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
            max_sn = len(all_rows)
            
            for log in logs:
                ai_result = ""
                if log.get('file') and analysis_output_path:
                    try:
                        file_path = log['file']
                        if file_path.startswith(current_path):
                            relative_path = os.path.relpath(file_path, current_path)
                        else:
                            relative_path = file_path
                        
                        analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                        
                        if os.path.exists(analyzed_file):
                            with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                                content = f.read()
                                ai_result = extract_ai_summary(content)
                        else:
                            ai_result = "找不到分析結果"
                    except Exception as e:
                        ai_result = f"讀取錯誤: {str(e)}"
                
                max_sn += 1
                row_idx = merged_ws.max_row + 1
                row_data = [
                    max_sn,
                    current_time,
                    log.get('problem_set', '-'),
                    log.get('type', ''),
                    log.get('process', ''),
                    ai_result,
                    log.get('filename', ''),
                    log.get('file', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = merged_ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = data_border
                    
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = data_alignment
                    
                    if col_idx == 4:
                        if value == 'ANR':
                            cell.fill = anr_fill
                        elif value == 'Tombstone':
                            cell.fill = tombstone_fill
        
        # 設定欄寬
        column_widths = {
            'A': 8,   # SN
            'B': 20,  # Date
            'C': 20,  # 問題 set
            'D': 12,  # Type
            'E': 30,  # Process
            'F': 60,  # AI result
            'G': 40,  # Filename
            'H': 80   # Folder Path
        }
        
        for col, width in column_widths.items():
            merged_ws.column_dimensions[col].width = width
        
        # 凍結標題列
        merged_ws.freeze_panes = 'A2'
        
        # 儲存到記憶體並回傳
        output = io.BytesIO()
        merged_wb.save(output)
        output.seek(0)
        
        # 生成檔名
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"{date_str}_merged_anr_tombstone_result.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in merge_multiple_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/export-html-to-folder/<analysis_id>', methods=['POST'])
def export_html_to_folder(analysis_id):
    """將 HTML 報告儲存到指定資料夾"""
    try:
        data = request.json
        output_path = data.get('output_path')
        base_url = data.get('base_url', '')
        
        if not output_path:
            return jsonify({'error': 'Missing output path'}), 400
        
        # 從快取獲取分析資料
        analysis_data = analysis_cache.get(analysis_id)
        if not analysis_data:
            return jsonify({'error': 'Analysis data not found'}), 404
        
        # 獲取基礎路徑
        base_path = ''
        if analysis_data.get('path'):
            base_path = analysis_data.get('path')
        elif analysis_data.get('base_path'):
            base_path = analysis_data.get('base_path')
        
        if not base_url:
            base_url = f"{request.scheme}://{request.host}"
        
        # 創建 HTML 報告（使用與 export 函數相同的邏輯）
        html_report = HTML_TEMPLATE
        
        # 在注入的腳本中修改檔案連結並自動載入資料
        static_script = f'''
<script>
    // 靜態頁面標記
    window.isStaticExport = true;
    window.exportBaseUrl = "{base_url}";
    window.basePath = "{base_path}";
    
    console.log('Initial basePath set to:', window.basePath);
    
    // 覆寫檔案連結點擊行為
    window.addEventListener('click', function(e) {{
        if (e.target.classList.contains('file-link') && !e.target.classList.contains('analyze-report-link')) {{
            e.preventDefault();
            const href = e.target.getAttribute('href');
            if (href) {{
                const fullUrl = href.startsWith('http') ? href : window.exportBaseUrl + href;
                window.open(fullUrl, '_blank');
            }}
            return false;
        }}
    }}, true);
</script>
'''
        
        # Inject the data directly into the HTML
        script_injection = static_script + f'''
<script>
    // 設定 base URL
    window.exportBaseUrl = "{base_url}";
    window.basePath = "{base_path}";
    
    // 保存分析報告相關資訊
    window.vpAnalyzeOutputPath = {json.dumps(analysis_data.get('vp_analyze_output_path'))};
    window.vpAnalyzeSuccess = {json.dumps(analysis_data.get('vp_analyze_success', False))};
    
    // 修改表格更新函數以使用完整 URL
    const originalUpdateFilesTable = window.updateFilesTable;
    const originalUpdateLogsTable = window.updateLogsTable;
    
    window.updateFilesTable = function() {{
        if (originalUpdateFilesTable) {{
            originalUpdateFilesTable.call(this);
        }}
        
        // 替換所有檔案連結為完整連結
        document.querySelectorAll('#filesTableBody .file-link').forEach(link => {{
            const href = link.getAttribute('href');
            if (href && !href.startsWith('http')) {{
                link.setAttribute('href', window.exportBaseUrl + href);
                link.setAttribute('target', '_blank');
            }}
        }});
        
        // 添加分析報告連結
        if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {{
            const startIndex = (filesPage - 1) * itemsPerPage;
            const pageData = filteredFiles.slice(startIndex, startIndex + itemsPerPage);
            
            document.querySelectorAll('#filesTableBody tr').forEach((row, index) => {{
                const fileCell = row.cells[5];
                if (fileCell && !fileCell.querySelector('.analyze-report-link') && index < pageData.length) {{
                    const fileData = pageData[index];
                    const filePath = fileData.filepath || fileData.file;
                    
                    if (filePath) {{
                        let relativePath = filePath;
                        const normalizedBasePath = window.basePath.replace(/\\/+$/, '');
                        const normalizedFilePath = filePath.replace(/\\/+$/, '');
                        
                        if (normalizedFilePath.includes(normalizedBasePath)) {{
                            const basePathIndex = normalizedFilePath.indexOf(normalizedBasePath);
                            relativePath = normalizedFilePath.substring(basePathIndex + normalizedBasePath.length);
                        }}
                        
                        relativePath = relativePath.replace(/^\\/+/, '');
                        const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                        const analyzeReportUrl = window.exportBaseUrl + '/view-file?path=' + encodeURIComponent(analyzedFilePath);
                        
                        const analyzeLink = document.createElement('a');
                        analyzeLink.href = analyzeReportUrl;
                        analyzeLink.target = '_blank';
                        analyzeLink.className = 'file-link analyze-report-link';
                        analyzeLink.textContent = ' (分析報告)';
                        analyzeLink.style.cssText = 'color: #28a745; font-size: 0.9em;';
                        fileCell.appendChild(analyzeLink);
                    }}
                }}
            }});
        }}
    }};
    
    window.updateLogsTable = function() {{
        if (originalUpdateLogsTable) {{
            originalUpdateLogsTable.call(this);
        }}
        
        // 替換所有檔案連結為完整連結
        document.querySelectorAll('#logsTableBody .file-link').forEach(link => {{
            const href = link.getAttribute('href');
            if (href && !href.startsWith('http')) {{
                link.setAttribute('href', window.exportBaseUrl + href);
                link.setAttribute('target', '_blank');
            }}
        }});
        
        // 添加分析報告連結
        if (window.vpAnalyzeOutputPath && window.vpAnalyzeSuccess) {{
            const startIndex = (logsPage - 1) * itemsPerPage;
            const pageData = filteredLogs.slice(startIndex, startIndex + itemsPerPage);
            
            document.querySelectorAll('#logsTableBody tr').forEach((row, index) => {{
                const fileCell = row.cells[6];
                if (fileCell && !fileCell.querySelector('.analyze-report-link') && index < pageData.length) {{
                    const logData = pageData[index];
                    const filePath = logData.file;
                    
                    if (filePath) {{
                        let relativePath = filePath;
                        const normalizedBasePath = window.basePath.replace(/\\/+$/, '');
                        const normalizedFilePath = filePath.replace(/\\/+$/, '');
                        
                        if (normalizedFilePath.includes(normalizedBasePath)) {{
                            const basePathIndex = normalizedFilePath.indexOf(normalizedBasePath);
                            relativePath = normalizedFilePath.substring(basePathIndex + normalizedBasePath.length);
                        }}
                        
                        relativePath = relativePath.replace(/^\\/+/, '');
                        const analyzedFilePath = window.vpAnalyzeOutputPath + '/' + relativePath + '.analyzed.txt';
                        const analyzeReportUrl = window.exportBaseUrl + '/view-file?path=' + encodeURIComponent(analyzedFilePath);
                        
                        const analyzeLink = document.createElement('a');
                        analyzeLink.href = analyzeReportUrl;
                        analyzeLink.target = '_blank';
                        analyzeLink.className = 'file-link analyze-report-link';
                        analyzeLink.textContent = ' (分析報告)';
                        analyzeLink.style.cssText = 'color: #28a745; font-size: 0.9em;';
                        fileCell.appendChild(analyzeLink);
                    }}
                }}
            }});
        }}
    }};
    
    // Injected analysis data
    window.injectedData = {json.dumps({
        'analysis_id': analysis_data.get('analysis_id', analysis_id),
        'total_files': analysis_data['total_files'],
        'files_with_cmdline': analysis_data['files_with_cmdline'],
        'anr_folders': analysis_data['anr_folders'],
        'tombstone_folders': analysis_data['tombstone_folders'],
        'statistics': analysis_data['statistics'],
        'file_statistics': analysis_data['file_statistics'],
        'logs': analysis_data['logs'],
        'analysis_time': analysis_data['analysis_time'],
        'used_grep': analysis_data['used_grep'],
        'zip_files_extracted': analysis_data.get('zip_files_extracted', 0),
        'anr_subject_count': analysis_data.get('anr_subject_count', 0),
        'vp_analyze_output_path': analysis_data.get('vp_analyze_output_path'),
        'vp_analyze_success': analysis_data.get('vp_analyze_success', False)
    })};
    
    // Auto-load the data when page loads
    window.addEventListener('DOMContentLoaded', function() {{
        // 隱藏控制面板並顯示結果
        document.querySelector('.control-panel').style.display = 'none';
        document.getElementById('results').style.display = 'block';
        
        // 隱藏歷史區塊
        const historySection = document.getElementById('historySection');
        if (historySection) historySection.style.display = 'none';
        
        // 隱藏不需要的按鈕
        document.getElementById('exportHtmlBtn').style.display = 'none';
        ['exportExcelBtn', 'exportExcelReportBtn', 'mergeExcelBtn', 'downloadCurrentZipBtn'].forEach(id => {{
            const btn = document.getElementById(id);
            if (btn) btn.style.display = 'none';
        }});
        
        // 設定並顯示分析結果按鈕
        if (window.injectedData.vp_analyze_success && window.injectedData.vp_analyze_output_path) {{
            const analysisBtn = document.getElementById('analysisResultBtn');
            if (analysisBtn) {{
                const reportUrl = window.exportBaseUrl + '/view-analysis-report?path=' + encodeURIComponent(window.injectedData.vp_analyze_output_path);
                analysisBtn.href = reportUrl;
                analysisBtn.target = '_blank';
                analysisBtn.onclick = null;
            }}
        }}
        
        // 載入資料
        currentAnalysisId = window.injectedData.analysis_id;
        allLogs = window.injectedData.logs.sort((a, b) => {{
            if (!a.timestamp && !b.timestamp) return 0;
            if (!a.timestamp) return 1;
            if (!b.timestamp) return -1;
            return a.timestamp.localeCompare(b.timestamp);
        }});
        allSummary = window.injectedData.statistics.type_process_summary || [];
        allFileStats = window.injectedData.file_statistics || [];
        
        // 生成程序統計資料
        const processOnlyData = {{}};
        window.injectedData.statistics.type_process_summary.forEach(item => {{
            if (!processOnlyData[item.process]) {{
                processOnlyData[item.process] = {{
                    count: 0,
                    problem_sets: new Set()
                }};
            }}
            processOnlyData[item.process].count += item.count;
            
            if (item.problem_sets && Array.isArray(item.problem_sets)) {{
                item.problem_sets.forEach(set => {{
                    processOnlyData[item.process].problem_sets.add(set);
                }});
            }}
        }});
        
        allProcessSummary = Object.entries(processOnlyData)
            .map(([process, data]) => ({{ 
                process, 
                count: data.count,
                problem_sets: Array.from(data.problem_sets).sort()
            }}))
            .sort((a, b) => b.count - a.count);
        
        // Reset filters and pagination
        resetFiltersAndPagination();
        
        // Update UI
        updateResults(window.injectedData);
        
        // 顯示導覽相關按鈕
        setTimeout(() => {{
            const navToggleBtn = document.getElementById('navToggleBtn');
            if (navToggleBtn) navToggleBtn.classList.add('show');
            
            const globalToggleBtn = document.getElementById('globalToggleBtn');
            if (globalToggleBtn) globalToggleBtn.classList.add('show');
            
            const backToTopBtn = document.getElementById('backToTop');
            if (backToTopBtn) backToTopBtn.classList.add('show');
            
            const analysisResultBtn = document.getElementById('analysisResultBtn');
            if (analysisResultBtn && window.injectedData.vp_analyze_success) {{
                analysisResultBtn.classList.add('show');
            }}
        }}, 300);
        
        // 監聽滾動事件
        window.addEventListener('scroll', function() {{
            const analysisResultBtn = document.getElementById('analysisResultBtn');
            if (analysisResultBtn && window.injectedData.vp_analyze_success) {{
                if (window.pageYOffset > 300) {{
                    analysisResultBtn.classList.add('show');
                }} else {{
                    analysisResultBtn.classList.remove('show');
                }}
            }}
        }});
        
        // Show analysis info message
        let message = `分析完成！共掃描 ${{window.injectedData.total_files}} 個檔案，找到 ${{window.injectedData.anr_subject_count || 0}} 個包含 ANR 的檔案，找到 ${{window.injectedData.files_with_cmdline - (window.injectedData.anr_subject_count || 0)}} 個包含 Tombstone 的檔案`;
        message += `<br>分析耗時: ${{window.injectedData.analysis_time}} 秒`;
        if (window.injectedData.used_grep) {{
            message += '<span class="grep-badge">使用 grep 加速</span>';
        }} else {{
            message += '<span class="grep-badge no-grep-badge">未使用 grep</span>';
        }}
        
        if (window.injectedData.vp_analyze_success) {{
            message += '<br><span style="color: #28a745;">✓ 詳細分析報告已生成</span>';
        }}
        
        message += `<br><br>報告生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`;
        
        const infoDiv = document.createElement('div');
        infoDiv.className = 'success';
        infoDiv.innerHTML = message;
        document.querySelector('.header').appendChild(infoDiv);
        
        document.getElementById('navBar').classList.remove('show');
    }});
</script>
'''

        # Insert the script before closing body tag
        html_report = html_report.replace('</body>', script_injection + '</body>')
        
        # 儲存 HTML 到分析資料夾
        html_save_path = os.path.join(output_path, 'all_anr_tombstone_result.html')
        with open(html_save_path, 'w', encoding='utf-8') as f:
            f.write(html_report)
        
        # print(f"已儲存 HTML 檔案到: {html_save_path}")
        
        return jsonify({
            'success': True,
            'saved_path': html_save_path
        })
        
    except Exception as e:
        print(f"Error in export_html_to_folder: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main_page_bp.route('/analyze-selected-items', methods=['POST'])
def analyze_selected_items():
    """處理選擇的檔案和資料夾，準備分析"""
    try:
        # 每次分析前清理舊的臨時檔案
        cleanup_old_temp_dirs()

        import tempfile
        import shutil
        import zipfile
        
        # 建立臨時目錄
        temp_dir = tempfile.mkdtemp(prefix='anr_analysis_')
        
        # 將臨時目錄加入追蹤集合
        temp_directories.add(temp_dir)
        
        # 處理上傳的檔案
        files = request.files.getlist('files')
        folder_files = request.files.getlist('folder_files')
        auto_group = request.form.get('auto_group') == 'true'
        
        # print(f"接收到 {len(files)} 個單獨檔案")
        # print(f"接收到 {len(folder_files)} 個資料夾檔案")
        # print(f"自動分組: {auto_group}")
        
        # === 修改：簡化排除邏輯 - 排除所有隱藏資料夾 ===
        def should_exclude_path(path):
            """檢查路徑是否應該被排除"""
            path_parts = path.split('/')
            for part in path_parts:
                # 排除隱藏資料夾（以 . 開頭的資料夾）
                if part.startswith('.'):
                    return True
            return False
        
        # 用於分組的計數器
        group_counter = 1
        anr_files = []
        tombstone_files = []
        zip_files_to_extract = []
        
        # === 新增：追蹤是否需要強制建立資料夾結構 ===
        needs_folder_structure = False
        has_anr_files = False
        has_tombstone_files = False
        
        # 處理單獨上傳的檔案
        for file in files:
            filename = file.filename
            file_lower = filename.lower()
            
            # === 新增：檢查是否為 ZIP 檔案 ===
            if filename.endswith('.zip'):
                temp_file_path = os.path.join(temp_dir, filename)
                file.save(temp_file_path)
                zip_files_to_extract.append(temp_file_path)
                continue
            
            # 檢查是否為 ANR 或 Tombstone 檔案
            is_anr_file = 'anr' in file_lower
            is_tombstone_file = 'tombstone' in file_lower
            
            if is_anr_file or is_tombstone_file:
                if is_anr_file:
                    has_anr_files = True
                else:
                    has_tombstone_files = True
                
                if auto_group:
                    # 自動分組：儲存到臨時目錄，稍後移動
                    temp_file_path = os.path.join(temp_dir, filename)
                    file.save(temp_file_path)
                    
                    if is_anr_file:
                        anr_files.append((temp_file_path, filename))
                    else:
                        tombstone_files.append((temp_file_path, filename))
                else:
                    # === 修改：不自動分組時，也建立適當的資料夾結構 ===
                    needs_folder_structure = True
                    if is_anr_file:
                        # 放入 anr 資料夾
                        anr_folder = os.path.join(temp_dir, 'anr')
                        os.makedirs(anr_folder, exist_ok=True)
                        file.save(os.path.join(anr_folder, filename))
                    else:
                        # 放入 tombstones 資料夾
                        tombstones_folder = os.path.join(temp_dir, 'tombstones')
                        os.makedirs(tombstones_folder, exist_ok=True)
                        file.save(os.path.join(tombstones_folder, filename))
            else:
                # 其他檔案直接儲存
                file.save(os.path.join(temp_dir, filename))
        
        # 處理資料夾中的檔案
        for file in folder_files:
            # 保持原始路徑結構
            relative_path = file.filename
            
            # === 新增：檢查是否應該排除此路徑 ===
            if should_exclude_path(relative_path):
                print(f"排除路徑: {relative_path}")
                continue
            
            file_path = os.path.join(temp_dir, relative_path)
            
            # 建立目錄
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 儲存檔案
            file.save(file_path)
            
            # === 新增：檢查是否為 ZIP 檔案 ===
            if file_path.endswith('.zip'):
                zip_files_to_extract.append(file_path)
                continue
            
            # 檢查是否需要分組（只對根目錄的檔案）
            path_parts = relative_path.split('/')
            filename = os.path.basename(relative_path)
            file_lower = filename.lower()
            
            # === 修改：即使不在根目錄，也檢查是否為 ANR/Tombstone 檔案 ===
            if 'anr' in file_lower or 'tombstone' in file_lower:
                if 'anr' in file_lower:
                    has_anr_files = True
                else:
                    has_tombstone_files = True
                    
                # 如果路徑中沒有 anr/ 或 tombstones/ 資料夾，且不是自動分組
                folder_lower = relative_path.lower()
                if not auto_group and '/anr/' not in folder_lower and '/tombstones/' not in folder_lower and '/tombstone/' not in folder_lower:
                    needs_folder_structure = True
                    
                if auto_group and len(path_parts) <= 2:  # 只對淺層檔案進行分組
                    if 'anr' in file_lower:
                        anr_files.append((file_path, filename))
                    else:
                        tombstone_files.append((file_path, filename))
        
        # === 新增：解壓所有 ZIP 檔案 ===
        for zip_path in zip_files_to_extract:
            try:
                print(f"解壓 ZIP 檔案: {zip_path}")
                extract_dir = os.path.splitext(zip_path)[0]  # 移除 .zip 副檔名
                
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    # 取得所有檔案列表
                    all_files = zip_ref.namelist()
                    
                    # 過濾並解壓檔案
                    for file_info in all_files:
                        # 檢查是否應該排除
                        if should_exclude_path(file_info):
                            print(f"解壓時排除: {file_info}")
                            continue
                        
                        # 解壓檔案
                        zip_ref.extract(file_info, extract_dir)
                        
                        # 檢查是否為 ANR 或 Tombstone 檔案
                        file_lower = file_info.lower()
                        if 'anr' in file_lower or 'tombstone' in file_lower:
                            if 'anr' in file_lower:
                                has_anr_files = True
                            else:
                                has_tombstone_files = True
                                
                            full_path = os.path.join(extract_dir, file_info)
                            if os.path.isfile(full_path):
                                filename = os.path.basename(file_info)
                                
                                # 檢查路徑結構
                                if not auto_group and '/anr/' not in file_lower and '/tombstones/' not in file_lower and '/tombstone/' not in file_lower:
                                    needs_folder_structure = True
                                
                                if auto_group:
                                    if 'anr' in file_lower:
                                        anr_files.append((full_path, filename))
                                    else:
                                        tombstone_files.append((full_path, filename))
                
                # 刪除原始 ZIP 檔案
                os.remove(zip_path)
                
            except Exception as e:
                print(f"解壓 ZIP 檔案失敗 {zip_path}: {str(e)}")
        
        # 如果啟用自動分組，將獨立的 ANR/Tombstone 檔案放入群組資料夾
        if auto_group:
            # 處理 ANR 檔案
            if anr_files:
                for file_path, filename in anr_files:
                    group_folder = os.path.join(temp_dir, f'Group{group_counter}', 'anr')
                    os.makedirs(group_folder, exist_ok=True)
                    
                    # 如果檔案存在，移動它
                    if os.path.exists(file_path):
                        shutil.move(file_path, os.path.join(group_folder, filename))
                        group_counter += 1
            
            # 處理 Tombstone 檔案
            if tombstone_files:
                for file_path, filename in tombstone_files:
                    group_folder = os.path.join(temp_dir, f'Group{group_counter}', 'tombstones')
                    os.makedirs(group_folder, exist_ok=True)
                    
                    # 如果檔案存在，移動它
                    if os.path.exists(file_path):
                        shutil.move(file_path, os.path.join(group_folder, filename))
                        group_counter += 1
        
        # === 新增：如果沒有自動分組，且需要建立資料夾結構 ===
        elif needs_folder_structure and not auto_group:
            # print("檢測到需要建立 anr/tombstones 資料夾結構")
            
            # 移動所有散落的 ANR/Tombstone 檔案到正確的資料夾
            for root, dirs, files in os.walk(temp_dir):
                for filename in files:
                    file_lower = filename.lower()
                    file_path = os.path.join(root, filename)
                    
                    # 跳過已經在正確資料夾中的檔案
                    if '/anr/' in root or '/tombstones/' in root or '/tombstone/' in root:
                        continue
                    
                    if 'anr' in file_lower:
                        # 移動到 anr 資料夾
                        anr_folder = os.path.join(temp_dir, 'anr')
                        os.makedirs(anr_folder, exist_ok=True)
                        dest_path = os.path.join(anr_folder, filename)
                        if not os.path.exists(dest_path):
                            shutil.move(file_path, dest_path)
                            # print(f"移動 ANR 檔案: {filename} -> anr/")
                    elif 'tombstone' in file_lower:
                        # 移動到 tombstones 資料夾
                        tombstones_folder = os.path.join(temp_dir, 'tombstones')
                        os.makedirs(tombstones_folder, exist_ok=True)
                        dest_path = os.path.join(tombstones_folder, filename)
                        if not os.path.exists(dest_path):
                            shutil.move(file_path, dest_path)
                            # print(f"移動 Tombstone 檔案: {filename} -> tombstones/")
        
        # === 新增：清理空目錄 ===
        def remove_empty_dirs(root_dir):
            """遞迴刪除空目錄"""
            for dirpath, dirnames, filenames in os.walk(root_dir, topdown=False):
                # 如果目錄為空，刪除它
                if not dirnames and not filenames and dirpath != root_dir:
                    try:
                        os.rmdir(dirpath)
                        # print(f"刪除空目錄: {dirpath}")
                    except:
                        pass
        
        remove_empty_dirs(temp_dir)
        
        # 返回臨時目錄路徑
        total_files = len(files) + len(folder_files)
        extracted_files = len(zip_files_to_extract)
        
        return jsonify({
            'success': True,
            'temp_path': temp_dir,
            'message': f'已準備 {total_files} 個檔案，解壓了 {extracted_files} 個 ZIP 檔案'
        })
        
    except Exception as e:
        print(f"Error in analyze_selected_items: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# 全域變數追蹤臨時目錄
temp_directories = set()

def cleanup_temp_dirs():
    """清理所有臨時目錄"""
    for temp_dir in temp_directories:
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                print(f"已清理臨時目錄: {temp_dir}")
        except Exception as e:
            print(f"清理臨時目錄失敗 {temp_dir}: {e}")

# 註冊清理函數
atexit.register(cleanup_temp_dirs)

def cleanup_old_temp_dirs():
    """清理超過 24 小時的臨時目錄"""
    import glob
    temp_base = tempfile.gettempdir()
    pattern = os.path.join(temp_base, 'anr_analysis_*')
    
    for temp_dir in glob.glob(pattern):
        try:
            # 檢查目錄修改時間
            mtime = os.path.getmtime(temp_dir)
            age_hours = (time.time() - mtime) / 3600
            
            if age_hours > 24:
                shutil.rmtree(temp_dir)
                print(f"已清理舊的臨時目錄: {temp_dir}")
        except Exception as e:
            print(f"清理舊臨時目錄失敗 {temp_dir}: {e}")

@main_page_bp.route('/check-analysis-lock', methods=['POST'])
def check_analysis_lock():
    """檢查分析路徑是否被鎖定"""
    try:
        data = request.json
        path = data.get('path', '')
        
        if not path:
            return jsonify({'locked': False})
        
        is_locked = analysis_lock_manager.is_locked(path)
        lock_info = analysis_lock_manager.get_lock_info(path) if is_locked else None
        
        response_data = {'locked': is_locked}
        
        if lock_info:
            elapsed_time = (datetime.now() - lock_info['start_time']).total_seconds()
            remaining_time = max(0, analysis_lock_manager._lock_timeout - elapsed_time)
            
            response_data.update({
                'owner': lock_info['owner'],
                'elapsed_time': int(elapsed_time),
                'remaining_time': int(remaining_time)
            })
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error checking lock: {str(e)}")
        return jsonify({'locked': False, 'error': str(e)}), 500
