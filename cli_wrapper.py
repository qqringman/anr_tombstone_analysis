#!/usr/bin/env python3.12
"""
Android ANR/Tombstone Analyzer CLI Wrapper

Usage:
    python3.12 cli_wrapper.py -input file1.zip,file2.txt,/path/to/folder -output result.zip
"""

import argparse
import os
import sys
import tempfile
import shutil
import zipfile
import json
import subprocess
from datetime import datetime
from pathlib import Path
import io
# 導入 JIRA 相關模組
from jira.jira_config import JiraConfig
from jira.jira_client import JiraClient
from jira.jira_file_manager import JiraFileManager

# 將當前目錄加入 Python 路徑
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 導入分析器和必要的模組
from routes.grep_analyzer import AndroidLogAnalyzer
from routes.main_page import extract_ai_summary, HTML_TEMPLATE

def parse_arguments():
    """解析命令列參數"""
    parser = argparse.ArgumentParser(
        description='Android ANR/Tombstone Analyzer CLI',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
    # 分析單一檔案
    python3.12 cli_wrapper.py -i anr.txt -o result.zip
    
    # 分析多個檔案和資料夾
    python3.12 cli_wrapper.py -i anr.zip,tombstone.txt,/logs/folder -o result.zip
    
    # 自動分組獨立檔案
    python3.12 cli_wrapper.py -i anr1.txt,anr2.txt -o result.zip --auto-group
    
    # 只從 JIRA 下載並分析
    python3.12 cli_wrapper.py --jira-instance rtk --jira-issues MAC8QQC-3660 -o result.zip
        '''
    )
    
    parser.add_argument(
        '-i', '--input',
        required=False,  # 改為非必需
        help='輸入檔案或資料夾，多個項目用逗號分隔'
    )
    
    parser.add_argument(
        '-o', '--output',
        required=True,
        help='輸出 ZIP 檔案名稱'
    )
    
    parser.add_argument(
        '--auto-group',
        action='store_true',
        help='自動將獨立的 ANR/Tombstone 檔案分組'
    )
    
    parser.add_argument(
        '--upload-issue',
        help='指定要上傳結果的 JIRA issue（如果與下載的 issue 不同）'
    )

    parser.add_argument(
        '--keep-temp',
        action='store_true',
        help='保留臨時檔案（用於除錯）'
    )

    parser.add_argument(
        '--no-auto-reopen',
        action='store_true',
        help='不要自動重新開啟已關閉的 JIRA issues'
    )

    # 新增 JIRA 相關參數
    jira_group = parser.add_argument_group('JIRA 選項')
    
    jira_group.add_argument(
        '--jira-instance',
        help='JIRA 實例名稱（在 jira_config.json 中配置）'
    )
    
    jira_group.add_argument(
        '--jira-issues',
        help='要下載附件的 JIRA issue keys，用逗號分隔（如 ANR-123,ANR-456）'
    )
    
    jira_group.add_argument(
        '--jira-file-patterns',
        help='要下載的檔案模式，用逗號分隔（如 *.zip,*.txt）。不指定則下載全部'
    )
    
    jira_group.add_argument(
        '--upload-to-jira',
        action='store_true',
        help='將分析結果上傳回 JIRA'
    )
    
    jira_group.add_argument(
        '--jira-config-file',
        default='jira_config.json',
        help='JIRA 配置檔案路徑（預設：jira_config.json）'
    )
    
    return parser.parse_args()

def prepare_analysis_directory(input_items, temp_dir, auto_group=True):
    """準備分析目錄結構"""
    group_counter = 1
    anr_files = []
    tombstone_files = []
    
    for item in input_items:
        item = item.strip()
        
        if not os.path.exists(item):
            print(f"警告：找不到 {item}，跳過")
            continue
        
        if os.path.isfile(item):
            # 處理單一檔案
            filename = os.path.basename(item)
            file_lower = filename.lower()
            
            # 檢查是否為 ZIP 檔案
            if item.endswith('.zip'):
                # 解壓縮 ZIP 檔案
                try:
                    with zipfile.ZipFile(item, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    print(f"已解壓縮: {item}")
                    continue
                except Exception as e:
                    print(f"解壓縮 {item} 失敗: {e}")
            
            # 檢查是否為 7z 檔案
            elif item.endswith('.7z'):
                # 使用系統命令解壓縮 7z 檔案
                try:
                    # 建立解壓縮目錄（以檔名命名，去掉 .7z）
                    extract_dir = os.path.join(temp_dir, os.path.splitext(filename)[0])
                    os.makedirs(extract_dir, exist_ok=True)
                    
                    # 使用 7z 命令解壓縮
                    cmd = ['7z', 'x', item, f'-o{extract_dir}', '-y']
                    result = subprocess.run(cmd, capture_output=True, text=True)
                    
                    if result.returncode == 0:
                        print(f"已解壓縮 7z: {item}")
                    else:
                        print(f"解壓縮 7z 失敗: {result.stderr}")
                        # 嘗試解壓到 temp_dir 根目錄
                        cmd = ['7z', 'x', item, f'-o{temp_dir}', '-y']
                        result = subprocess.run(cmd, capture_output=True, text=True)
                        if result.returncode == 0:
                            print(f"已解壓縮 7z 到根目錄: {item}")
                        else:
                            print(f"第二次嘗試解壓縮 7z 失敗: {result.stderr}")
                    continue
                except Exception as e:
                    print(f"執行 7z 命令失敗: {e}")
            
            # 檢查是否為 ANR 或 Tombstone 檔案
            elif 'anr' in file_lower or 'tombstone' in file_lower:
                dest_path = os.path.join(temp_dir, filename)
                shutil.copy2(item, dest_path)
                
                if 'anr' in file_lower:
                    anr_files.append((dest_path, filename))
                else:
                    tombstone_files.append((dest_path, filename))
            else:
                # 其他檔案直接複製
                shutil.copy2(item, os.path.join(temp_dir, filename))
        
        elif os.path.isdir(item):
            # 複製整個資料夾
            dest_folder = os.path.join(temp_dir, os.path.basename(item))
            shutil.copytree(item, dest_folder)
            print(f"已複製資料夾: {item}")
    
    # 自動分組處理
    if auto_group:

        # 重新掃描 temp_dir 中的所有 ANR 和 Tombstone 檔案
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_lower = file.lower()
                if 'anr' in file_lower or 'tombstone' in file_lower:
                    file_path = os.path.join(root, file)
                    # 檢查是否已經在合適的目錄結構中
                    rel_path = os.path.relpath(file_path, temp_dir)
                    path_parts = rel_path.split(os.sep)
                    
                    # 如果不在 GroupX/anr 或 GroupX/tombstones 結構中
                    if not (len(path_parts) >= 2 and 
                            path_parts[0].startswith('Group') and 
                            path_parts[1] in ['anr', 'tombstones']):
                        # 加入待分組列表
                        if 'anr' in file_lower:
                            anr_files.append((file_path, file))
                        else:
                            tombstone_files.append((file_path, file))
                                    
        # 處理 ANR 檔案
        if anr_files:
            print(f"找到 {len(anr_files)} 個獨立的 ANR 檔案，自動分組中...")
            for file_path, filename in anr_files:
                group_folder = os.path.join(temp_dir, f'Group{group_counter}', 'anr')
                os.makedirs(group_folder, exist_ok=True)
                shutil.move(file_path, os.path.join(group_folder, filename))
                print(f"  {filename} -> Group{group_counter}/anr/")
                group_counter += 1
        
        # 處理 Tombstone 檔案
        if tombstone_files:
            print(f"找到 {len(tombstone_files)} 個獨立的 Tombstone 檔案，自動分組中...")
            for file_path, filename in tombstone_files:
                group_folder = os.path.join(temp_dir, f'Group{group_counter}', 'tombstones')
                os.makedirs(group_folder, exist_ok=True)
                shutil.move(file_path, os.path.join(group_folder, filename))
                print(f"  {filename} -> Group{group_counter}/tombstones/")
                group_counter += 1

def run_vp_analyze(analysis_path, output_path):
    """執行 vp_analyze_logs.py"""
    try:
        vp_script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'routes', 'vp_analyze_logs.py')
        
        if not os.path.exists(vp_script_path):
            print("警告：找不到 vp_analyze_logs.py，跳過詳細分析")
            return False
        
        print("執行詳細分析...")
        cmd = ['python3.12', vp_script_path, analysis_path, output_path]
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300
        )
        
        if result.returncode == 0:
            print("詳細分析完成")
            return True
        else:
            print(f"詳細分析失敗: {result.stderr}")
            return False
            
    except Exception as e:
        print(f"執行 vp_analyze 時發生錯誤: {e}")
        return False

def generate_excel_file(output_path, results, analysis_output_path, base_path):
    """生成 Excel 檔案"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 準備 Excel 資料
        excel_data = []
        sn = 1
        current_time = datetime.now().strftime('%Y%m%d %H:%M:%S')
        
        for log in results.get('logs', []):
            # 讀取對應的 AI 分析結果
            ai_result = ""
            if log.get('file') and analysis_output_path and os.path.exists(analysis_output_path):
                try:
                    file_path = log['file']
                    if file_path.startswith(base_path):
                        relative_path = os.path.relpath(file_path, base_path)
                    else:
                        relative_path = file_path
                    
                    analyzed_file = os.path.join(analysis_output_path, relative_path + '.analyzed.txt')
                    
                    if os.path.exists(analyzed_file):
                        with open(analyzed_file, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            ai_result = extract_ai_summary(content)
                    else:
                        ai_result = "找不到分析結果"
                except Exception as e:
                    ai_result = f"讀取錯誤: {str(e)}"
            
            excel_data.append({
                'SN': sn,
                'Date': current_time,
                'Problem set': log.get('problem_set', '-'),
                'Type': log.get('type', ''),
                'Process': log.get('process', ''),
                'AI result': ai_result,
                'Filename': log.get('filename', ''),
                'Folder Path': log.get('file', '')
            })
            sn += 1
        
        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "ANR Tombstone Analysis"
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題
        headers = ['SN', 'Date', 'Problem set', 'Type', 'Process', 'AI result', 'Filename', 'Folder Path']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # 設定資料樣式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        anr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tombstone_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 寫入資料
        for row_idx, row_data in enumerate(excel_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.font = data_font
                cell.border = data_border
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment
                
                if col_idx == 4:  # Type 欄位
                    if row_data.get('Type') == 'ANR':
                        cell.fill = anr_fill
                    elif row_data.get('Type') == 'Tombstone':
                        cell.fill = tombstone_fill
        
        # 調整欄寬
        column_widths = {
            'A': 8,   # SN
            'B': 20,  # Date
            'C': 20,  # 問題 set
            'D': 12,  # Type
            'E': 30,  # Process
            'F': 60,  # AI result
            'G': 40,  # Filename
            'H': 80   # Folder Path
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 凍結標題列
        ws.freeze_panes = 'A2'
        
        # 儲存檔案
        excel_path = os.path.join(output_path, 'all_anr_tombstone_result.xlsx')
        wb.save(excel_path)
        print(f"已生成 Excel 檔案: all_anr_tombstone_result.xlsx")
        
        return excel_data
        
    except Exception as e:
        print(f"生成 Excel 檔案失敗: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def generate_html_report(output_path, results, analysis_output_path, base_path):
    """生成 HTML 統計報告"""
    try:
        # 準備資料，模擬網頁版的資料結構
        data = {
            'analysis_id': f'cli_{datetime.now().strftime("%Y%m%d%H%M%S")}',
            'total_files': results['total_files'],
            'files_with_cmdline': results['files_with_cmdline'],
            'anr_folders': results['anr_folders'],
            'tombstone_folders': results['tombstone_folders'],
            'statistics': results['statistics'],
            'file_statistics': results['file_statistics'],
            'logs': results['logs'],
            'analysis_time': results['analysis_time'],
            'used_grep': results['used_grep'],
            'zip_files_extracted': results.get('zip_files_extracted', 0),
            'anr_subject_count': results.get('anr_subject_count', 0),
            'vp_analyze_output_path': analysis_output_path,
            'vp_analyze_success': os.path.exists(analysis_output_path) if analysis_output_path else False
        }
        
        # 生成程序統計（不分類型）
        process_only_data = {}
        for item in data['statistics']['type_process_summary']:
            if not process_only_data.get(item['process']):
                process_only_data[item['process']] = {
                    'count': 0,
                    'problem_sets': set()
                }
            process_only_data[item['process']]['count'] += item['count']
            
            if item.get('problem_sets'):
                for ps in item['problem_sets']:
                    process_only_data[item['process']]['problem_sets'].add(ps)
        
        # 使用 HTML 模板
        html_report = HTML_TEMPLATE
        
        # 注入資料的腳本
        script_injection = f'''
<script>
    // 靜態頁面標記
    window.isStaticExport = true;
    window.exportBaseUrl = "";
    window.basePath = "{base_path}";
    
    // 保存分析報告相關資訊
    window.vpAnalyzeOutputPath = "{analysis_output_path}";
    window.vpAnalyzeSuccess = {str(data['vp_analyze_success']).lower()};
    
    // Injected analysis data
    window.injectedData = {json.dumps(data)};
    
    // Auto-load the data when page loads
    window.addEventListener('DOMContentLoaded', function() {{
        // 隱藏控制面板並顯示結果
        document.querySelector('.control-panel').style.display = 'none';
        document.getElementById('results').style.display = 'block';
        
        // 隱藏不需要的按鈕
        document.getElementById('exportHtmlBtn').style.display = 'none';
        ['exportExcelBtn', 'exportExcelReportBtn', 'mergeExcelBtn', 'downloadCurrentZipBtn'].forEach(id => {{
            const btn = document.getElementById(id);
            if (btn) btn.style.display = 'none';
        }});
        
        // 載入資料
        currentAnalysisId = window.injectedData.analysis_id;
        allLogs = window.injectedData.logs;
        allSummary = window.injectedData.statistics.type_process_summary || [];
        allFileStats = window.injectedData.file_statistics || [];
        
        // 生成程序統計資料
        const processOnlyData = {json.dumps({k: {'count': v['count'], 'problem_sets': list(v['problem_sets'])} for k, v in process_only_data.items()})};
        
        allProcessSummary = Object.entries(processOnlyData)
            .map(([process, data]) => ({{ 
                process, 
                count: data.count,
                problem_sets: data.problem_sets || []
            }}))
            .sort((a, b) => b.count - a.count);
        
        // Reset filters and pagination
        resetFiltersAndPagination();
        
        // Update UI
        updateResults(window.injectedData);
        
        // Show analysis info message
        let message = `分析完成！共掃描 ${{window.injectedData.total_files}} 個檔案`;
        message += `<br><br>報告生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`;
        
        const infoDiv = document.createElement('div');
        infoDiv.className = 'success';
        infoDiv.innerHTML = message;
        document.querySelector('.header').appendChild(infoDiv);
    }});
</script>
'''
        
        # 插入腳本
        html_report = html_report.replace('</body>', script_injection + '</body>')
        
        # 儲存 HTML
        html_path = os.path.join(output_path, 'all_anr_tombstone_result.html')
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_report)
        
        print(f"已生成 HTML 報告: all_anr_tombstone_result.html")
        
    except Exception as e:
        print(f"生成 HTML 報告失敗: {str(e)}")
        import traceback
        traceback.print_exc()

def generate_excel_report_html(output_path, excel_data, base_path):
    """生成 Excel 報表 HTML"""
    try:
        from routes.excel_report import EXCEL_REPORT_TEMPLATE
        
        # 準備模板資料
        template_data = {
            'filename': 'CLI Analysis Result',
            'filepath': base_path,
            'load_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data': json.dumps(excel_data),
            'excel_data_base64': '',
            'filename_list': ['CLI Analysis Result'],
            'path_list': [base_path]
        }
        
        # 生成 HTML 內容
        html_content = EXCEL_REPORT_TEMPLATE
        
        # 替換模板變數
        for key, value in template_data.items():
            if key in ['data', 'filename_list', 'path_list']:
                if key == 'data':
                    html_content = html_content.replace('{{ data | tojson }}', value)
                elif key == 'filename_list':
                    html_content = html_content.replace('{{ filename_list | tojson }}', json.dumps(value))
                elif key == 'path_list':
                    html_content = html_content.replace('{{ path_list | tojson }}', json.dumps(value))
            else:
                html_content = html_content.replace(f'{{{{ {key} }}}}', str(value))
        
        # 儲存檔案
        report_path = os.path.join(output_path, 'all_anr_tombstone_excel_result.html')
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"已生成 Excel 報表: all_anr_tombstone_excel_result.html")
        
    except Exception as e:
        print(f"生成 Excel 報表失敗: {str(e)}")
        import traceback
        traceback.print_exc()

def create_output_zip(analysis_output_path, output_file):
    """建立輸出 ZIP 檔案"""
    print(f"\n建立輸出檔案: {output_file}")
    
    with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(analysis_output_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, os.path.dirname(analysis_output_path))
                zipf.write(file_path, arcname)
                
    print(f"已建立 ZIP 檔案: {output_file}")
    print(f"檔案大小: {os.path.getsize(output_file) / 1024 / 1024:.2f} MB")

def main():
    """主程式"""
    args = parse_arguments()
    
    # 檢查是否至少有一個輸入來源
    if not args.input and not (args.jira_instance and args.jira_issues):
        print("錯誤：必須提供至少一個輸入來源")
        print("  使用 -i/--input 指定本地檔案/資料夾")
        print("  或使用 --jira-instance 和 --jira-issues 從 JIRA 下載")
        return 1
            
    # 解析輸入項目
    input_items = [item.strip() for item in args.input.split(',') if item.strip()] if args.input else []
    
    # 處理 JIRA 下載
    jira_downloads = []
    jira_client = None
    jira_file_manager = None
    
    if args.jira_instance and args.jira_issues:
        print("\n處理 JIRA 下載...")
        
        # 載入 JIRA 配置
        jira_config = JiraConfig(args.jira_config_file)
        config = jira_config.get_jira_config(args.jira_instance)
        
        if not config:
            print(f"錯誤：找不到 JIRA 實例 '{args.jira_instance}' 的配置")
            print(f"可用的實例：{', '.join(jira_config.list_instances())}")
            return 1
        
        # 建立 JIRA 客戶端
        jira_client = JiraClient(
            config['url'], 
            config['token'], 
            config.get('username')
        )
        jira_file_manager = JiraFileManager(jira_client)
        
        # 解析 issue keys 和檔案模式
        issue_keys = [key.strip() for key in args.jira_issues.split(',')]
        file_patterns = None
        if args.jira_file_patterns:
            file_patterns = [p.strip() for p in args.jira_file_patterns.split(',')]
        
        # 下載檔案
        temp_jira_dir = tempfile.mkdtemp(prefix='jira_downloads_')
        for issue_key in issue_keys:
            downloaded = jira_file_manager.download_issue_attachments(
                issue_key, file_patterns, 
                os.path.join(temp_jira_dir, issue_key)
            )
            jira_downloads.extend(downloaded)
        
        # 將下載的檔案加入輸入項目
        if jira_downloads:
            print(f"\n從 JIRA 下載了 {len(jira_downloads)} 個檔案")
            input_items.extend(jira_downloads)

    print("="*60)
    print("Android ANR/Tombstone Analyzer CLI")
    print("="*60)
    print(f"輸入項目: {len(input_items)} 個")
    print(f"輸出檔案: {args.output}")
    print(f"自動分組: {'是' if args.auto_group else '否'}")
    print("="*60)
    
    # 建立臨時目錄
    temp_dir = tempfile.mkdtemp(prefix='anr_cli_')
    print(f"\n臨時目錄: {temp_dir}")
    
    try:
        # 準備分析目錄
        print("\n準備分析資料...")
        prepare_analysis_directory(input_items, temp_dir, args.auto_group)
        
        # 執行分析
        print("\n開始分析...")
        analyzer = AndroidLogAnalyzer()
        results = analyzer.analyze_logs(temp_dir)
        
        # 為每個 log 添加 problem_set
        for log in results['logs']:
            if 'problem_set' not in log and log.get('file'):
                try:
                    file_path = log['file']
                    relative_path = file_path.replace(temp_dir, '').lstrip('/')
                    parts = relative_path.split('/')
                    if len(parts) > 1:
                        log['problem_set'] = parts[0]
                    else:
                        log['problem_set'] = '未分類'
                except:
                    log['problem_set'] = '未分類'
        
        # 為每個 file_stat 添加 problem_set
        for file_stat in results['file_statistics']:
            if 'problem_set' not in file_stat and file_stat.get('filepath'):
                try:
                    file_path = file_stat['filepath']
                    relative_path = file_path.replace(temp_dir, '').lstrip('/')
                    parts = relative_path.split('/')
                    if len(parts) > 1:
                        file_stat['problem_set'] = parts[0]
                    else:
                        file_stat['problem_set'] = '未分類'
                except:
                    file_stat['problem_set'] = '未分類'
        
        print(f"\n分析完成:")
        print(f"  總掃描檔案數: {results['total_files']}")
        print(f"  ANR 檔案數: {results.get('anr_subject_count', 0)}")
        print(f"  Tombstone 檔案數: {results['files_with_cmdline'] - results.get('anr_subject_count', 0)}")
        print(f"  分析耗時: {results['analysis_time']:.2f} 秒")
        
        # 建立輸出目錄
        output_dir_name = f".cli_anr_tombstones_analyze"
        output_path = os.path.join(temp_dir, output_dir_name)
        os.makedirs(output_path, exist_ok=True)
        
        # 執行 vp_analyze
        vp_analyze_success = run_vp_analyze(temp_dir, output_path)
        
        # 生成所有報告檔案
        print("\n生成報告檔案...")
        
        # 1. 生成 Excel 檔案
        excel_data = generate_excel_file(output_path, results, output_path, temp_dir)
        
        # 2. 生成 HTML 統計報告
        generate_html_report(output_path, results, output_path, temp_dir)
        
        # 3. 生成 Excel 報表 HTML
        if excel_data:
            generate_excel_report_html(output_path, excel_data, temp_dir)
        
        # 儲存原始分析結果
        with open(os.path.join(output_path, 'analysis_results.json'), 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        # 建立輸出 ZIP
        output_path_abs = os.path.abspath(args.output)
        create_output_zip(output_path, output_path_abs)
        
        # 上傳結果到 JIRA（如果需要）
        if args.upload_to_jira and jira_client and jira_file_manager:
            print("\n上傳分析結果到 JIRA...")
            
            # 決定上傳到哪些 issues
            if args.upload_issue:
                # 如果指定了特定的上傳 issue
                upload_issues = [args.upload_issue.strip()]
                print(f"將上傳到指定的 issue: {args.upload_issue}")
            else:
                # 否則上傳到下載來源的 issues
                if args.jira_issues:
                    upload_issues = [key.strip() for key in args.jira_issues.split(',')]
                    print(f"將上傳到來源 issues: {', '.join(upload_issues)}")
                else:
                    upload_issues = []
                    print("警告：沒有指定要上傳的 issue")
            
            # 執行上傳
            upload_failed = []
            upload_success = []
            
            for issue_key in upload_issues:
                try:
                    success = jira_file_manager.upload_analysis_result(
                        issue_key, 
                        output_path_abs, 
                        add_comment=True,
                        auto_reopen=not args.no_auto_reopen
                    )
                    if success:
                        upload_success.append(issue_key)
                    else:
                        upload_failed.append(issue_key)
                except Exception as e:
                    print(f"錯誤：處理 {issue_key} 時發生異常: {e}")
                    upload_failed.append(issue_key)
                    
                    # 詳細的錯誤處理
                    if "403" in str(e) or "permission" in str(e).lower():
                        print(f"  ⚠️  權限問題：您沒有權限上傳附件到 {issue_key}")
                        print(f"     請聯繫 JIRA 管理員或專案負責人授予附件上傳權限")
                    elif "401" in str(e):
                        print(f"  ⚠️  認證問題：請檢查 JIRA token 是否正確")
                    elif "404" in str(e):
                        print(f"  ⚠️  找不到 issue: {issue_key}")
                    elif "415" in str(e):
                        print(f"  ⚠️  不支援的檔案類型或 issue 狀態問題")
            
            # 顯示上傳結果摘要
            print("\n上傳結果摘要:")
            if upload_success:
                print(f"✓ 成功上傳到: {', '.join(upload_success)}")
            if upload_failed:
                print(f"✗ 上傳失敗: {', '.join(upload_failed)}")
                if args.upload_issue and args.upload_issue in upload_failed:
                    print(f"\n提示：指定的上傳 issue {args.upload_issue} 上傳失敗")
                    print("     可能原因：")
                    print("     1. Issue 已關閉且您沒有 reopen 權限")
                    print("     2. 您沒有該 issue 的附件上傳權限")
                    print("     3. Issue 不存在或 key 錯誤")
            
            if not upload_success and not upload_failed:
                print("沒有執行任何上傳操作")
            
            print(f"\n分析結果已儲存在本地：{output_path_abs}")
        
        print("\n分析完成！")
        print(f"已生成以下檔案：")
        print(f"  - all_anr_tombstone_result.xlsx")
        print(f"  - all_anr_tombstone_result.html")
        print(f"  - all_anr_tombstone_excel_result.html")
        print(f"  - analysis_results.json")
        print(f"  - vp_analyze 分析結果（如果成功）")
        
    except Exception as e:
        print(f"\n錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
        
    finally:
        # 清理臨時檔案
        if not args.keep_temp:
            print(f"\n清理臨時檔案: {temp_dir}")
            shutil.rmtree(temp_dir, ignore_errors=True)
        else:
            print(f"\n保留臨時檔案: {temp_dir}")
    
    return 0

if __name__ == '__main__':
    sys.exit(main())